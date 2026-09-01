import os
import random
import os
import io
import pandas as pd
import uuid
from datetime import datetime
from flask import jsonify
from urllib.parse import quote
from io import BytesIO
import requests
from datetime import datetime
from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak
)

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import mm
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash
)

import io
import os

from datetime import datetime

from flask import (
    send_file,
    redirect,
    url_for
)

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from openpyxl.drawing.image import Image as XLImage
from flask import send_from_directory
import os
from datetime import datetime

current_time = datetime.now()

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_from_directory
)
from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)
from dotenv import load_dotenv

from werkzeug.utils import secure_filename

from database import get_db, init_db
from flask import Flask, request, session, redirect, url_for, jsonify, render_template, flash, send_file
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

NIGERIA_TZ = ZoneInfo("Africa/Lagos")

# ============================================================
# LOAD ENVIRONMENT VARIABLES
# ============================================================

load_dotenv()

import os
import cloudinary
import cloudinary.uploader

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    secure=True
)
# ============================================================
# APPLICATION
# ============================================================

app = Flask(__name__)


app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY",
    "development-secret-key"
)


# ============================================================
# UPLOAD CONFIGURATION
# ============================================================

UPLOAD_FOLDER = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "static",
    "uploads"
)


os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)


app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER


ALLOWED_IMAGE_EXTENSIONS = {
    "jpg",
    "jpeg",
    "png",
    "webp"
}


ALLOWED_DOCUMENT_EXTENSIONS = {
    "pdf",
    "doc",
    "docx"
}


# ============================================================
# COMPANY INFORMATION
# ============================================================

COMPANY_NAME = "AV KING VET DRUG VENTURE"

COMPANY_MOTTO = "Your Needs, Our Priority"

COMPANY_EMAIL = "avkingvetdrug@gmail.com"

COMPANY_PHONE = "09058842501"

COMPANY_ADDRESS = (
    "NO11, HALLELUJAH SHOPPING COMPLEX, "
    "OPPOSITE POULTRY ASSOCIATION, "
    "EGBEDA, IYANA-AJIA, IBADAN"
)


# ============================================================
# AVAILABLE POSITIONS
# ============================================================

POSITIONS = [

    "Sales Representative",

    "Veterinary Drug Sales Assistant",

    "Store Assistant",

    "Account / Cashier",

    "Administrative Assistant",

    "Field Sales Representative",

    "Other"

]


# ============================================================
# GLOBAL TEMPLATE VARIABLES
# ============================================================

@app.context_processor
def inject_company_details():

    return {

        "company_name": COMPANY_NAME,

        "company_motto": COMPANY_MOTTO,

        "company_email": COMPANY_EMAIL,

        "company_phone": COMPANY_PHONE,

        "company_address": COMPANY_ADDRESS,

        "positions": POSITIONS

    }


# ============================================================
# FILE VALIDATION
# ============================================================

def allowed_file(filename, allowed_extensions):

    if not filename:
        return False

    if "." not in filename:
        return False

    extension = (
        filename.rsplit(".", 1)[1]
        .lower()
    )

    return extension in allowed_extensions


# ============================================================
# APPLICATION NUMBER
# ============================================================

# =========================================================
# GENERATE RANDOM APPLICATION NUMBER
# =========================================================

def generate_application_number(conn):

    while True:

        random_number = random.randint(
            1000,
            9999
        )

        application_number = (
            f"AV-APP-{random_number}"
        )

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT id
                FROM applications
                WHERE application_number = %s
                LIMIT 1
                """,
                (application_number,)
            )

            existing = cur.fetchone()

        if not existing:

            return application_number


# ============================================================
# HOME
# ============================================================
@app.route("/")
def home():
    conn = get_db()

    try:
        cursor = conn.cursor()

        # Get company settings
        cursor.execute("""
            SELECT
                company_name,
                company_email,
                company_phone,
                company_address,
                company_website,
                footer_text,
                logo
            FROM company_settings
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if settings:
            company_name = settings["company_name"] or "AV KING"
            company_logo = settings["logo"]
            footer_text = settings["footer_text"] or ""
        else:
            company_name = "AV KING"
            company_logo = None
            footer_text = ""

        return render_template(
            "careers.html",
            company_name=company_name,
            company_logo=company_logo,
            footer_text=footer_text,
            current_year=datetime.now().year
        )

    except Exception as e:
        app.logger.exception(
            "Error loading public home page"
        )

        # Fallback so the public homepage does not crash
        return render_template(
            "careers.html",
            company_name="AV KING",
            company_logo=None,
            footer_text="",
            current_year=datetime.now().year
        )

    finally:
        conn.close()

@app.context_processor
def utility_processor():

    def endpoint_exists(endpoint):

        return endpoint in app.view_functions

    return {
        "endpoint_exists": endpoint_exists
    }
    
@app.route("/setup-admin")
def setup_admin():

    username = "admin"
    password = "AVKing@2026"
    full_name = "AV KING Administrator"

    password_hash = generate_password_hash(password)

    conn = get_db()

    try:
        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT id
                FROM admin_users
                WHERE username = %s
                LIMIT 1
                """,
                (username,)
            )

            existing = cur.fetchone()

            if existing:
                return "Admin account already exists."

            cur.execute(
                """
                INSERT INTO admin_users (
                    username,
                    password_hash,
                    full_name
                )
                VALUES (%s, %s, %s)
                """,
                (
                    username,
                    password_hash,
                    full_name
                )
            )

        conn.commit()

        return """
        <h2>Admin account created successfully.</h2>
        <p>Username: admin</p>
        <p>Password: AVKing@2026</p>
        <p>You can now go to /admin/login</p>
        """

    except Exception as e:

        conn.rollback()

        return f"Error: {e}", 500

    finally:
        conn.close()

@app.route("/admin/interviews")
def admin_interviews():

    if not admin_required():
        return redirect(url_for("admin_login"))

    conn = get_db()

    try:
        with conn.cursor() as cur:

            cur.execute("""
                SELECT *
                FROM applications
                WHERE interview_date IS NOT NULL
                ORDER BY interview_date ASC
            """)

            interviews = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "admin_interviews.html",
        interviews=interviews
    )

# ============================================================
# APPLICATION FORM
# ============================================================
@app.route(
    "/apply",
    methods=["GET", "POST"]
)
def apply():

    # ========================================================
    # CHECK APPLICATION SETTINGS
    # ========================================================

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    application_status,
                    application_deadline
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            settings = cur.fetchone()

    except Exception:

        app.logger.exception(
            "Error checking application settings"
        )

        settings = None

    finally:

        conn.close()


    # ========================================================
    # DEFAULT SETTINGS
    # ========================================================

    application_status = "Closed"
    application_deadline = None

    if settings:

        application_status = (
            settings["application_status"]
            or "Closed"
        ).strip().capitalize()

        application_deadline = (
            settings["application_deadline"]
        )


    # ========================================================
    # APPLICATIONS MANUALLY CLOSED
    # ========================================================

    if application_status == "Closed":

        return render_template(
            "applications_closed.html"
        ), 403


    # ========================================================
    # CHECK APPLICATION DEADLINE
    # ========================================================

    if application_deadline:

        try:

            if hasattr(
                application_deadline,
                "date"
            ):

                deadline_date = (
                    application_deadline.date()
                )

            elif isinstance(
                application_deadline,
                str
            ):

                deadline_date = datetime.strptime(
                    application_deadline,
                    "%Y-%m-%d"
                ).date()

            else:

                deadline_date = application_deadline


            if datetime.now().date() > deadline_date:

                return render_template(
                    "applications_closed.html"
                ), 403


        except Exception:

            app.logger.exception(
                "Error checking application deadline"
            )

            return render_template(
                "applications_closed.html"
            ), 403


    # ========================================================
    # DISPLAY APPLICATION FORM
    # ========================================================

    if request.method == "GET":

        conn = get_db()

        try:

            with conn.cursor() as cur:

                cur.execute(
                    """
                    SELECT
                        id,
                        position_name
                    FROM available_positions
                    WHERE is_active = TRUE
                    ORDER BY position_name ASC
                    """
                )

                positions = cur.fetchall()

        except Exception:

            app.logger.exception(
                "Error loading available positions"
            )

            positions = []

        finally:

            conn.close()


        return render_template(
            "apply.html",
            positions=positions
        )


    # ========================================================
    # POST APPLICATION
    # ========================================================

    first_name = request.form.get(
        "first_name",
        ""
    ).strip()

    middle_name = request.form.get(
        "middle_name",
        ""
    ).strip()

    last_name = request.form.get(
        "last_name",
        ""
    ).strip()

    gender = request.form.get(
        "gender",
        ""
    ).strip()

    date_of_birth = request.form.get(
        "date_of_birth",
        ""
    ).strip()


    # ========================================================
    # CONTACT INFORMATION
    # ========================================================

    phone = request.form.get(
        "phone",
        ""
    ).strip()

    email = request.form.get(
        "email",
        ""
    ).strip().lower()

    address = request.form.get(
        "address",
        ""
    ).strip()

    state = request.form.get(
        "state",
        ""
    ).strip()

    lga = request.form.get(
        "lga",
        ""
    ).strip()


    # ========================================================
    # APPLICATION INFORMATION
    # ========================================================

    position_applied = request.form.get(
        "position_applied",
        ""
    ).strip()


    # ========================================================
    # EDUCATION
    # ========================================================

    highest_qualification = request.form.get(
        "highest_qualification",
        ""
    ).strip()

    course_of_study = request.form.get(
        "course_of_study",
        ""
    ).strip()

    institution = request.form.get(
        "institution",
        ""
    ).strip()

    graduation_year = request.form.get(
        "graduation_year",
        ""
    ).strip()


    # ========================================================
    # WORK EXPERIENCE
    # ========================================================

    work_experience = request.form.get(
        "work_experience",
        ""
    ).strip()

    previous_employer = request.form.get(
        "previous_employer",
        ""
    ).strip()

    previous_position = request.form.get(
        "previous_position",
        ""
    ).strip()


    # ========================================================
    # ADDITIONAL INFORMATION
    # ========================================================

    reason_for_applying = request.form.get(
        "reason_for_applying",
        ""
    ).strip()

    additional_information = request.form.get(
        "additional_information",
        ""
    ).strip()


    # ========================================================
    # APPLICANT PORTAL PASSWORD
    # ========================================================

    portal_password = request.form.get(
        "portal_password",
        ""
    ).strip()

    confirm_portal_password = request.form.get(
        "confirm_portal_password",
        ""
    ).strip()


    # ========================================================
    # DECLARATION
    # ========================================================

    declaration = request.form.get(
        "declaration"
    )


    # ========================================================
    # FILES
    # ========================================================

    passport = request.files.get(
        "passport"
    )

    cv = request.files.get(
        "cv"
    )

    qualification = request.files.get(
        "qualification"
    )


    # ========================================================
    # BASIC VALIDATION
    # ========================================================

    if not first_name:

        flash(
            "Please enter your first name.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    if not last_name:

        flash(
            "Please enter your last name.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    if not phone:

        flash(
            "Please enter your phone number.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    if not position_applied:

        flash(
            "Please select the position you are applying for.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    # ========================================================
    # PASSWORD VALIDATION
    # ========================================================

    if not portal_password:

        flash(
            "Please create a password for your applicant portal.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    if len(portal_password) < 6:

        flash(
            "Your portal password must contain at least 6 characters.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    if portal_password != confirm_portal_password:

        flash(
            "The portal passwords do not match.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    # ========================================================
    # DECLARATION VALIDATION
    # ========================================================

    if not declaration:

        flash(
            "You must accept the declaration before submitting.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    # ========================================================
    # PASSPORT VALIDATION
    # ========================================================

    if not passport or not passport.filename:

        flash(
            "Please upload your passport photograph.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    if not allowed_file(
        passport.filename,
        ALLOWED_IMAGE_EXTENSIONS
    ):

        flash(
            "Passport must be JPG, JPEG, PNG or WEBP.",
            "error"
        )

        return render_template(
            "apply.html"
        )


    # ========================================================
    # CV VALIDATION
    # ========================================================

    if cv and cv.filename:

        if not allowed_file(
            cv.filename,
            ALLOWED_DOCUMENT_EXTENSIONS
        ):

            flash(
                "CV must be PDF, DOC or DOCX.",
                "error"
            )

            return render_template(
                "apply.html"
            )


    # ========================================================
    # QUALIFICATION VALIDATION
    # ========================================================

    if qualification and qualification.filename:

        if not allowed_file(
            qualification.filename,
            ALLOWED_DOCUMENT_EXTENSIONS
        ):

            flash(
                "Qualification document must be PDF, DOC or DOCX.",
                "error"
            )

            return render_template(
                "apply.html"
            )


    # ========================================================
    # DATABASE CONNECTION
    # ========================================================

    conn = get_db()

    uploaded_cloudinary_public_ids = []

    try:

        with conn.cursor() as cur:

            # ==================================================
            # FINAL APPLICATION STATUS CHECK
            # ==================================================

            cur.execute(
                """
                SELECT
                    application_status,
                    application_deadline
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            current_settings = cur.fetchone()


            if not current_settings:

                conn.rollback()

                return render_template(
                    "applications_closed.html"
                ), 403


            # ==================================================
            # FINAL STATUS CHECK
            # ==================================================

            current_status = (
                current_settings["application_status"]
                or "Closed"
            ).strip().capitalize()


            if current_status == "Closed":

                conn.rollback()

                return render_template(
                    "applications_closed.html"
                ), 403


            # ==================================================
            # FINAL DEADLINE CHECK
            # ==================================================

            current_deadline = (
                current_settings["application_deadline"]
            )


            if current_deadline:

                if hasattr(
                    current_deadline,
                    "date"
                ):

                    current_deadline_date = (
                        current_deadline.date()
                    )

                elif isinstance(
                    current_deadline,
                    str
                ):

                    current_deadline_date = (
                        datetime.strptime(
                            current_deadline,
                            "%Y-%m-%d"
                        ).date()
                    )

                else:

                    current_deadline_date = (
                        current_deadline
                    )


                if datetime.now().date() > current_deadline_date:

                    conn.rollback()

                    return render_template(
                        "applications_closed.html"
                    ), 403


            # ==================================================
            # GENERATE APPLICATION NUMBER
            # ==================================================

            application_number = (
                generate_application_number(conn)
            )


            app.logger.info(
                "Generated application number: %s",
                application_number
            )


            # ==================================================
            # HASH PASSWORD
            # ==================================================

            password_hash = generate_password_hash(
                portal_password
            )


            # ==================================================
            # INITIALIZE CLOUDINARY URL VARIABLES
            # ==================================================

            passport_filename = None
            cv_filename = None
            qualification_filename = None


            # ==================================================
            # VERIFY CLOUDINARY CONFIGURATION
            # ==================================================

            cloudinary_cloud_name = os.getenv(
                "CLOUDINARY_CLOUD_NAME"
            )

            cloudinary_api_key = os.getenv(
                "CLOUDINARY_API_KEY"
            )

            cloudinary_api_secret = os.getenv(
                "CLOUDINARY_API_SECRET"
            )


            if not cloudinary_cloud_name:

                raise RuntimeError(
                    "CLOUDINARY_CLOUD_NAME is not configured."
                )


            if not cloudinary_api_key:

                raise RuntimeError(
                    "CLOUDINARY_API_KEY is not configured."
                )


            if not cloudinary_api_secret:

                raise RuntimeError(
                    "CLOUDINARY_API_SECRET is not configured."
                )


            # ==================================================
            # UPLOAD PASSPORT TO CLOUDINARY
            # ==================================================

            if passport and passport.filename:

                try:

                    passport_result = (
                        cloudinary.uploader.upload(
                            passport,
                            folder="av_king/applicants/passports",
                            public_id=(
                                f"{application_number}_passport"
                            ),
                            resource_type="image",
                            overwrite=False
                        )
                    )


                    passport_filename = (
                        passport_result.get(
                            "secure_url"
                        )
                    )


                    passport_public_id = (
                        passport_result.get(
                            "public_id"
                        )
                    )


                    if passport_public_id:

                        uploaded_cloudinary_public_ids.append(
                            (
                                passport_public_id,
                                "image"
                            )
                        )


                    if not passport_filename:

                        raise RuntimeError(
                            "Cloudinary did not return a passport URL."
                        )


                    app.logger.info(
                        "PASSPORT UPLOADED TO CLOUDINARY: %s",
                        passport_filename
                    )


                except Exception:

                    app.logger.exception(
                        "Error uploading applicant passport to Cloudinary"
                    )

                    raise RuntimeError(
                        "Unable to upload passport photograph. Please try again."
                    )


            # ==================================================
            # UPLOAD CV TO CLOUDINARY
            # ==================================================

            if cv and cv.filename:

                try:

                    cv_result = (
                        cloudinary.uploader.upload(
                            cv,
                            folder="av_king/applicants/cvs",
                            public_id=(
                                f"{application_number}_cv"
                            ),
                            resource_type="raw",
                            overwrite=False
                        )
                    )


                    cv_filename = (
                        cv_result.get(
                            "secure_url"
                        )
                    )


                    cv_public_id = (
                        cv_result.get(
                            "public_id"
                        )
                    )


                    if cv_public_id:

                        uploaded_cloudinary_public_ids.append(
                            (
                                cv_public_id,
                                "raw"
                            )
                        )


                    if not cv_filename:

                        raise RuntimeError(
                            "Cloudinary did not return a CV URL."
                        )


                    app.logger.info(
                        "CV UPLOADED TO CLOUDINARY: %s",
                        cv_filename
                    )


                except Exception:

                    app.logger.exception(
                        "Error uploading applicant CV to Cloudinary"
                    )

                    raise RuntimeError(
                        "Unable to upload CV. Please try again."
                    )


            # ==================================================
            # UPLOAD QUALIFICATION TO CLOUDINARY
            # ==================================================

            if qualification and qualification.filename:

                try:

                    qualification_result = (
                        cloudinary.uploader.upload(
                            qualification,
                            folder="av_king/applicants/qualifications",
                            public_id=(
                                f"{application_number}_qualification"
                            ),
                            resource_type="raw",
                            overwrite=False
                        )
                    )


                    qualification_filename = (
                        qualification_result.get(
                            "secure_url"
                        )
                    )


                    qualification_public_id = (
                        qualification_result.get(
                            "public_id"
                        )
                    )


                    if qualification_public_id:

                        uploaded_cloudinary_public_ids.append(
                            (
                                qualification_public_id,
                                "raw"
                            )
                        )


                    if not qualification_filename:

                        raise RuntimeError(
                            "Cloudinary did not return a qualification URL."
                        )


                    app.logger.info(
                        "QUALIFICATION UPLOADED TO CLOUDINARY: %s",
                        qualification_filename
                    )


                except Exception:

                    app.logger.exception(
                        "Error uploading qualification to Cloudinary"
                    )

                    raise RuntimeError(
                        "Unable to upload qualification document. Please try again."
                    )


            # ==================================================
            # INSERT APPLICATION
            # ==================================================

            cur.execute(
                """
                INSERT INTO applications (

                    application_number,

                    first_name,
                    middle_name,
                    last_name,

                    gender,
                    date_of_birth,

                    phone,
                    email,

                    address,
                    state,
                    lga,

                    position_applied,

                    highest_qualification,
                    course_of_study,
                    institution,
                    graduation_year,

                    work_experience,
                    previous_employer,
                    previous_position,

                    reason_for_applying,
                    additional_information,

                    passport_filename,
                    cv_filename,
                    qualification_filename,

                    password_hash,

                    portal_active,

                    status

                )

                VALUES (

                    %s,

                    %s,
                    %s,
                    %s,

                    %s,
                    %s,

                    %s,
                    %s,

                    %s,
                    %s,
                    %s,

                    %s,

                    %s,
                    %s,
                    %s,
                    %s,

                    %s,
                    %s,
                    %s,

                    %s,
                    %s,

                    %s,
                    %s,
                    %s,

                    %s,

                    TRUE,

                    'Pending'

                )

                RETURNING id
                """,

                (
                    application_number,

                    first_name,
                    middle_name,
                    last_name,

                    gender,
                    date_of_birth or None,

                    phone,
                    email or None,

                    address,
                    state,
                    lga,

                    position_applied,

                    highest_qualification,
                    course_of_study,
                    institution,
                    graduation_year,

                    work_experience,
                    previous_employer,
                    previous_position,

                    reason_for_applying,
                    additional_information,

                    passport_filename,
                    cv_filename,
                    qualification_filename,

                    password_hash
                )
            )


            application = cur.fetchone()


        # ====================================================
        # COMMIT DATABASE TRANSACTION
        # ====================================================

        conn.commit()


        app.logger.info(
            "APPLICATION SUBMITTED SUCCESSFULLY: %s",
            application_number
        )


    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error submitting application"
        )


        # ====================================================
        # CLEAN UP CLOUDINARY FILES IF DATABASE INSERT FAILS
        # ====================================================

        for public_id, resource_type in (
            uploaded_cloudinary_public_ids
        ):

            try:

                cloudinary.uploader.destroy(
                    public_id,
                    resource_type=resource_type
                )

                app.logger.info(
                    "Removed Cloudinary upload after failed application: %s",
                    public_id
                )

            except Exception:

                app.logger.exception(
                    "Unable to clean up Cloudinary file: %s",
                    public_id
                )


        flash(
            "We could not submit your application at this time. Please try again.",
            "error"
        )


        return render_template(
            "apply.html"
        ), 500


    finally:

        conn.close()


    # ========================================================
    # SUCCESS
    # ========================================================

    return redirect(
        url_for(
            "application_success",
            application_number=application_number
        )
    )


@app.route(
    "/admin/application_status",
    methods=["POST"]
)
def admin_application_status():

    if session.get("role") != "admin":
        return redirect("/")

    status = request.form.get(
        "status",
        ""
    ).strip().lower()

    if status not in ["open", "closed"]:

        flash(
            "Invalid application status.",
            "error"
        )

        return redirect(
            url_for("settings")
        )

    applications_open = (
        status == "open"
    )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                UPDATE company_settings
                SET applications_open = %s
                """,
                (applications_open,)
            )

        conn.commit()

        if applications_open:

            flash(
                "Applications are now OPEN.",
                "success"
            )

        else:

            flash(
                "Applications are now CLOSED.",
                "success"
            )

    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error updating application status"
        )

        flash(
            "Unable to update application status.",
            "error"
        )

    finally:

        conn.close()

    return redirect(
        url_for("settings")
    )

# ============================================================
# CREATE INITIAL ADMIN
# ============================================================

def create_initial_admin():

    username = os.getenv(
        "ADMIN_USERNAME"
    )

    password = os.getenv(
        "ADMIN_PASSWORD"
    )

    if not username or not password:
        return

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT id
                FROM admin_users
                WHERE username = %s
                """,
                (username,)
            )

            existing = cur.fetchone()

            if existing:
                return


            password_hash = generate_password_hash(
                password
            )


            cur.execute(
                """
                INSERT INTO admin_users (
                    username,
                    password_hash,
                    full_name
                )

                VALUES (
                    %s,
                    %s,
                    %s
                )
                """,
                (
                    username,
                    password_hash,
                    "System Administrator"
                )
            )

        conn.commit()

    finally:

        conn.close()

# =========================================================
# COMPANY SETTINGS HELPER
# =========================================================

def get_company_settings():
    """
    Get the current company settings.

    Returns:
        dict-like row or None
    """

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT *
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            settings = cur.fetchone()

            return settings

    except Exception:

        app.logger.exception(
            "Unable to load company settings."
        )

        return None

    finally:

        conn.close()
# ============================================================
# ADMIN LOGIN
# ============================================================
# =========================================================
# ADMIN LOGIN
@app.route(
    "/admin/login",
    methods=["GET", "POST"]
)
def admin_login():

    # =========================================================
    # LOAD COMPANY SETTINGS
    # =========================================================

    def get_company_settings():

        conn = get_db()

        try:

            with conn.cursor() as cur:

                cur.execute(
                    """
                    SELECT
                        id,
                        company_name,
                        logo,
                        admin_username,
                        admin_password_hash
                    FROM company_settings
                    ORDER BY id ASC
                    LIMIT 1
                    """
                )

                return cur.fetchone()

        except Exception:

            app.logger.exception(
                "Unable to load company settings for admin login."
            )

            return None

        finally:

            conn.close()


    # =========================================================
    # ALREADY LOGGED IN
    # =========================================================

    if session.get("admin_id"):

        # Make sure old admin sessions also have the role
        session["role"] = "admin"
        session["admin_logged_in"] = True

        return redirect(
            url_for("admin_dashboard")
        )


    # =========================================================
    # GET REQUEST
    # =========================================================

    if request.method == "GET":

        settings = get_company_settings()

        return render_template(
            "admin_login.html",
            settings=settings
        )


    # =========================================================
    # LOGIN DETAILS
    # =========================================================

    username = (
        request.form.get(
            "username",
            ""
        )
        .strip()
    )

    password = request.form.get(
        "password",
        ""
    )


    # =========================================================
    # VALIDATE LOGIN FIELDS
    # =========================================================

    if not username or not password:

        flash(
            "Please enter your username and password.",
            "error"
        )

        return render_template(
            "admin_login.html",
            settings=get_company_settings()
        )


    # =========================================================
    # GET ADMIN ACCOUNT
    # =========================================================

    settings = get_company_settings()

    if not settings:

        flash(
            "Administrator account has not been configured.",
            "error"
        )

        return render_template(
            "admin_login.html",
            settings=None
        )


    # =========================================================
    # STORED ADMIN DETAILS
    # =========================================================

    stored_username = (
        settings.get("admin_username")
        or ""
    ).strip()

    stored_password_hash = (
        settings.get("admin_password_hash")
    )


    # =========================================================
    # CHECK USERNAME
    # =========================================================

    if (
        not stored_username
        or username != stored_username
    ):

        flash(
            "Invalid username or password.",
            "error"
        )

        return render_template(
            "admin_login.html",
            settings=settings
        )


    # =========================================================
    # CHECK PASSWORD CONFIGURATION
    # =========================================================

    if not stored_password_hash:

        flash(
            "Administrator password has not been configured.",
            "error"
        )

        return render_template(
            "admin_login.html",
            settings=settings
        )


    # =========================================================
    # VERIFY PASSWORD
    # =========================================================

    try:

        password_valid = check_password_hash(
            stored_password_hash,
            password
        )

    except Exception:

        app.logger.exception(
            "Error verifying administrator password."
        )

        password_valid = False


    # =========================================================
    # INVALID PASSWORD
    # =========================================================

    if not password_valid:

        flash(
            "Invalid username or password.",
            "error"
        )

        return render_template(
            "admin_login.html",
            settings=settings
        )


    # =========================================================
    # LOGIN SUCCESS
    # =========================================================

    session.clear()


    # =========================================================
    # ADMIN SESSION
    # =========================================================

    # Company settings ID
    session["admin_id"] = settings["id"]

    # Admin username
    session["admin_username"] = stored_username

    # Admin display name
    session["admin_name"] = stored_username

    # Admin authentication flag
    session["admin_logged_in"] = True

    # IMPORTANT:
    # This is what the base.html sidebar uses
    session["role"] = "admin"


    # =========================================================
    # REDIRECT TO ADMIN DASHBOARD
    # =========================================================

    return redirect(
        url_for("admin_dashboard")
    )
# ============================================================
# ADMIN LOGOUT
# ============================================================
@app.route("/admin/logout")
def admin_logout():

    session.clear()

    flash(
        "You have been signed out successfully.",
        "success"
    )

    return redirect(
        url_for("admin_login")
    )
# ============================================================
# ADMIN AUTHENTICATION
# ============================================================

def admin_required():

    return bool(
        session.get("admin_id")
    )

# ============================================================
# APPLICATIONS MANAGEMENT
# ============================================================
# ============================================================
# ADMIN APPLICATIONS
# ============================================================

@app.route("/admin/applications")
def admin_applications():

    # ============================================================
    # ADMIN LOGIN CHECK
    # ============================================================

    if not admin_required():
        return redirect(url_for("admin_login"))


    # ============================================================
    # FILTERS
    # ============================================================

    search = request.args.get(
        "search",
        ""
    ).strip()

    status = request.args.get(
        "status",
        ""
    ).strip()

    position = request.args.get(
        "position",
        ""
    ).strip()


    # ============================================================
    # DATABASE CONNECTION
    # ============================================================

    conn = get_db()

    applications = []
    summary = None

    try:

        with conn.cursor() as cur:

            # ====================================================
            # APPLICATION QUERY
            # ====================================================

            query = """
                SELECT
                    id,
                    application_number,

                    first_name,
                    middle_name,
                    last_name,

                    phone,
                    email,

                    position_applied,
                    highest_qualification,

                    status,

                    passport_filename,
                    cv_filename,
                    qualification_filename,

                    submitted_at

                FROM applications

                WHERE 1=1
            """

            params = []


            # ====================================================
            # SEARCH
            # ====================================================

            if search:

                query += """
                    AND (
                        application_number ILIKE %s
                        OR first_name ILIKE %s
                        OR middle_name ILIKE %s
                        OR last_name ILIKE %s
                        OR phone ILIKE %s
                        OR email ILIKE %s
                    )
                """

                search_value = f"%{search}%"

                params.extend([
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value
                ])


            # ====================================================
            # STATUS FILTER
            # ====================================================

            if status:

                query += """
                    AND status = %s
                """

                params.append(status)


            # ====================================================
            # POSITION FILTER
            # ============================================================

            if position:

                query += """
                    AND position_applied ILIKE %s
                """

                params.append(
                    f"%{position}%"
                )


            # ====================================================
            # ORDER BY
            # ====================================================

            query += """
                ORDER BY
                    submitted_at DESC NULLS LAST,
                    id DESC
            """


            # ====================================================
            # EXECUTE APPLICATION QUERY
            # ====================================================

            cur.execute(
                query,
                params
            )

            applications = cur.fetchall()


            # ====================================================
            # SUMMARY COUNTS
            # ====================================================

            cur.execute(
                """
                SELECT
                    COUNT(*) AS total,

                    COUNT(*) FILTER (
                        WHERE status = 'Pending'
                    ) AS pending,

                    COUNT(*) FILTER (
                        WHERE status = 'Under Review'
                    ) AS under_review,

                    COUNT(*) FILTER (
                        WHERE status = 'Shortlisted'
                    ) AS shortlisted,

                    COUNT(*) FILTER (
                        WHERE status = 'Approved'
                    ) AS approved,

                    COUNT(*) FILTER (
                        WHERE status = 'Rejected'
                    ) AS rejected

                FROM applications
                """
            )

            summary = cur.fetchone()


    except Exception:

        app.logger.exception(
            "Error loading admin applications"
        )

        flash(
            "Unable to load applications at this time.",
            "error"
        )

        applications = []

        summary = {
            "total": 0,
            "pending": 0,
            "under_review": 0,
            "shortlisted": 0,
            "approved": 0,
            "rejected": 0
        }


    finally:

        conn.close()


    # ============================================================
    # SUMMARY VALUES
    # ============================================================

    total = (
        summary["total"]
        if summary
        else 0
    )

    pending = (
        summary["pending"]
        if summary
        else 0
    )

    under_review = (
        summary["under_review"]
        if summary
        else 0
    )

    shortlisted = (
        summary["shortlisted"]
        if summary
        else 0
    )

    approved = (
        summary["approved"]
        if summary
        else 0
    )

    rejected = (
        summary["rejected"]
        if summary
        else 0
    )


    # ============================================================
    # RENDER ADMIN APPLICATIONS
    # ============================================================

    return render_template(
        "admin_applications.html",

        applications=applications,

        total=total,
        pending=pending,
        under_review=under_review,
        shortlisted=shortlisted,
        approved=approved,
        rejected=rejected,

        search=search,
        selected_status=status,
        selected_position=position
    )


# ============================================================
# COMPANY SETTINGS CONTEXT
# ============================================================

@app.context_processor
def inject_company_settings():

    conn = None
    cur = None

    try:

        conn = get_db()

        cur = conn.cursor()

        cur.execute(
            """
            SELECT
                company_name,
                company_email,
                company_phone,
                company_address,
                company_website,
                footer_text,
                logo
            FROM company_settings
            ORDER BY id DESC
            LIMIT 1
            """
        )

        settings = cur.fetchone()


        if settings:

            return {
                "settings": settings,

                "company_name": (
                    settings["company_name"]
                    or "AV KING VETERINARY BUSINESS SUITE"
                ),

                "company_email": (
                    settings["company_email"]
                    or ""
                ),

                "company_phone": (
                    settings["company_phone"]
                    or ""
                ),

                "company_address": (
                    settings["company_address"]
                    or ""
                ),

                "company_website": (
                    settings["company_website"]
                    or ""
                ),

                "company_logo": (
                    settings["logo"]
                    or None
                )
            }


    except Exception:

        app.logger.exception(
            "Company settings context error"
        )


    finally:

        if cur:

            try:
                cur.close()
            except Exception:
                pass

        if conn:

            try:
                conn.close()
            except Exception:
                pass


    return {
        "settings": None,

        "company_name":
            "AV KING VETERINARY BUSINESS SUITE",

        "company_email":
            "",

        "company_phone":
            "",

        "company_address":
            "",

        "company_website":
            "",

        "company_logo":
            None
    }


# ============================================================
# VIEW APPLICATION DETAILS
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>"
)
def admin_application_details(application_id):

    # ============================================================
    # ADMIN LOGIN CHECK
    # ============================================================

    if not admin_required():
        return redirect(url_for("admin_login"))


    # ============================================================
    # DATABASE CONNECTION
    # ============================================================

    conn = get_db()

    application = None

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT *
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            application = cur.fetchone()


    except Exception:

        app.logger.exception(
            "Error loading application details: %s",
            application_id
        )

        flash(
            "Unable to load application details.",
            "error"
        )

        return redirect(
            url_for("admin_applications")
        )


    finally:

        conn.close()


    # ============================================================
    # APPLICATION NOT FOUND
    # ============================================================

    if not application:

        flash(
            "Application not found.",
            "error"
        )

        return redirect(
            url_for("admin_applications")
        )


    # ============================================================
    # RENDER APPLICATION DETAILS
    # ============================================================

    return render_template(
        "admin_application_details.html",
        application=application
    )
# ============================================================
# ADMIN SEND MESSAGE TO APPLICANT
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/message",
    methods=["POST"]
)
def send_applicant_message(application_id):

    # --------------------------------------------------------
    # CHECK ADMIN LOGIN
    # --------------------------------------------------------

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )


    # --------------------------------------------------------
    # GET FORM DATA
    # --------------------------------------------------------

    subject = (
        request.form.get(
            "subject",
            ""
        )
        .strip()
    )

    message = (
        request.form.get(
            "message",
            ""
        )
        .strip()
    )

    message_type = (
        request.form.get(
            "message_type",
            "General"
        )
        .strip()
    )


    # --------------------------------------------------------
    # VALIDATION
    # --------------------------------------------------------

    if not subject:

        flash(
            "Please enter a message subject.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )


    if not message:

        flash(
            "Please enter the message.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )


    # --------------------------------------------------------
    # SAVE MESSAGE
    # --------------------------------------------------------

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # Check applicant exists

            cur.execute(
                """
                SELECT id
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            application = cur.fetchone()

            if not application:

                flash(
                    "Application not found.",
                    "error"
                )

                return redirect(
                    url_for("admin_applications")
                )


            # Insert message

            cur.execute(
                """
                INSERT INTO applicant_messages (

                    application_id,
                    subject,
                    message,
                    message_type,
                    is_read,
                    created_at

                )

                VALUES (

                    %s,
                    %s,
                    %s,
                    %s,
                    FALSE,
                    CURRENT_TIMESTAMP

                )
                """,

                (
                    application_id,
                    subject,
                    message,
                    message_type
                )
            )


        conn.commit()


    except Exception:

        conn.rollback()

        raise


    finally:

        conn.close()


    # --------------------------------------------------------
    # SUCCESS
    # --------------------------------------------------------

    flash(
        "Message sent to applicant successfully.",
        "success"
    )

    return redirect(
        url_for(
            "admin_application_details",
            application_id=application_id
        )
    )
# ============================================================
# UPDATE ADMIN NOTES
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/notes",
    methods=["POST"]
)
def update_application_notes(application_id):

    if not admin_required():
        return redirect(url_for("admin_login"))

    notes = (
        request.form.get(
            "admin_notes",
            ""
        )
        .strip()
    )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                UPDATE applications

                SET
                    admin_notes = %s,
                    updated_at = CURRENT_TIMESTAMP

                WHERE id = %s
                """,
                (
                    notes,
                    application_id
                )
            )

        conn.commit()

    except Exception:

        conn.rollback()

        raise

    finally:

        conn.close()

    flash(
        "Recruitment notes saved successfully.",
        "success"
    )

    return redirect(
        url_for(
            "admin_application_details",
            application_id=application_id
        )
    )

# ============================================================
# DOWNLOAD APPLICATION DOCUMENT
# ============================================================

# ============================================================
# SECURE APPLICATION DOCUMENT
# ============================================================

@app.route(
    "/admin/application-file/<path:filename>"
)
def admin_application_file(filename):

    if not admin_required():
        return redirect(url_for("admin_login"))

    return send_from_directory(
        app.config["UPLOAD_FOLDER"],
        filename,
        as_attachment=False
    )

def send_shortlisted_whatsapp(application):

    phone = (
        application["phone"]
        or ""
    ).strip()

    if not phone:
        raise ValueError(
            "Applicant has no phone number."
        )


    # --------------------------------------------------------
    # COMPANY WHATSAPP NUMBER
    # --------------------------------------------------------

    company_whatsapp = os.getenv(
        "COMPANY_WHATSAPP",
        "2349058842501"
    )


    # --------------------------------------------------------
    # MESSAGE
    # --------------------------------------------------------

    message = f"""
Dear {application["first_name"]},

Congratulations!

We are pleased to inform you that you have been shortlisted for the next stage of the recruitment process at AV KING VET DRUG VENTURE.

Application Number:
{application["application_number"]}

Position:
{application["position_applied"]}

Please monitor your applicant portal for further information regarding the interview.

For enquiries, please contact us on WhatsApp:
+{company_whatsapp}

Thank you.

AV KING VET DRUG VENTURE
Your Needs, Our Priority
""".strip()


    # --------------------------------------------------------
    # TEMPORARY DEVELOPMENT MODE
    # --------------------------------------------------------

    app.logger.info(
        "SHORTLISTED WHATSAPP MESSAGE:\n%s",
        message
    )

    return True

# ============================================================
# ADMIN - SCHEDULE / UPDATE INTERVIEW
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/interview",
    methods=["POST"]
)
def update_application_interview(application_id):

    # ========================================================
    # ADMIN SECURITY
    # ========================================================

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )


    # ========================================================
    # GET FORM DATA
    # ========================================================

    interview_date = (
        request.form.get(
            "interview_date",
            ""
        ).strip()
    )

    interview_location = (
        request.form.get(
            "interview_location",
            ""
        ).strip()
    )

    interview_notes = (
        request.form.get(
            "interview_notes",
            ""
        ).strip()
    )

    interview_status = (
        request.form.get(
            "interview_status",
            "Scheduled"
        ).strip()
    )


    # ========================================================
    # VALIDATION
    # ========================================================

    allowed_interview_statuses = {
        "Scheduled",
        "Completed",
        "Rescheduled",
        "Cancelled"
    }

    if interview_status not in allowed_interview_statuses:

        flash(
            "Invalid interview status.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )


    # ========================================================
    # INTERVIEW DATE VALIDATION
    # ========================================================

    if interview_date:

        try:

            datetime.strptime(
                interview_date,
                "%Y-%m-%dT%H:%M"
            )

        except ValueError:

            flash(
                "Invalid interview date and time.",
                "error"
            )

            return redirect(
                url_for(
                    "admin_application_details",
                    application_id=application_id
                )
            )


    # ========================================================
    # DATABASE
    # ========================================================

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # ------------------------------------------------
            # CHECK APPLICATION
            # ------------------------------------------------

            cur.execute(
                """
                SELECT
                    id,
                    first_name,
                    last_name,
                    application_number,
                    status
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            application = cur.fetchone()


            if not application:

                flash(
                    "Application not found.",
                    "error"
                )

                return redirect(
                    url_for("admin_applications")
                )


            # ------------------------------------------------
            # SAVE INTERVIEW
            # ------------------------------------------------

            cur.execute(
                """
                UPDATE applications

                SET

                    interview_date = %s,

                    interview_location = %s,

                    interview_notes = %s,

                    interview_status = %s,

                    updated_at = CURRENT_TIMESTAMP

                WHERE id = %s
                """,
                (
                    interview_date or None,
                    interview_location or None,
                    interview_notes or None,
                    interview_status,
                    application_id
                )
            )


            # ------------------------------------------------
            # CREATE PORTAL MESSAGE
            # ------------------------------------------------

            if interview_date:

                cur.execute(
                    """
                    INSERT INTO applicant_messages (

                        application_id,

                        subject,

                        message,

                        message_type

                    )

                    VALUES (

                        %s,

                        %s,

                        %s,

                        %s

                    )
                    """,
                    (
                        application_id,

                        "Interview Scheduled",

                        (
                            "Your interview has been scheduled.\n\n"
                            f"Date & Time: {interview_date}\n"
                            f"Location: {interview_location or 'To be confirmed'}\n\n"
                            f"Instructions: "
                            f"{interview_notes or 'Please check your applicant portal for further updates.'}"
                        ),

                        "Interview"
                    )
                )


        conn.commit()


    except Exception:

        conn.rollback()

        raise


    finally:

        conn.close()


    # ========================================================
    # SUCCESS
    # ========================================================

    flash(
        "Interview details saved successfully.",
        "success"
    )


    return redirect(
        url_for(
            "admin_application_details",
            application_id=application_id
        )
    )
@app.route("/applicant/dashboard")
def applicant_dashboard():

    if not session.get("applicant_id"):
        return redirect(url_for("applicant_login"))

    applicant_id = session["applicant_id"]

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT *
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (applicant_id,)
            )

            application = cur.fetchone()

    finally:

        conn.close()


    if not application:

        session.clear()

        return redirect(
            url_for("applicant_login")
        )


    return render_template(
        "applicant_dashboard.html",
        application=application
    )

@app.route(
    "/admin/positions",
    methods=["GET", "POST"]
)
def admin_positions():

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # ==================================================
            # ADD POSITION
            # ==================================================

            if request.method == "POST":

                position_name = (
                    request.form.get(
                        "position_name",
                        ""
                    ).strip()
                )

                description = (
                    request.form.get(
                        "description",
                        ""
                    ).strip()
                )

                if not position_name:

                    flash(
                        "Please enter a position name.",
                        "error"
                    )

                else:

                    cur.execute(
                        """
                        INSERT INTO available_positions
                        (
                            position_name,
                            description,
                            is_active
                        )
                        VALUES
                        (
                            %s,
                            %s,
                            TRUE
                        )
                        ON CONFLICT (position_name)
                        DO UPDATE SET
                            description = EXCLUDED.description,
                            is_active = TRUE
                        """,
                        (
                            position_name,
                            description or None
                        )
                    )

                    conn.commit()

                    flash(
                        "Position added successfully.",
                        "success"
                    )

                    return redirect(
                        url_for(
                            "admin_positions"
                        )
                    )


            # ==================================================
            # GET POSITIONS
            # ==================================================

            cur.execute(
                """
                SELECT
                    id,
                    position_name,
                    description,
                    is_active,
                    created_at
                FROM available_positions
                ORDER BY
                    is_active DESC,
                    position_name ASC
                """
            )

            positions = cur.fetchall()


    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error managing available positions"
        )

        flash(
            "Unable to manage positions.",
            "error"
        )

        positions = []


    finally:

        conn.close()


    return render_template(
        "admin_positions.html",
        positions=positions
    )

@app.route(
    "/admin/positions/<int:position_id>/edit",
    methods=["POST"]
)
def admin_edit_position(position_id):

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    position_name = (
        request.form.get(
            "position_name",
            ""
        ).strip()
    )

    description = (
        request.form.get(
            "description",
            ""
        ).strip()
    )

    if not position_name:

        flash(
            "Position name is required.",
            "error"
        )

        return redirect(
            url_for("admin_positions")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                UPDATE available_positions

                SET
                    position_name = %s,
                    description = %s

                WHERE id = %s
                """,
                (
                    position_name,
                    description or None,
                    position_id
                )
            )

        conn.commit()

        flash(
            "Position updated successfully.",
            "success"
        )

    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error editing position"
        )

        flash(
            "Unable to update position.",
            "error"
        )

    finally:

        conn.close()

    return redirect(
        url_for("admin_positions")
    )

@app.route(
    "/admin/positions/<int:position_id>/toggle",
    methods=["POST"]
)
def admin_toggle_position(position_id):

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                UPDATE available_positions

                SET is_active =
                    NOT is_active

                WHERE id = %s
                """,
                (position_id,)
            )

        conn.commit()

        flash(
            "Position status updated.",
            "success"
        )

    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error changing position status"
        )

        flash(
            "Unable to change position status.",
            "error"
        )

    finally:

        conn.close()

    return redirect(
        url_for("admin_positions")
    )

# ============================================================
# SEND WHATSAPP NOTIFICATION
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/whatsapp"
)
def send_application_whatsapp(application_id):

    if not admin_required():
        return redirect(url_for("admin_login"))

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    id,
                    first_name,
                    last_name,
                    phone,
                    application_number,
                    position_applied,
                    status
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            application = cur.fetchone()

    finally:

        conn.close()


    if not application:

        flash(
            "Application not found.",
            "error"
        )

        return redirect(
            url_for("admin_applications")
        )


    # ========================================================
    # ONLY SHORTLISTED APPLICANTS
    # ========================================================

    if application["status"] != "Shortlisted":

        flash(
            "WhatsApp notification is only available "
            "for shortlisted applicants.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )


    phone = (
        application["phone"] or ""
    ).strip()


    if not phone:

        flash(
            "This applicant does not have a valid phone number.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )


    # ========================================================
    # CONVERT NIGERIAN NUMBER TO INTERNATIONAL FORMAT
    # ========================================================

    whatsapp_phone = phone

    if whatsapp_phone.startswith("0"):

        whatsapp_phone = (
            "234"
            + whatsapp_phone[1:]
        )

    elif whatsapp_phone.startswith("+234"):

        whatsapp_phone = whatsapp_phone[1:]


    # ========================================================
    # MESSAGE
    # ========================================================

    applicant_name = (
        f"{application['first_name']} "
        f"{application['last_name']}"
    )


    message = (
        f"Dear {applicant_name},\n\n"

        f"Congratulations!\n\n"

        f"We are pleased to inform you that your "
        f"application for the position of "
        f"{application['position_applied']} at "
        f"{COMPANY_NAME} has been shortlisted "
        f"for the next stage of our recruitment process.\n\n"

        f"Application Number: "
        f"{application['application_number']}\n\n"

        f"Please log in to your Applicant Portal regularly "
        f"to check for further recruitment updates and "
        f"interview information.\n\n"

        f"Regards,\n"
        f"{COMPANY_NAME}\n"
        f"{COMPANY_PHONE}"
    )


    # ========================================================
    # WHATSAPP URL
    # ========================================================

    whatsapp_url = (
        "https://wa.me/"
        + whatsapp_phone
        + "?text="
        + quote(message)
    )


    return redirect(whatsapp_url)
# ============================================================
# UPDATE APPLICATION STATUS
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/status",
    methods=["POST"]
)
def update_application_status(application_id):

    if not admin_required():
        return redirect(url_for("admin_login"))

    new_status = (
        request.form.get("status", "")
        .strip()
    )

    allowed_statuses = {
        "Pending",
        "Under Review",
        "Shortlisted",
        "Approved",
        "Rejected"
    }

    if new_status not in allowed_statuses:

        flash(
            "Invalid application status.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET APPLICATION
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    first_name,
                    last_name,
                    phone,
                    application_number,
                    position_applied,
                    status,
                    notification_sent
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            application = cur.fetchone()


            if not application:

                flash(
                    "Application not found.",
                    "error"
                )

                return redirect(
                    url_for("admin_applications")
                )


            # =================================================
            # SHORTLISTED
            # =================================================

            if new_status == "Shortlisted":

                cur.execute(
                    """
                    UPDATE applications

                    SET
                        status = %s,

                        shortlisted_at =
                            COALESCE(
                                shortlisted_at,
                                CURRENT_TIMESTAMP
                            ),

                        updated_at =
                            CURRENT_TIMESTAMP

                    WHERE id = %s
                    """,
                    (
                        new_status,
                        application_id
                    )
                )


            # =================================================
            # OTHER STATUSES
            # =================================================

            else:

                cur.execute(
                    """
                    UPDATE applications

                    SET
                        status = %s,

                        updated_at =
                            CURRENT_TIMESTAMP

                    WHERE id = %s
                    """,
                    (
                        new_status,
                        application_id
                    )
                )


        conn.commit()


    except Exception:

        conn.rollback()

        raise

    finally:

        conn.close()


    # =========================================================
    # SHORTLISTED WHATSAPP NOTIFICATION
    # =========================================================

    if new_status == "Shortlisted":

        applicant_name = (
            f"{application['first_name']} "
            f"{application['last_name']}"
        )

        phone = (
            application["phone"]
            or ""
        ).strip()

        application_number = (
            application["application_number"]
        )

        position = (
            application["position_applied"]
        )


        # -----------------------------------------------------
        # NORMALIZE NIGERIAN PHONE NUMBER
        # -----------------------------------------------------

        whatsapp_phone = phone

        if whatsapp_phone.startswith("0"):

            whatsapp_phone = (
                "234"
                + whatsapp_phone[1:]
            )

        elif whatsapp_phone.startswith("+234"):

            whatsapp_phone = (
                whatsapp_phone[1:]
            )


        # -----------------------------------------------------
        # CREATE WHATSAPP MESSAGE
        # -----------------------------------------------------

        whatsapp_message = (
            f"Dear {applicant_name},\n\n"

            f"Congratulations!\n\n"

            f"We are pleased to inform you that your "
            f"application for the position of "
            f"{position} at {COMPANY_NAME} "
            f"has been shortlisted for the next stage "
            f"of our recruitment process.\n\n"

            f"Application Number: "
            f"{application_number}\n\n"

            f"Please log in to your Applicant Portal "
            f"regularly to check for further recruitment "
            f"updates and interview information.\n\n"

            f"Regards,\n"
            f"{COMPANY_NAME}\n"
            f"{COMPANY_PHONE}"
        )


        # -----------------------------------------------------
        # WHATSAPP LINK
        # -----------------------------------------------------

        from urllib.parse import quote

        whatsapp_url = (
            f"https://wa.me/"
            f"{whatsapp_phone}"
            f"?text="
            f"{quote(whatsapp_message)}"
        )


        # =====================================================
        # RECORD NOTIFICATION STATE
        # =====================================================

        conn = get_db()

        try:

            with conn.cursor() as cur:

                cur.execute(
                    """
                    UPDATE applications

                    SET
                        notification_sent = TRUE,
                        notification_sent_at =
                            CURRENT_TIMESTAMP,

                        updated_at =
                            CURRENT_TIMESTAMP

                    WHERE id = %s
                    """,
                    (application_id,)
                )

            conn.commit()

        except Exception:

            conn.rollback()

            raise

        finally:

            conn.close()


        # =====================================================
        # SHOW ADMIN WHATSAPP ACTION
        # =====================================================

        flash(
            "Applicant shortlisted successfully. "
            "WhatsApp notification is ready to send.",
            "success"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id,
                whatsapp_url=whatsapp_url
            )
        )


    # =========================================================
    # OTHER STATUS
    # =========================================================

    flash(
        "Application status updated successfully.",
        "success"
    )

    return redirect(
        url_for(
            "admin_application_details",
            application_id=application_id
        )
    )

# =========================================================
# NIGERIA DATE / TIME HELPER
# =========================================================

from datetime import datetime
from zoneinfo import ZoneInfo


def get_nigeria_now():
    """
    Return the current date and time in Nigeria.

    Nigeria uses the Africa/Lagos timezone.
    """

    return datetime.now(
        ZoneInfo("Africa/Lagos")
    )
# ============================================================
# ADMIN DASHBOARD
# ============================================================

@app.route("/admin")
@app.route("/admin/dashboard")
def admin_dashboard():

    if not admin_required():

        return redirect(
            url_for("admin_login")
        )


    conn = get_db()

    try:

        with conn.cursor() as cur:

            # ----------------------------------------------
            # TOTAL APPLICATIONS
            # ----------------------------------------------

            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM applications
                """
            )

            total = cur.fetchone()["total"]


            # ----------------------------------------------
            # PENDING
            # ----------------------------------------------

            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Pending'
                """
            )

            pending = cur.fetchone()["total"]


            # ----------------------------------------------
            # UNDER REVIEW
            # ----------------------------------------------

            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Under Review'
                """
            )

            under_review = (
                cur.fetchone()["total"]
            )


            # ----------------------------------------------
            # SHORTLISTED
            # ----------------------------------------------

            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Shortlisted'
                """
            )

            shortlisted = (
                cur.fetchone()["total"]
            )


            # ----------------------------------------------
            # APPROVED
            # ----------------------------------------------

            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Approved'
                """
            )

            approved = (
                cur.fetchone()["total"]
            )


            # ----------------------------------------------
            # REJECTED
            # ----------------------------------------------

            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Rejected'
                """
            )

            rejected = (
                cur.fetchone()["total"]
            )


            # ----------------------------------------------
            # RECENT APPLICATIONS
            # ----------------------------------------------

            cur.execute(
                """
                SELECT

                    id,

                    application_number,

                    first_name,

                    middle_name,

                    last_name,

                    phone,

                    position_applied,

                    status,

                    submitted_at

                FROM applications

                ORDER BY submitted_at DESC

                LIMIT 10
                """
            )

            recent_applications = cur.fetchall()


    finally:

        conn.close()


    return render_template(
        "admin_dashboard.html",

        total=total,

        pending=pending,

        under_review=under_review,

        shortlisted=shortlisted,

        approved=approved,

        rejected=rejected,

        recent_applications=recent_applications
    )

# =========================================================
# ATTENDANCE STATUS HELPER
# =========================================================

def calculate_attendance_status(
    attendance,
    clock_in_start=None
):
    """
    Determine the official attendance status.

    Possible statuses:
        Present
        Late
        Incomplete
        Absent
    """

    # -----------------------------------------------------
    # NO ATTENDANCE RECORD
    # -----------------------------------------------------

    if not attendance:
        return "Absent"


    clock_in = attendance.get("clock_in")
    clock_out = attendance.get("clock_out")


    # -----------------------------------------------------
    # CLOCKED IN BUT NOT CLOCKED OUT
    # -----------------------------------------------------

    if clock_in and not clock_out:
        return "Incomplete"


    # -----------------------------------------------------
    # NO VALID CLOCK-IN
    # -----------------------------------------------------

    if not clock_in:
        return "Absent"


    # -----------------------------------------------------
    # CLOCKED IN + CLOCKED OUT
    # -----------------------------------------------------

    if clock_in and clock_out:

        # -----------------------------------------------
        # CHECK LATE
        # -----------------------------------------------

        if clock_in_start:

            try:

                if clock_in.time() > clock_in_start:
                    return "Late"

            except Exception:

                pass


        return "Present"


    # -----------------------------------------------------
    # FALLBACK
    # -----------------------------------------------------

    return (
        attendance.get("status")
        or "Absent"
    )

# =========================================================
# GET ATTENDANCE SETTINGS
# =========================================================

def get_attendance_settings(conn):

    with conn.cursor() as cur:

        cur.execute(
            """
            SELECT
                attendance_enabled,
                company_latitude,
                company_longitude,
                attendance_radius,
                clock_in_start,
                clock_out_end

            FROM company_settings

            ORDER BY id ASC

            LIMIT 1
            """
        )

        settings = cur.fetchone()


    # -----------------------------------------------------
    # DEFAULT SETTINGS
    # -----------------------------------------------------

    if not settings:

        return {
            "attendance_enabled": True,
            "company_latitude": None,
            "company_longitude": None,
            "attendance_radius": 200,
            "clock_in_start": None,
            "clock_out_end": None
        }


    return settings
# ============================================================
# ADMIN — ATTENDANCE MANAGEMENT
# ============================================================
@app.route("/admin/attendance")
def admin_attendance():

    # =========================================================
    # ADMIN ACCESS
    # =========================================================

    if not admin_required():
        return redirect(url_for("admin_login"))

    # =========================================================
    # SELECTED DATE
    # =========================================================

    selected_date = (
        request.args.get("date", "").strip()
        or datetime.now().strftime("%Y-%m-%d")
    )

    search = (
        request.args.get("search", "").strip()
    )

    conn = get_db()

    attendance_records = []

    total_staff = 0
    present_count = 0
    absent_count = 0
    incomplete_count = 0
    total_hours = 0

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET APPROVED / ACTIVE WORKERS
            # =================================================

            worker_query = """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    phone,
                    email,
                    position_applied,
                    status
                FROM applications
                WHERE
                    status = 'Approved'
                    AND portal_active = TRUE
            """

            worker_params = []

            # -------------------------------------------------
            # SEARCH
            # -------------------------------------------------

            if search:

                worker_query += """
                    AND (
                        application_number ILIKE %s
                        OR first_name ILIKE %s
                        OR middle_name ILIKE %s
                        OR last_name ILIKE %s
                        OR phone ILIKE %s
                        OR email ILIKE %s
                        OR position_applied ILIKE %s
                    )
                """

                search_value = f"%{search}%"

                worker_params.extend([
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value
                ])

            worker_query += """
                ORDER BY first_name ASC, last_name ASC
            """

            cur.execute(
                worker_query,
                worker_params
            )

            workers = cur.fetchall()

            total_staff = len(workers)

            # =================================================
            # GET ATTENDANCE FOR SELECTED DATE
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    worker_id,
                    attendance_date,
                    clock_in,
                    clock_out,

                    clock_in_latitude,
                    clock_in_longitude,

                    clock_out_latitude,
                    clock_out_longitude,

                    clock_in_location_verified,
                    clock_out_location_verified,

                    total_hours,
                    status,
                    notes

                FROM attendance

                WHERE attendance_date = %s
                """,
                (selected_date,)
            )

            attendance_rows = cur.fetchall()

            # =================================================
            # CONVERT TO DICTIONARY
            # =================================================

            attendance_map = {}

            for row in attendance_rows:

                attendance_map[
                    row["worker_id"]
                ] = row

            # =================================================
            # BUILD COMPLETE ATTENDANCE LIST
            # =================================================

            for worker in workers:

                attendance = attendance_map.get(
                    worker["id"]
                )

                # -------------------------------------------------
                # PRESENT
                # -------------------------------------------------

                if attendance:

                    clock_in = attendance["clock_in"]
                    clock_out = attendance["clock_out"]

                    if clock_in and clock_out:

                        record_status = "Present"

                        present_count += 1

                    elif clock_in:

                        record_status = "Incomplete"

                        incomplete_count += 1

                    else:

                        record_status = "Absent"

                        absent_count += 1

                    hours = (
                        float(
                            attendance["total_hours"] or 0
                        )
                    )

                    total_hours += hours

                    attendance_records.append({

                        "worker_id": worker["id"],

                        "application_number":
                            worker["application_number"],

                        "first_name":
                            worker["first_name"],

                        "middle_name":
                            worker["middle_name"],

                        "last_name":
                            worker["last_name"],

                        "phone":
                            worker["phone"],

                        "email":
                            worker["email"],

                        "position_applied":
                            worker["position_applied"],

                        "attendance_id":
                            attendance["id"],

                        "clock_in":
                            clock_in,

                        "clock_out":
                            clock_out,

                        "clock_in_latitude":
                            attendance[
                                "clock_in_latitude"
                            ],

                        "clock_in_longitude":
                            attendance[
                                "clock_in_longitude"
                            ],

                        "clock_out_latitude":
                            attendance[
                                "clock_out_latitude"
                            ],

                        "clock_out_longitude":
                            attendance[
                                "clock_out_longitude"
                            ],

                        "clock_in_location_verified":
                            attendance[
                                "clock_in_location_verified"
                            ],

                        "clock_out_location_verified":
                            attendance[
                                "clock_out_location_verified"
                            ],

                        "total_hours":
                            hours,

                        "status":
                            record_status,

                        "notes":
                            attendance["notes"]

                    })

                # -------------------------------------------------
                # NO ATTENDANCE RECORD = ABSENT
                # -------------------------------------------------

                else:

                    absent_count += 1

                    attendance_records.append({

                        "worker_id":
                            worker["id"],

                        "application_number":
                            worker["application_number"],

                        "first_name":
                            worker["first_name"],

                        "middle_name":
                            worker["middle_name"],

                        "last_name":
                            worker["last_name"],

                        "phone":
                            worker["phone"],

                        "email":
                            worker["email"],

                        "position_applied":
                            worker["position_applied"],

                        "attendance_id":
                            None,

                        "clock_in":
                            None,

                        "clock_out":
                            None,

                        "clock_in_latitude":
                            None,

                        "clock_in_longitude":
                            None,

                        "clock_out_latitude":
                            None,

                        "clock_out_longitude":
                            None,

                        "clock_in_location_verified":
                            False,

                        "clock_out_location_verified":
                            False,

                        "total_hours":
                            0,

                        "status":
                            "Absent",

                        "notes":
                            None

                    })

    except Exception:

        app.logger.exception(
            "Error loading admin attendance"
        )

        flash(
            "Unable to load attendance records.",
            "error"
        )

    finally:

        conn.close()

    # =========================================================
    # RENDER
    # =========================================================

    return render_template(

        "admin_attendance.html",

        attendance_records=attendance_records,

        selected_date=selected_date,

        search=search,

        total_staff=total_staff,

        present_count=present_count,

        absent_count=absent_count,

        incomplete_count=incomplete_count,

        total_hours=round(
            total_hours,
            2
        )

    )

@app.route("/admin/attendance/monthly")
def admin_monthly_attendance():

    # =========================================================
    # ADMIN ACCESS
    # =========================================================

    if not admin_required():
        return redirect(url_for("admin_login"))

    # =========================================================
    # SELECTED MONTH
    # =========================================================

    selected_month = (
        request.args.get("month", "").strip()
        or datetime.now().strftime("%Y-%m")
    )

    # Validate month
    try:

        month_start = datetime.strptime(
            selected_month,
            "%Y-%m"
        ).date()

    except ValueError:

        selected_month = datetime.now().strftime("%Y-%m")

        month_start = datetime.strptime(
            selected_month,
            "%Y-%m"
        ).date()

    # First day of next month
    if month_start.month == 12:

        next_month = month_start.replace(
            year=month_start.year + 1,
            month=1,
            day=1
        )

    else:

        next_month = month_start.replace(
            month=month_start.month + 1,
            day=1
        )

    month_end = next_month - timedelta(days=1)

    conn = get_db()

    attendance_records = []

    total_staff = 0
    total_present = 0
    total_absent = 0
    total_late = 0
    total_incomplete = 0
    total_hours = 0

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET APPROVED / ACTIVE STAFF
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    phone,
                    email,
                    position_applied

                FROM applications

                WHERE
                    status = 'Approved'
                    AND portal_active = TRUE

                ORDER BY
                    first_name ASC,
                    last_name ASC
                """
            )

            workers = cur.fetchall()

            total_staff = len(workers)

            # =================================================
            # GET ATTENDANCE FOR THE MONTH
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    worker_id,
                    attendance_date,
                    clock_in,
                    clock_out,
                    total_hours,
                    status,
                    clock_in_location_verified,
                    clock_out_location_verified

                FROM attendance

                WHERE
                    attendance_date >= %s
                    AND attendance_date < %s

                ORDER BY
                    attendance_date ASC
                """,
                (
                    month_start,
                    next_month
                )
            )

            attendance_rows = cur.fetchall()

            # =================================================
            # ORGANIZE ATTENDANCE
            # =================================================

            attendance_map = {}

            for row in attendance_rows:

                worker_id = row["worker_id"]
                attendance_date = row["attendance_date"]

                attendance_map[
                    (
                        worker_id,
                        attendance_date
                    )
                ] = row

            # =================================================
            # WORKING DAYS
            # =================================================

            working_days = []

            current_date = month_start

            while current_date < next_month:

                # Monday = 0
                # Sunday = 6

                if current_date.weekday() < 5:

                    working_days.append(
                        current_date
                    )

                current_date += timedelta(days=1)

            total_working_days = len(
                working_days
            )

            # =================================================
            # BUILD EMPLOYEE MONTHLY REPORT
            # =================================================

            for worker in workers:

                worker_present = 0
                worker_absent = 0
                worker_late = 0
                worker_incomplete = 0
                worker_hours = 0

                daily_records = []

                for attendance_date in working_days:

                    attendance = attendance_map.get(
                        (
                            worker["id"],
                            attendance_date
                        )
                    )

                    # -----------------------------------------
                    # NO RECORD = ABSENT
                    # -----------------------------------------

                    if not attendance:

                        status = "Absent"

                        worker_absent += 1

                        total_absent += 1

                        daily_records.append({

                            "date": attendance_date,

                            "clock_in": None,

                            "clock_out": None,

                            "total_hours": 0,

                            "status": "Absent"

                        })

                        continue

                    # -----------------------------------------
                    # GET ATTENDANCE DATA
                    # -----------------------------------------

                    clock_in = attendance["clock_in"]
                    clock_out = attendance["clock_out"]

                    hours = float(
                        attendance["total_hours"] or 0
                    )

                    worker_hours += hours
                    total_hours += hours

                    # -----------------------------------------
                    # DETERMINE STATUS
                    # -----------------------------------------

                    if clock_in and clock_out:

                        raw_status = (
                            attendance["status"]
                            or "Present"
                        )

                        if raw_status.lower() == "late":

                            status = "Late"

                            worker_late += 1
                            total_late += 1

                        else:

                            status = "Present"

                            worker_present += 1
                            total_present += 1

                    elif clock_in:

                        status = "Incomplete"

                        worker_incomplete += 1
                        total_incomplete += 1

                    else:

                        status = "Absent"

                        worker_absent += 1
                        total_absent += 1

                    daily_records.append({

                        "date": attendance_date,

                        "clock_in": clock_in,

                        "clock_out": clock_out,

                        "total_hours": hours,

                        "status": status

                    })

                # =================================================
                # ATTENDANCE PERCENTAGE
                # =================================================

                if total_working_days > 0:

                    attendance_percentage = (
                        (
                            worker_present
                            + worker_late
                        )
                        / total_working_days
                    ) * 100

                else:

                    attendance_percentage = 0

                # =================================================
                # ADD EMPLOYEE REPORT
                # =================================================

                attendance_records.append({

                    "worker_id":
                        worker["id"],

                    "application_number":
                        worker["application_number"],

                    "first_name":
                        worker["first_name"],

                    "middle_name":
                        worker["middle_name"],

                    "last_name":
                        worker["last_name"],

                    "phone":
                        worker["phone"],

                    "email":
                        worker["email"],

                    "position_applied":
                        worker["position_applied"],

                    "present":
                        worker_present,

                    "absent":
                        worker_absent,

                    "late":
                        worker_late,

                    "incomplete":
                        worker_incomplete,

                    "total_hours":
                        round(
                            worker_hours,
                            2
                        ),

                    "attendance_percentage":
                        round(
                            attendance_percentage,
                            1
                        ),

                    "daily_records":
                        daily_records

                })

    except Exception:

        app.logger.exception(
            "Error loading monthly attendance"
        )

        flash(
            "Unable to load monthly attendance report.",
            "error"
        )

    finally:

        conn.close()

    # =========================================================
    # RENDER
    # =========================================================

    return render_template(
        "admin_monthly_attendance.html",

        attendance_records=attendance_records,

        selected_month=selected_month,

        month_start=month_start,

        month_end=month_end,

        total_working_days=total_working_days,

        total_staff=total_staff,

        total_present=total_present,

        total_absent=total_absent,

        total_late=total_late,

        total_incomplete=total_incomplete,

        total_hours=round(
            total_hours,
            2
        )
    )
@app.route("/admin/attendance/monthly/pdf")
def admin_monthly_attendance_pdf():

    # =========================================================
    # ADMIN ACCESS
    # =========================================================

    if not admin_required():
        return redirect(url_for("admin_login"))

    # =========================================================
    # SELECTED MONTH
    # =========================================================

    selected_month = (
        request.args.get("month", "").strip()
        or datetime.now().strftime("%Y-%m")
    )

    try:

        month_start = datetime.strptime(
            selected_month,
            "%Y-%m"
        ).date()

    except ValueError:

        selected_month = datetime.now().strftime("%Y-%m")

        month_start = datetime.strptime(
            selected_month,
            "%Y-%m"
        ).date()

    # =========================================================
    # NEXT MONTH
    # =========================================================

    if month_start.month == 12:

        next_month = month_start.replace(
            year=month_start.year + 1,
            month=1,
            day=1
        )

    else:

        next_month = month_start.replace(
            month=month_start.month + 1,
            day=1
        )

    month_end = next_month - timedelta(days=1)

    # =========================================================
    # WORKING DAYS
    # =========================================================

    working_days = []

    current_date = month_start

    while current_date < next_month:

        if current_date.weekday() < 5:

            working_days.append(current_date)

        current_date += timedelta(days=1)

    total_working_days = len(working_days)

    # =========================================================
    # DATABASE
    # =========================================================

    conn = get_db()

    workers = []
    attendance_rows = []
    settings = None

    try:

        with conn.cursor() as cur:

            # =================================================
            # COMPANY SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT *
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            settings = cur.fetchone()

            # =================================================
            # APPROVED STAFF
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    phone,
                    email,
                    position_applied

                FROM applications

                WHERE
                    status = 'Approved'
                    AND portal_active = TRUE

                ORDER BY
                    first_name ASC,
                    last_name ASC
                """
            )

            workers = cur.fetchall()

            # =================================================
            # ATTENDANCE
            # =================================================

            cur.execute(
                """
                SELECT
                    worker_id,
                    attendance_date,
                    clock_in,
                    clock_out,
                    total_hours,
                    status

                FROM attendance

                WHERE
                    attendance_date >= %s
                    AND attendance_date < %s

                ORDER BY
                    attendance_date ASC
                """,
                (
                    month_start,
                    next_month
                )
            )

            attendance_rows = cur.fetchall()

    except Exception:

        app.logger.exception(
            "Error preparing monthly attendance PDF"
        )

        flash(
            "Unable to generate monthly attendance PDF.",
            "error"
        )

        return redirect(
            url_for(
                "admin_monthly_attendance",
                month=selected_month
            )
        )

    finally:

        conn.close()

    # =========================================================
    # ATTENDANCE MAP
    # =========================================================

    attendance_map = {}

    for row in attendance_rows:

        attendance_map[
            (
                row["worker_id"],
                row["attendance_date"]
            )
        ] = row

    # =========================================================
    # BUILD REPORT DATA
    # =========================================================

    report_rows = []

    total_present = 0
    total_absent = 0
    total_late = 0
    total_incomplete = 0
    total_hours = 0

    for worker in workers:

        present = 0
        absent = 0
        late = 0
        incomplete = 0
        worker_hours = 0

        for day in working_days:

            attendance = attendance_map.get(
                (
                    worker["id"],
                    day
                )
            )

            if not attendance:

                absent += 1
                total_absent += 1

                continue

            clock_in = attendance["clock_in"]
            clock_out = attendance["clock_out"]

            hours = float(
                attendance["total_hours"] or 0
            )

            worker_hours += hours
            total_hours += hours

            if clock_in and clock_out:

                if (
                    attendance["status"]
                    and attendance["status"].lower() == "late"
                ):

                    late += 1
                    total_late += 1

                else:

                    present += 1
                    total_present += 1

            elif clock_in:

                incomplete += 1
                total_incomplete += 1

            else:

                absent += 1
                total_absent += 1

        # =====================================================
        # ATTENDANCE %
        # =====================================================

        if total_working_days:

            attendance_percentage = (
                (
                    present
                    + late
                )
                / total_working_days
            ) * 100

        else:

            attendance_percentage = 0

        report_rows.append({

            "application_number":
                worker["application_number"],

            "name":
                " ".join(
                    filter(
                        None,
                        [
                            worker["first_name"],
                            worker["middle_name"],
                            worker["last_name"]
                        ]
                    )
                ),

            "position":
                worker["position_applied"] or "—",

            "present":
                present,

            "late":
                late,

            "absent":
                absent,

            "incomplete":
                incomplete,

            "hours":
                round(
                    worker_hours,
                    2
                ),

            "percentage":
                round(
                    attendance_percentage,
                    1
                )

        })

    # =========================================================
    # PDF IMPORTS
    # =========================================================

    from io import BytesIO

    from reportlab.lib import colors

    from reportlab.lib.pagesizes import A4

    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle
    )

    from reportlab.lib.enums import (
        TA_CENTER,
        TA_LEFT
    )

    from reportlab.lib.units import mm

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image
    )

    # =========================================================
    # PDF BUFFER
    # =========================================================

    buffer = BytesIO()

    # =========================================================
    # DOCUMENT
    # =========================================================

    doc = SimpleDocTemplate(

        buffer,

        pagesize=A4,

        rightMargin=12 * mm,

        leftMargin=12 * mm,

        topMargin=12 * mm,

        bottomMargin=15 * mm

    )

    # =========================================================
    # STYLES
    # =========================================================

    styles = getSampleStyleSheet()

    company_style = ParagraphStyle(

        "CompanyName",

        parent=styles["Heading1"],

        fontSize=16,

        leading=20,

        alignment=TA_CENTER,

        spaceAfter=3

    )

    address_style = ParagraphStyle(

        "CompanyAddress",

        parent=styles["Normal"],

        fontSize=8.5,

        leading=11,

        alignment=TA_CENTER

    )

    title_style = ParagraphStyle(

        "ReportTitle",

        parent=styles["Heading2"],

        fontSize=13,

        leading=16,

        alignment=TA_CENTER,

        spaceBefore=8,

        spaceAfter=3

    )

    subtitle_style = ParagraphStyle(

        "ReportSubtitle",

        parent=styles["Normal"],

        fontSize=8.5,

        leading=11,

        alignment=TA_CENTER,

        spaceAfter=10

    )

    normal_style = ParagraphStyle(

        "NormalSmall",

        parent=styles["Normal"],

        fontSize=7.5,

        leading=9

    )

    # =========================================================
    # STORY
    # =========================================================

    story = []

    # =========================================================
    # LOGO
    # =========================================================

    logo_path = None

    if settings and settings["logo"]:

        possible_logo = os.path.join(

            app.root_path,

            "static",

            "uploads",

            settings["logo"]

        )

        if os.path.exists(possible_logo):

            logo_path = possible_logo

    if logo_path:

        try:

            logo = Image(
                logo_path,
                width=24 * mm,
                height=24 * mm
            )

            logo.hAlign = "CENTER"

            story.append(logo)

            story.append(
                Spacer(
                    1,
                    3 * mm
                )
            )

        except Exception:

            app.logger.warning(
                "Unable to load company logo for PDF.",
                exc_info=True
            )

    # =========================================================
    # COMPANY INFORMATION
    # =========================================================

    company_name = (
        settings["company_name"]
        if settings
        and settings["company_name"]
        else "AV KING VET DRUG VENTURE"
    )

    company_address = (
        settings["company_address"]
        if settings
        and settings["company_address"]
        else ""
    )

    company_phone = (
        settings["company_phone"]
        if settings
        and settings["company_phone"]
        else ""
    )

    company_email = (
        settings["company_email"]
        if settings
        and settings["company_email"]
        else ""
    )

    story.append(
        Paragraph(
            str(company_name),
            company_style
        )
    )

    contact_line = " | ".join(
        filter(
            None,
            [
                str(company_address),
                str(company_phone),
                str(company_email)
            ]
        )
    )

    if contact_line:

        story.append(
            Paragraph(
                contact_line,
                address_style
            )
        )

    # =========================================================
    # TITLE
    # =========================================================

    story.append(
        Paragraph(
            "MONTHLY ATTENDANCE REPORT",
            title_style
        )
    )

    story.append(
        Paragraph(
            month_start.strftime("%B %Y"),
            subtitle_style
        )
    )

    # =========================================================
    # SUMMARY TABLE
    # =========================================================

    summary_data = [

        [
            "ACTIVE STAFF",
            "WORKING DAYS",
            "PRESENT",
            "LATE",
            "ABSENT",
            "INCOMPLETE",
            "TOTAL HOURS"
        ],

        [
            str(len(workers)),
            str(total_working_days),
            str(total_present),
            str(total_late),
            str(total_absent),
            str(total_incomplete),
            f"{total_hours:.2f}"
        ]

    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            25 * mm,
            27 * mm,
            22 * mm,
            20 * mm,
            22 * mm,
            25 * mm,
            27 * mm
        ]
    )

    summary_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1f2937")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.HexColor("#d1d5db")
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            )

        ])
    )

    story.append(summary_table)

    story.append(
        Spacer(
            1,
            7 * mm
        )
    )

    # =========================================================
    # EMPLOYEE TABLE
    # =========================================================

    table_data = [

        [
            "#",
            "Application No.",
            "Employee",
            "Position",
            "Present",
            "Late",
            "Absent",
            "Incomplete",
            "Hours",
            "Attendance"
        ]

    ]

    for index, row in enumerate(
        report_rows,
        start=1
    ):

        table_data.append([

            str(index),

            row["application_number"],

            Paragraph(
                row["name"],
                normal_style
            ),

            Paragraph(
                row["position"],
                normal_style
            ),

            str(row["present"]),

            str(row["late"]),

            str(row["absent"]),

            str(row["incomplete"]),

            f'{row["hours"]:.2f}',

            f'{row["percentage"]:.1f}%'

        ])

    employee_table = Table(

        table_data,

        repeatRows=1,

        colWidths=[

            8 * mm,

            27 * mm,

            37 * mm,

            30 * mm,

            14 * mm,

            12 * mm,

            14 * mm,

            18 * mm,

            17 * mm,

            20 * mm

        ]

    )

    employee_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#111827")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, 0),
                6.5
            ),

            (
                "FONTSIZE",
                (0, 1),
                (-1, -1),
                6.8
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "ALIGN",
                (0, 0),
                (0, -1),
                "CENTER"
            ),

            (
                "ALIGN",
                (4, 1),
                (-1, -1),
                "CENTER"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.35,
                colors.HexColor("#d1d5db")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f9fafb")
                ]
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                5
            )

        ])
    )

    story.append(employee_table)

    story.append(
        Spacer(
            1,
            8 * mm
        )
    )

    # =========================================================
    # FOOTER
    # =========================================================

    footer_text = (

        settings["footer_text"]

        if settings
        and settings["footer_text"]

        else "Your needs, Our Priority"

    )

    story.append(
        Paragraph(
            footer_text,
            address_style
        )
    )

    story.append(
        Spacer(
            1,
            2 * mm
        )
    )

    story.append(
        Paragraph(
            "Generated on "
            + datetime.now().strftime(
                "%d %B %Y at %I:%M %p"
            ),
            address_style
        )
    )

    # =========================================================
    # BUILD PDF
    # =========================================================

    doc.build(story)

    buffer.seek(0)

    # =========================================================
    # DOWNLOAD
    # =========================================================

    return send_file(

        buffer,

        mimetype="application/pdf",

        as_attachment=True,

        download_name=(
            "monthly_attendance_"
            + selected_month
            + ".pdf"
        )
    )

# =========================================================
# NORMALIZE DATABASE TIME
# =========================================================

from datetime import time


def normalize_db_time(value, default=None):
    """
    Convert a PostgreSQL time/datetime/string value
    into a Python datetime.time object.

    Supports:
    - datetime.time
    - datetime.datetime
    - strings such as:
        08:00
        08:00:00
        08:00 AM
        08:00:00 AM
    """

    # -----------------------------------------------------
    # Already a time object
    # -----------------------------------------------------

    if isinstance(value, time):
        return value


    # -----------------------------------------------------
    # Datetime object
    # -----------------------------------------------------

    if isinstance(value, datetime):
        return value.time()


    # -----------------------------------------------------
    # Empty database value
    # -----------------------------------------------------

    if value is None:

        return default


    # -----------------------------------------------------
    # Convert to string
    # -----------------------------------------------------

    value = str(value).strip()


    if not value:

        return default


    # -----------------------------------------------------
    # Try common time formats
    # -----------------------------------------------------

    formats = [
        "%H:%M:%S",
        "%H:%M",
        "%I:%M %p",
        "%I:%M:%S %p",
    ]


    for fmt in formats:

        try:

            return datetime.strptime(
                value,
                fmt
            ).time()

        except ValueError:

            continue


    # -----------------------------------------------------
    # Unable to understand database value
    # -----------------------------------------------------

    app.logger.warning(
        "Unable to normalize database time value: %r",
        value
    )

    return default

@app.route("/admin/fix-applicant-messages")
def fix_applicant_messages():

    # =========================================================
    # ADMIN AUTHENTICATION
    # =========================================================

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # CREATE TABLE IF IT DOES NOT EXIST
            # =================================================

            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS applicant_messages (

                    id BIGSERIAL PRIMARY KEY,

                    application_id BIGINT NOT NULL,

                    message_type VARCHAR(30)
                        NOT NULL DEFAULT 'message',

                    subject VARCHAR(255),

                    message TEXT,

                    interview_date DATE,

                    interview_time TIME,

                    interview_location VARCHAR(500),

                    is_read BOOLEAN
                        NOT NULL DEFAULT FALSE,

                    created_at TIMESTAMP
                        NOT NULL DEFAULT CURRENT_TIMESTAMP,

                    CONSTRAINT fk_applicant_messages_application
                        FOREIGN KEY (application_id)
                        REFERENCES applications(id)
                        ON DELETE CASCADE
                )
                """
            )


            # =================================================
            # ADD INTERVIEW DATE
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS interview_date DATE
                """
            )


            # =================================================
            # ADD INTERVIEW TIME
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS interview_time TIME
                """
            )


            # =================================================
            # ADD INTERVIEW LOCATION
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS interview_location VARCHAR(500)
                """
            )


            # =================================================
            # ENSURE OTHER COLUMNS EXIST
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS message_type
                VARCHAR(30)
                NOT NULL DEFAULT 'message'
                """
            )


            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS subject
                VARCHAR(255)
                """
            )


            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS message
                TEXT
                """
            )


            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS is_read
                BOOLEAN
                NOT NULL DEFAULT FALSE
                """
            )


            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS created_at
                TIMESTAMP
                NOT NULL DEFAULT CURRENT_TIMESTAMP
                """
            )


            # =================================================
            # CREATE INDEXES
            # =================================================

            cur.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_applicant_messages_application

                ON applicant_messages(application_id)
                """
            )


            cur.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_applicant_messages_unread

                ON applicant_messages(
                    application_id,
                    is_read
                )
                """
            )


            # =================================================
            # COMMIT CHANGES
            # =================================================

            conn.commit()


            # =================================================
            # VERIFY COLUMNS
            # =================================================

            cur.execute(
                """
                SELECT
                    column_name

                FROM information_schema.columns

                WHERE table_schema = 'public'

                AND table_name = 'applicant_messages'

                ORDER BY ordinal_position
                """
            )

            columns = cur.fetchall()


            column_names = [
                row["column_name"]
                for row in columns
            ]


        # =====================================================
        # VERIFY REQUIRED COLUMNS
        # =====================================================

        required_columns = [
            "interview_date",
            "interview_time",
            "interview_location"
        ]


        missing_columns = [
            column
            for column in required_columns
            if column not in column_names
        ]


        if missing_columns:

            return (
                "Migration failed. Missing columns: "
                + ", ".join(missing_columns),
                500
            )


        return (
            """
            <h2>Applicant Messages Database Fixed</h2>

            <p>
                The applicant_messages table has been
                successfully updated.
            </p>

            <h3>Available columns:</h3>

            <ul>
                """
                + "".join(
                    f"<li>{column}</li>"
                    for column in column_names
                )
                + """
            </ul>

            <p>
                You can now return to the applicant portal.
            </p>
            """
        )


    except Exception as e:

        conn.rollback()

        app.logger.exception(
            "Error fixing applicant_messages table"
        )

        return (
            f"""
            <h2>Database Migration Failed</h2>

            <pre>{str(e)}</pre>
            """,
            500
        )


    finally:

        conn.close()
# ============================================================
# DAILY ATTENDANCE PDF
# ============================================================

@app.route("/admin/attendance/pdf")
def admin_attendance_pdf():

    # ---------------------------------------------------------
    # ADMIN ACCESS
    # ---------------------------------------------------------

    if not admin_required():
        return redirect(url_for("admin_login"))

    # ---------------------------------------------------------
    # SELECTED DATE
    # ---------------------------------------------------------

    selected_date = (
        request.args.get("date", "").strip()
        or datetime.now().strftime("%Y-%m-%d")
    )

    try:

        attendance_date = datetime.strptime(
            selected_date,
            "%Y-%m-%d"
        ).date()

    except ValueError:

        attendance_date = datetime.now().date()

        selected_date = (
            attendance_date.strftime("%Y-%m-%d")
        )

    # ---------------------------------------------------------
    # DATABASE
    # ---------------------------------------------------------

    conn = get_db()

    attendance_records = []
    absent_staff = []
    settings = None

    try:

        with conn.cursor() as cur:

            # =================================================
            # COMPANY SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT
                    company_name,
                    company_email,
                    company_phone,
                    company_address,
                    footer_text,
                    logo

                FROM company_settings

                ORDER BY id ASC

                LIMIT 1
                """
            )

            settings = cur.fetchone()

            # =================================================
            # ATTENDANCE RECORDS
            # =================================================

            cur.execute(
                """
                SELECT
                    a.id,
                    a.worker_id,
                    a.attendance_date,
                    a.clock_in,
                    a.clock_out,
                    a.total_hours,
                    a.status,
                    a.clock_in_location_verified,
                    a.clock_out_location_verified,

                    ap.application_number,
                    ap.first_name,
                    ap.middle_name,
                    ap.last_name,
                    ap.phone,
                    ap.email,
                    ap.position_applied

                FROM attendance a

                INNER JOIN applications ap
                    ON ap.id = a.worker_id

                WHERE
                    a.attendance_date = %s

                AND ap.status = 'Approved'

                AND ap.portal_active = TRUE

                ORDER BY
                    ap.first_name ASC,
                    ap.last_name ASC
                """,
                (
                    attendance_date,
                )
            )

            attendance_records = cur.fetchall()

            # =================================================
            # ABSENT STAFF
            # =================================================

            cur.execute(
                """
                SELECT
                    ap.id,
                    ap.application_number,
                    ap.first_name,
                    ap.middle_name,
                    ap.last_name,
                    ap.phone,
                    ap.email,
                    ap.position_applied

                FROM applications ap

                WHERE
                    ap.status = 'Approved'

                AND ap.portal_active = TRUE

                AND NOT EXISTS (

                    SELECT 1

                    FROM attendance a

                    WHERE
                        a.worker_id = ap.id

                    AND a.attendance_date = %s
                )

                ORDER BY
                    ap.first_name ASC,
                    ap.last_name ASC
                """,
                (
                    attendance_date,
                )
            )

            absent_staff = cur.fetchall()

    except Exception:

        app.logger.exception(
            "Error generating daily attendance PDF"
        )

        flash(
            "Unable to generate daily attendance PDF.",
            "error"
        )

        return redirect(
            url_for(
                "admin_attendance",
                date=selected_date
            )
        )

    finally:

        conn.close()

    # ---------------------------------------------------------
    # CALCULATE SUMMARY
    # ---------------------------------------------------------

    total_staff = (
        len(attendance_records)
        + len(absent_staff)
    )

    present_count = 0
    clocked_in_count = 0
    completed_count = 0
    total_hours = 0

    for record in attendance_records:

        clock_in = record["clock_in"]
        clock_out = record["clock_out"]

        if clock_in:

            clocked_in_count += 1

            present_count += 1

        if clock_in and clock_out:

            completed_count += 1

        total_hours += float(
            record["total_hours"] or 0
        )

    absent_count = len(
        absent_staff
    )

    # ---------------------------------------------------------
    # REPORTLAB IMPORTS
    # ---------------------------------------------------------

    from io import BytesIO

    from reportlab.lib import colors

    from reportlab.lib.pagesizes import A4

    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle
    )

    from reportlab.lib.enums import TA_CENTER

    from reportlab.lib.units import mm

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image
    )

    # ---------------------------------------------------------
    # PDF BUFFER
    # ---------------------------------------------------------

    buffer = BytesIO()

    # ---------------------------------------------------------
    # DOCUMENT
    # ---------------------------------------------------------

    doc = SimpleDocTemplate(

        buffer,

        pagesize=A4,

        rightMargin=10 * mm,

        leftMargin=10 * mm,

        topMargin=10 * mm,

        bottomMargin=14 * mm

    )

    # ---------------------------------------------------------
    # STYLES
    # ---------------------------------------------------------

    styles = getSampleStyleSheet()

    company_style = ParagraphStyle(

        "DailyCompanyName",

        parent=styles["Heading1"],

        fontSize=16,

        leading=19,

        alignment=TA_CENTER,

        spaceAfter=3

    )

    contact_style = ParagraphStyle(

        "DailyContact",

        parent=styles["Normal"],

        fontSize=7.5,

        leading=10,

        alignment=TA_CENTER

    )

    title_style = ParagraphStyle(

        "DailyTitle",

        parent=styles["Heading2"],

        fontSize=13,

        leading=16,

        alignment=TA_CENTER,

        spaceBefore=7,

        spaceAfter=3

    )

    subtitle_style = ParagraphStyle(

        "DailySubtitle",

        parent=styles["Normal"],

        fontSize=8,

        leading=10,

        alignment=TA_CENTER,

        spaceAfter=8

    )

    small_style = ParagraphStyle(

        "DailySmall",

        parent=styles["Normal"],

        fontSize=6.5,

        leading=8

    )

    # ---------------------------------------------------------
    # STORY
    # ---------------------------------------------------------

    story = []

    # ---------------------------------------------------------
    # LOGO
    # ---------------------------------------------------------

    logo_path = None

    if settings and settings["logo"]:

        possible_logo = os.path.join(

            app.root_path,

            "static",

            "uploads",

            settings["logo"]

        )

        if os.path.exists(
            possible_logo
        ):

            logo_path = possible_logo

    if logo_path:

        try:

            logo = Image(

                logo_path,

                width=20 * mm,

                height=20 * mm

            )

            logo.hAlign = "CENTER"

            story.append(
                logo
            )

            story.append(
                Spacer(
                    1,
                    2 * mm
                )
            )

        except Exception:

            app.logger.warning(
                "Unable to load daily attendance logo.",
                exc_info=True
            )

    # ---------------------------------------------------------
    # COMPANY INFORMATION
    # ---------------------------------------------------------

    company_name = (

        settings["company_name"]

        if settings
        and settings["company_name"]

        else "AV KING VET DRUG VENTURE"

    )

    company_address = (

        settings["company_address"]

        if settings
        and settings["company_address"]

        else ""

    )

    company_phone = (

        settings["company_phone"]

        if settings
        and settings["company_phone"]

        else ""

    )

    company_email = (

        settings["company_email"]

        if settings
        and settings["company_email"]

        else ""

    )

    story.append(
        Paragraph(
            str(company_name),
            company_style
        )
    )

    contact_line = " | ".join(
        filter(
            None,
            [
                str(company_address),
                str(company_phone),
                str(company_email)
            ]
        )
    )

    if contact_line:

        story.append(
            Paragraph(
                contact_line,
                contact_style
            )
        )

    # ---------------------------------------------------------
    # TITLE
    # ---------------------------------------------------------

    story.append(
        Paragraph(
            "DAILY ATTENDANCE REPORT",
            title_style
        )
    )

    story.append(
        Paragraph(
            attendance_date.strftime(
                "%A, %d %B %Y"
            ),
            subtitle_style
        )
    )

    # ---------------------------------------------------------
    # SUMMARY TABLE
    # ---------------------------------------------------------

    summary_data = [

        [
            "TOTAL STAFF",
            "PRESENT",
            "ABSENT",
            "CLOCKED IN",
            "COMPLETED",
            "TOTAL HOURS"
        ],

        [
            str(total_staff),
            str(present_count),
            str(absent_count),
            str(clocked_in_count),
            str(completed_count),
            f"{total_hours:.2f}"
        ]

    ]

    summary_table = Table(

        summary_data,

        colWidths=[

            30 * mm,
            25 * mm,
            25 * mm,
            30 * mm,
            30 * mm,
            35 * mm

        ]

    )

    summary_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1f2937")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.HexColor("#d1d5db")
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            )

        ])
    )

    story.append(
        summary_table
    )

    story.append(
        Spacer(
            1,
            6 * mm
        )
    )

    # ---------------------------------------------------------
    # ATTENDANCE TABLE
    # ---------------------------------------------------------

    table_data = [

        [
            "#",
            "Application No.",
            "Employee",
            "Position",
            "Clock In",
            "Clock Out",
            "Hours",
            "Status"
        ]

    ]

    for index, record in enumerate(
        attendance_records,
        start=1
    ):

        clock_in_text = (

            record["clock_in"].strftime(
                "%I:%M %p"
            )

            if record["clock_in"]

            else "—"

        )

        clock_out_text = (

            record["clock_out"].strftime(
                "%I:%M %p"
            )

            if record["clock_out"]

            else "Still Clocked In"

        )

        employee_name = " ".join(
            filter(
                None,
                [
                    record["first_name"],
                    record["middle_name"],
                    record["last_name"]
                ]
            )
        )

        if record["clock_in"] and record["clock_out"]:

            status_text = "Completed"

        elif record["clock_in"]:

            status_text = "Clocked In"

        else:

            status_text = "Not Recorded"

        table_data.append([

            str(index),

            Paragraph(
                str(
                    record["application_number"]
                ),
                small_style
            ),

            Paragraph(
                employee_name,
                small_style
            ),

            Paragraph(
                str(
                    record["position_applied"]
                    or "—"
                ),
                small_style
            ),

            clock_in_text,

            clock_out_text,

            f'{float(record["total_hours"] or 0):.2f}',

            status_text

        ])

    # ---------------------------------------------------------
    # ABSENT STAFF
    # ---------------------------------------------------------

    for staff in absent_staff:

        employee_name = " ".join(
            filter(
                None,
                [
                    staff["first_name"],
                    staff["middle_name"],
                    staff["last_name"]
                ]
            )
        )

        table_data.append([

            str(
                len(table_data)
            ),

            Paragraph(
                str(
                    staff["application_number"]
                ),
                small_style
            ),

            Paragraph(
                employee_name,
                small_style
            ),

            Paragraph(
                str(
                    staff["position_applied"]
                    or "—"
                ),
                small_style
            ),

            "—",

            "—",

            "0.00",

            "Absent"

        ])

    # ---------------------------------------------------------
    # TABLE
    # ---------------------------------------------------------

    attendance_table = Table(

        table_data,

        repeatRows=1,

        colWidths=[

            8 * mm,
            27 * mm,
            39 * mm,
            29 * mm,
            20 * mm,
            27 * mm,
            17 * mm,
            20 * mm

        ]

    )

    attendance_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#111827")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, 0),
                6.5
            ),

            (
                "FONTSIZE",
                (0, 1),
                (-1, -1),
                6.5
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "ALIGN",
                (0, 0),
                (0, -1),
                "CENTER"
            ),

            (
                "ALIGN",
                (4, 1),
                (-1, -1),
                "CENTER"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.35,
                colors.HexColor("#d1d5db")
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f9fafb")
                ]
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                4
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                4
            )

        ])
    )

    story.append(
        attendance_table
    )

    story.append(
        Spacer(
            1,
            7 * mm
        )
    )

    # ---------------------------------------------------------
    # FOOTER
    # ---------------------------------------------------------

    footer_text = (

        settings["footer_text"]

        if settings
        and settings["footer_text"]

        else "Your needs, Our Priority"

    )

    story.append(
        Paragraph(
            str(footer_text),
            contact_style
        )
    )

    story.append(
        Spacer(
            1,
            2 * mm
        )
    )

    story.append(
        Paragraph(
            "Generated on "
            + datetime.now().strftime(
                "%d %B %Y at %I:%M %p"
            ),
            contact_style
        )
    )

    # ---------------------------------------------------------
    # BUILD
    # ---------------------------------------------------------

    doc.build(
        story
    )

    buffer.seek(0)

    # ---------------------------------------------------------
    # DOWNLOAD
    # ---------------------------------------------------------

    return send_file(

        buffer,

        mimetype="application/pdf",

        as_attachment=True,

        download_name=(
            f"daily_attendance_"
            f"{selected_date}.pdf"
        )

    )

@app.route("/admin/attendance/monthly/export/excel")
def export_monthly_attendance_excel():

    # =========================================================
    # ADMIN ACCESS
    # =========================================================

    if not admin_required():
        return redirect(url_for("admin_login"))

    # =========================================================
    # FILTERS
    # =========================================================

    selected_month = (
        request.args.get("month", "").strip()
        or datetime.now().strftime("%Y-%m")
    )

    search = (
        request.args.get("search", "").strip()
    )

    # =========================================================
    # VALIDATE MONTH
    # =========================================================

    try:

        month_date = datetime.strptime(
            selected_month,
            "%Y-%m"
        )

    except ValueError:

        month_date = datetime.now()

        selected_month = month_date.strftime(
            "%Y-%m"
        )

    # =========================================================
    # MONTH RANGE
    # =========================================================

    month_start = month_date.replace(
        day=1
    )

    if month_date.month == 12:

        next_month = datetime(
            month_date.year + 1,
            1,
            1
        )

    else:

        next_month = datetime(
            month_date.year,
            month_date.month + 1,
            1
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET COMPANY INFORMATION
            # =================================================

            cur.execute(
                """
                SELECT
                    company_name,
                    company_email,
                    company_phone,
                    company_address

                FROM company_settings

                ORDER BY id ASC

                LIMIT 1
                """
            )

            company = cur.fetchone()

            company_name = (
                company["company_name"]
                if company
                else "AV KING VET DRUG VENTURE"
            )

            company_email = (
                company["company_email"]
                if company
                else ""
            )

            company_phone = (
                company["company_phone"]
                if company
                else ""
            )

            company_address = (
                company["company_address"]
                if company
                else ""
            )

            # =================================================
            # GET STAFF
            # =================================================

            worker_query = """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    phone,
                    email,
                    position_applied

                FROM applications

                WHERE
                    status = 'Approved'
                    AND portal_active = TRUE
            """

            params = []

            if search:

                worker_query += """
                    AND (
                        application_number ILIKE %s
                        OR first_name ILIKE %s
                        OR middle_name ILIKE %s
                        OR last_name ILIKE %s
                        OR phone ILIKE %s
                        OR email ILIKE %s
                        OR position_applied ILIKE %s
                    )
                """

                search_value = f"%{search}%"

                params = [
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value,
                    search_value
                ]

            worker_query += """
                ORDER BY
                    first_name ASC,
                    last_name ASC
            """

            cur.execute(
                worker_query,
                params
            )

            workers = cur.fetchall()

            # =================================================
            # GET ATTENDANCE
            # =================================================

            cur.execute(
                """
                SELECT
                    worker_id,
                    attendance_date,
                    clock_in,
                    clock_out,
                    total_hours,
                    status

                FROM attendance

                WHERE
                    attendance_date >= %s
                    AND attendance_date < %s

                ORDER BY attendance_date ASC
                """,
                (
                    month_start.strftime("%Y-%m-%d"),
                    next_month.strftime("%Y-%m-%d")
                )
            )

            attendance_rows = cur.fetchall()

    finally:

        conn.close()

    # =========================================================
    # ORGANIZE ATTENDANCE
    # =========================================================

    attendance_map = {}

    for row in attendance_rows:

        worker_id = row["worker_id"]

        attendance_map.setdefault(
            worker_id,
            []
        ).append(row)

    # =========================================================
    # CALCULATE WORKING DAYS
    # =========================================================

    working_days = 0

    current_day = month_start

    while current_day < next_month:

        if current_day.weekday() < 5:
            working_days += 1

        current_day += timedelta(days=1)

    # =========================================================
    # BUILD EXCEL DATA
    # =========================================================

    excel_rows = []

    for worker in workers:

        records = attendance_map.get(
            worker["id"],
            []
        )

        present_days = 0
        absent_days = 0
        late_days = 0
        incomplete_days = 0
        total_hours = 0

        for record in records:

            clock_in = record["clock_in"]
            clock_out = record["clock_out"]

            status = (
                str(record["status"] or "")
                .strip()
                .lower()
            )

            hours = float(
                record["total_hours"] or 0
            )

            total_hours += hours

            if (
                clock_in
                and not clock_out
            ):

                incomplete_days += 1

            elif status == "late":

                late_days += 1
                present_days += 1

            elif (
                clock_in
                and clock_out
            ):

                present_days += 1

            elif status == "absent":

                absent_days += 1

        # =====================================================
        # MISSING WORKING DAYS = ABSENT
        # =====================================================

        absent_days += max(
            working_days
            - present_days
            - incomplete_days
            - absent_days,
            0
        )

        # =====================================================
        # ATTENDANCE %
        # =====================================================

        if working_days:

            attendance_percentage = (
                (
                    present_days
                    + incomplete_days
                )
                / working_days
            ) * 100

        else:

            attendance_percentage = 0

        # =====================================================
        # EMPLOYEE NAME
        # =====================================================

        full_name = " ".join(
            part
            for part in [
                worker["first_name"],
                worker["middle_name"],
                worker["last_name"]
            ]
            if part
        )

        excel_rows.append({

            "Employee Name":
                full_name,

            "Application Number":
                worker["application_number"],

            "Position":
                worker["position_applied"] or "",

            "Phone":
                worker["phone"] or "",

            "Email":
                worker["email"] or "",

            "Working Days":
                working_days,

            "Present Days":
                present_days,

            "Absent Days":
                absent_days,

            "Late Days":
                late_days,

            "Incomplete Days":
                incomplete_days,

            "Total Hours":
                round(
                    total_hours,
                    2
                ),

            "Attendance %":
                round(
                    attendance_percentage,
                    1
                )
        })

    # =========================================================
    # DATAFRAME
    # =========================================================

    df = pd.DataFrame(
        excel_rows
    )

    # =========================================================
    # CREATE EXCEL FILE
    # =========================================================

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        # ---------------------------------------------
        # COMPANY HEADER
        # ---------------------------------------------

        header_data = pd.DataFrame({

            "A": [
                company_name,
                company_address,
                company_phone,
                company_email,
                "",
                f"MONTHLY ATTENDANCE REPORT — {selected_month}"
            ]

        })

        header_data.to_excel(
            writer,
            index=False,
            header=False,
            sheet_name="Attendance Report"
        )

        # ---------------------------------------------
        # ATTENDANCE TABLE
        # ---------------------------------------------

        start_row = len(header_data) + 2

        df.to_excel(
            writer,
            index=False,
            sheet_name="Attendance Report",
            startrow=start_row
        )

        worksheet = writer.book[
            "Attendance Report"
        ]

        # ---------------------------------------------
        # COLUMN WIDTHS
        # ---------------------------------------------

        widths = {

            "A": 28,
            "B": 20,
            "C": 25,
            "D": 18,
            "E": 32,
            "F": 16,
            "G": 16,
            "H": 16,
            "I": 14,
            "J": 18,
            "K": 16,
            "L": 16

        }

        for column, width in widths.items():

            worksheet.column_dimensions[
                column
            ].width = width

        # ---------------------------------------------
        # FREEZE TABLE HEADER
        # ---------------------------------------------

        worksheet.freeze_panes = (
            f"A{start_row + 2}"
        )

        # ---------------------------------------------
        # AUTO FILTER
        # ---------------------------------------------

        if len(df) > 0:

            first_row = start_row + 1

            last_row = (
                start_row
                + len(df)
                + 1
            )

            worksheet.auto_filter.ref = (
                f"A{first_row}:L{last_row}"
            )

    output.seek(0)

    # =========================================================
    # DOWNLOAD
    # =========================================================

    filename = (
        f"monthly_attendance_"
        f"{selected_month}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

# =========================================================
# ATTENDANCE STATUS PROCESSOR
# =========================================================

def process_attendance_status(
    attendance_date=None
):
    """
    Automatically creates Absent records for approved active
    applicants who did not clock in on a working day.

    Also updates existing attendance records to Present/Late/
    Incomplete according to the company attendance settings.
    """

    if attendance_date is None:
        attendance_date = datetime.now().date()

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET COMPANY ATTENDANCE SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT
                    attendance_enabled,
                    clock_in_end,
                    clock_out_start,
                    clock_out_end,
                    late_after_minutes,
                    early_before_minutes

                FROM company_settings

                ORDER BY id ASC

                LIMIT 1
                """
            )

            settings = cur.fetchone()

            if not settings:
                return

            if not settings["attendance_enabled"]:
                return

            # =================================================
            # ONLY PROCESS WORKING DAYS
            # MONDAY = 0
            # SUNDAY = 6
            # =================================================

            if attendance_date.weekday() >= 5:
                return

            # =================================================
            # GET APPROVED ACTIVE APPLICANTS
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    last_name

                FROM applications

                WHERE status = 'Approved'
                AND portal_active = TRUE
                """
            )

            workers = cur.fetchall()

            # =================================================
            # PROCESS EACH WORKER
            # =================================================

            for worker in workers:

                worker_id = worker["id"]

                # -------------------------------------------------
                # GET EXISTING ATTENDANCE
                # -------------------------------------------------

                cur.execute(
                    """
                    SELECT
                        id,
                        clock_in,
                        clock_out,
                        status,
                        total_hours

                    FROM attendance

                    WHERE worker_id = %s
                    AND attendance_date = %s

                    LIMIT 1
                    """,
                    (
                        worker_id,
                        attendance_date
                    )
                )

                attendance = cur.fetchone()

                # =================================================
                # NO ATTENDANCE = ABSENT
                # =================================================

                if not attendance:

                    # Only create Absent records for dates that
                    # have already passed.

                    today = datetime.now().date()

                    if attendance_date < today:

                        cur.execute(
                            """
                            INSERT INTO attendance
                            (
                                worker_id,
                                attendance_date,
                                status,
                                total_hours,
                                clock_in_location_verified,
                                clock_out_location_verified,
                                created_at,
                                updated_at
                            )

                            VALUES
                            (
                                %s,
                                %s,
                                'Absent',
                                0,
                                FALSE,
                                FALSE,
                                CURRENT_TIMESTAMP,
                                CURRENT_TIMESTAMP
                            )

                            ON CONFLICT
                            (worker_id, attendance_date)
                            DO NOTHING
                            """,
                            (
                                worker_id,
                                attendance_date
                            )
                        )

                    continue

                # =================================================
                # EXISTING ATTENDANCE
                # =================================================

                clock_in = attendance["clock_in"]
                clock_out = attendance["clock_out"]

                # -------------------------------------------------
                # NO CLOCK-IN
                # -------------------------------------------------

                if not clock_in:

                    cur.execute(
                        """
                        UPDATE attendance

                        SET
                            status = 'Absent',
                            total_hours = 0,
                            updated_at = CURRENT_TIMESTAMP

                        WHERE id = %s
                        """,
                        (
                            attendance["id"],
                        )
                    )

                    continue

                # =================================================
                # CALCULATE WORKING HOURS
                # =================================================

                total_hours = 0

                if clock_in and clock_out:

                    duration = (
                        clock_out - clock_in
                    )

                    total_seconds = (
                        duration.total_seconds()
                    )

                    if total_seconds > 0:

                        total_hours = round(
                            total_seconds / 3600,
                            2
                        )

                # =================================================
                # DETERMINE LATE STATUS
                # =================================================

                record_status = "Present"

                if settings["clock_in_end"]:

                    clock_in_end = settings[
                        "clock_in_end"
                    ]

                    late_after_minutes = int(
                        settings[
                            "late_after_minutes"
                        ] or 0
                    )

                    allowed_time = (
                        datetime.combine(
                            attendance_date,
                            clock_in_end
                        )
                        + timedelta(
                            minutes=late_after_minutes
                        )
                    )

                    if clock_in > allowed_time:

                        record_status = "Late"

                # =================================================
                # INCOMPLETE
                # =================================================

                if clock_in and not clock_out:

                    record_status = "Incomplete"

                # =================================================
                # UPDATE RECORD
                # =================================================

                cur.execute(
                    """
                    UPDATE attendance

                    SET
                        status = %s,
                        total_hours = %s,
                        updated_at = CURRENT_TIMESTAMP

                    WHERE id = %s
                    """,
                    (
                        record_status,
                        total_hours,
                        attendance["id"]
                    )
                )

        conn.commit()

    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error processing attendance status"
        )

        raise

    finally:

        conn.close()
# ============================================================
# MARK WHATSAPP NOTIFICATION AS SENT
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/whatsapp-sent",
    methods=["POST"]
)
def mark_whatsapp_notification_sent(application_id):

    if not admin_required():
        return redirect(url_for("admin_login"))

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    id,
                    status
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            application = cur.fetchone()

            if not application:

                flash(
                    "Application not found.",
                    "error"
                )

                return redirect(
                    url_for("admin_applications")
                )


            if application["status"] != "Shortlisted":

                flash(
                    "Only shortlisted applicants can have "
                    "a shortlist notification recorded.",
                    "error"
                )

                return redirect(
                    url_for(
                        "admin_application_details",
                        application_id=application_id
                    )
                )


            cur.execute(
                """
                UPDATE applications

                SET
                    notification_sent = TRUE,

                    notification_sent_at =
                        CURRENT_TIMESTAMP,

                    updated_at =
                        CURRENT_TIMESTAMP

                WHERE id = %s
                """,
                (application_id,)
            )


        conn.commit()


    except Exception:

        conn.rollback()
        raise


    finally:

        conn.close()


    flash(
        "WhatsApp notification marked as sent.",
        "success"
    )


    return redirect(
        url_for(
            "admin_application_details",
            application_id=application_id
        )
    )
# ============================================================
# APPLICATION SUCCESS
# ============================================================

@app.route(
    "/application-success/<application_number>"
)
def application_success(
    application_number
):

    return render_template(
        "application_success.html",
        application_number=application_number
    )

# ============================================================
# APPLICANT AUTHENTICATION
# ============================================================

def applicant_required():

    return bool(
        session.get("applicant_id")
    )

# ============================================================
# APPLICANT PORTAL LOGIN
# ============================================================

# =========================================================
# APPLICANT LOGIN
# =========================================================
@app.route("/applicant/login", methods=["GET", "POST"])
def applicant_login():

    # =========================================================
    # GET COMPANY LOGO
    # =========================================================

    def get_company_logo():

        conn = get_db()

        try:

            with conn.cursor() as cur:

                cur.execute(
                    """
                    SELECT logo
                    FROM company_settings
                    ORDER BY id ASC
                    LIMIT 1
                    """
                )

                row = cur.fetchone()

                if row and row["logo"]:
                    return str(row["logo"]).strip()

                return None

        except Exception:

            app.logger.exception(
                "Unable to load company logo."
            )

            return None

        finally:

            conn.close()

    # =========================================================
    # GET REQUEST
    # =========================================================

    if request.method == "GET":

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    # =========================================================
    # FORM DATA
    # =========================================================

    application_number = (
        request.form.get(
            "application_number",
            ""
        )
        .strip()
        .upper()
    )

    password = request.form.get(
        "password",
        ""
    )

    # =========================================================
    # VALIDATE INPUT
    # =========================================================

    if not application_number or not password:

        flash(
            "Please enter your application number and password.",
            "error"
        )

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    # =========================================================
    # FIND APPLICANT
    # =========================================================

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    last_name,
                    password_hash,
                    portal_active,
                    status
                FROM applications
                WHERE UPPER(application_number) = %s
                LIMIT 1
                """,
                (application_number,)
            )

            applicant = cur.fetchone()

    except Exception:

        app.logger.exception(
            "Error loading applicant during login."
        )

        flash(
            "Unable to process your login right now. Please try again.",
            "error"
        )

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    finally:

        conn.close()

    # =========================================================
    # APPLICANT NOT FOUND
    # =========================================================

    if not applicant:

        flash(
            "Invalid application number or password.",
            "error"
        )

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    # =========================================================
    # PORTAL DISABLED
    # =========================================================

    if not applicant["portal_active"]:

        flash(
            "Your applicant portal has been disabled.",
            "error"
        )

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    # =========================================================
    # PASSWORD NOT ACTIVATED
    # =========================================================

    if not applicant["password_hash"]:

        flash(
            "Your applicant portal account has not been activated yet.",
            "error"
        )

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    # =========================================================
    # CHECK PASSWORD
    # =========================================================

    try:

        password_valid = check_password_hash(
            applicant["password_hash"],
            password
        )

    except Exception:

        app.logger.exception(
            "Applicant password verification failed."
        )

        password_valid = False

    if not password_valid:

        flash(
            "Invalid application number or password.",
            "error"
        )

        return render_template(
            "applicant_login.html",
            company_logo=get_company_logo()
        )

    # =========================================================
    # LOGIN SUCCESS
    # =========================================================

    session.clear()

    session["applicant_logged_in"] = True

    session["applicant_id"] = applicant["id"]

    session["applicant_application_number"] = (
        applicant["application_number"]
    )

    session["applicant_name"] = (
        f"{applicant['first_name']} "
        f"{applicant['last_name']}"
    ).strip()

    # =========================================================
    # UPDATE LAST LOGIN
    # =========================================================

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                UPDATE applications
                SET
                    last_login = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (applicant["id"],)
            )

        conn.commit()

    except Exception:

        conn.rollback()

        app.logger.exception(
            "Unable to update applicant last login."
        )

    finally:

        conn.close()

    # =========================================================
    # REDIRECT TO APPLICANT PORTAL
    # =========================================================

    return redirect(
        url_for("applicant_portal")
    )



# ============================================================
# APPLICANT LOGOUT
# ============================================================

@app.route("/applicant/logout")
def applicant_logout():

    session.pop(
        "applicant_id",
        None
    )

    session.pop(
        "applicant_application_number",
        None
    )

    session.pop(
        "applicant_name",
        None
    )

    flash(
        "You have been logged out.",
        "success"
    )

    return redirect(
        url_for("applicant_login")
    )

# ============================================================
# ADMIN - APPLICANT COMMUNICATION
# ============================================================

@app.route(
    "/admin/applicant-communication",
    methods=["GET", "POST"]
)
def admin_applicant_communication():

    # --------------------------------------------------------
    # ADMIN AUTHENTICATION
    # --------------------------------------------------------

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # SEND MESSAGE / INTERVIEW
            # =================================================

            if request.method == "POST":

                application_id = (
                    request.form.get(
                        "application_id",
                        ""
                    ).strip()
                )

                message_type = (
                    request.form.get(
                        "message_type",
                        "message"
                    ).strip().lower()
                )

                subject = (
                    request.form.get(
                        "subject",
                        ""
                    ).strip()
                )

                message = (
                    request.form.get(
                        "message",
                        ""
                    ).strip()
                )

                interview_date = (
                    request.form.get(
                        "interview_date",
                        ""
                    ).strip()
                )

                interview_time = (
                    request.form.get(
                        "interview_time",
                        ""
                    ).strip()
                )

                interview_location = (
                    request.form.get(
                        "interview_location",
                        ""
                    ).strip()
                )


                # =================================================
                # VALIDATE APPLICATION
                # =================================================

                try:

                    application_id = int(
                        application_id
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    flash(
                        "Please select a valid applicant.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "admin_applicant_communication"
                        )
                    )


                # =================================================
                # VALIDATE MESSAGE TYPE
                # =================================================

                if message_type not in [
                    "message",
                    "interview"
                ]:

                    message_type = "message"


                # =================================================
                # CHECK APPLICANT
                # =================================================

                cur.execute(
                    """
                    SELECT
                        id,
                        application_number,
                        first_name,
                        middle_name,
                        last_name,
                        email,
                        phone
                    FROM applications
                    WHERE id = %s
                    LIMIT 1
                    """,
                    (
                        application_id,
                    )
                )

                applicant = cur.fetchone()


                if not applicant:

                    flash(
                        "Applicant not found.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "admin_applicant_communication"
                        )
                    )


                # =================================================
                # MESSAGE VALIDATION
                # =================================================

                if not message:

                    flash(
                        "Please enter a message.",
                        "error"
                    )

                    return redirect(
                        url_for(
                            "admin_applicant_communication"
                        )
                    )


                # =================================================
                # INTERVIEW VALIDATION
                # =================================================

                if message_type == "interview":

                    if not interview_date:

                        flash(
                            "Please select the interview date.",
                            "error"
                        )

                        return redirect(
                            url_for(
                                "admin_applicant_communication"
                            )
                        )


                    if not interview_time:

                        flash(
                            "Please select the interview time.",
                            "error"
                        )

                        return redirect(
                            url_for(
                                "admin_applicant_communication"
                            )
                        )


                    if not interview_location:

                        flash(
                            "Please enter the interview location.",
                            "error"
                        )

                        return redirect(
                            url_for(
                                "admin_applicant_communication"
                            )
                        )


                    # ---------------------------------------------
                    # VALIDATE DATE
                    # ---------------------------------------------

                    try:

                        datetime.strptime(
                            interview_date,
                            "%Y-%m-%d"
                        )

                    except ValueError:

                        flash(
                            "Invalid interview date.",
                            "error"
                        )

                        return redirect(
                            url_for(
                                "admin_applicant_communication"
                            )
                        )


                    # ---------------------------------------------
                    # VALIDATE TIME
                    # ---------------------------------------------

                    try:

                        datetime.strptime(
                            interview_time,
                            "%H:%M"
                        )

                    except ValueError:

                        flash(
                            "Invalid interview time.",
                            "error"
                        )

                        return redirect(
                            url_for(
                                "admin_applicant_communication"
                            )
                        )


                    if not subject:

                        subject = (
                            "Interview Invitation"
                        )


                else:

                    # ------------------------------------------------
                    # NORMAL MESSAGE
                    # ------------------------------------------------

                    if not subject:

                        subject = (
                            "Application Update"
                        )

                    interview_date = None
                    interview_time = None
                    interview_location = None


                # =================================================
                # INSERT COMMUNICATION
                # =================================================

                cur.execute(
                    """
                    INSERT INTO applicant_messages
                    (
                        application_id,
                        message_type,
                        subject,
                        message,
                        interview_date,
                        interview_time,
                        interview_location,
                        is_read,
                        created_at
                    )
                    VALUES
                    (
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        FALSE,
                        CURRENT_TIMESTAMP
                    )
                    """,
                    (
                        application_id,
                        message_type,
                        subject,
                        message,
                        interview_date or None,
                        interview_time or None,
                        interview_location or None
                    )
                )


                conn.commit()


                # =================================================
                # SUCCESS
                # =================================================

                flash(
                    f"Communication sent successfully to "
                    f"{applicant['first_name']} "
                    f"{applicant['last_name']}.",
                    "success"
                )

                return redirect(
                    url_for(
                        "admin_applicant_communication"
                    )
                )


            # =================================================
            # GET APPLICANTS
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    email,
                    phone,
                    position_applied,
                    application_status
                FROM applications
                ORDER BY created_at DESC, id DESC
                """
            )

            applicants = cur.fetchall()


            # =================================================
            # RECENT COMMUNICATIONS
            # =================================================

            cur.execute(
                """
                SELECT
                    am.id,
                    am.application_id,
                    am.message_type,
                    am.subject,
                    am.message,
                    am.interview_date,
                    am.interview_time,
                    am.interview_location,
                    am.is_read,
                    am.created_at,

                    a.application_number,
                    a.first_name,
                    a.middle_name,
                    a.last_name

                FROM applicant_messages am

                INNER JOIN applications a
                    ON a.id = am.application_id

                ORDER BY
                    am.created_at DESC,
                    am.id DESC

                LIMIT 100
                """
            )

            communications = cur.fetchall()


    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error managing applicant communication"
        )

        flash(
            "Unable to process applicant communication.",
            "error"
        )

        applicants = []
        communications = []


    finally:

        conn.close()


    return render_template(
        "admin_applicant_communication.html",
        applicants=applicants,
        communications=communications
    )
# ============================================================
# APPLICANT PORTAL
# ============================================================
@app.route("/applicant/portal")
def applicant_portal():

    # =========================================================
    # CHECK APPLICANT LOGIN
    # =========================================================

    applicant_id = session.get("applicant_id")

    if not applicant_id:

        return redirect(
            url_for("applicant_login")
        )


    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # ENSURE APPLICANT MESSAGES TABLE EXISTS
            # =================================================

            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS applicant_messages (

                    id BIGSERIAL PRIMARY KEY,

                    application_id BIGINT NOT NULL,

                    message_type VARCHAR(30)
                        NOT NULL DEFAULT 'message',

                    subject VARCHAR(255),

                    message TEXT,

                    interview_date DATE,

                    interview_time TIME,

                    interview_location VARCHAR(500),

                    is_read BOOLEAN
                        NOT NULL DEFAULT FALSE,

                    created_at TIMESTAMP
                        NOT NULL DEFAULT CURRENT_TIMESTAMP,

                    CONSTRAINT fk_applicant_messages_application
                        FOREIGN KEY (application_id)
                        REFERENCES applications(id)
                        ON DELETE CASCADE
                )
                """
            )


            # =================================================
            # IMPORTANT DATABASE MIGRATION
            #
            # This updates an EXISTING applicant_messages table.
            #
            # CREATE TABLE IF NOT EXISTS does not add columns
            # to a table that already exists.
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS interview_date DATE
                """
            )


            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS interview_time TIME
                """
            )


            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS interview_location
                VARCHAR(500)
                """
            )


            # =================================================
            # ENSURE MESSAGE TYPE EXISTS
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS message_type
                VARCHAR(30)
                NOT NULL DEFAULT 'message'
                """
            )


            # =================================================
            # ENSURE SUBJECT EXISTS
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS subject
                VARCHAR(255)
                """
            )


            # =================================================
            # ENSURE MESSAGE EXISTS
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS message
                TEXT
                """
            )


            # =================================================
            # ENSURE READ STATUS EXISTS
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS is_read
                BOOLEAN
                NOT NULL DEFAULT FALSE
                """
            )


            # =================================================
            # ENSURE CREATED AT EXISTS
            # =================================================

            cur.execute(
                """
                ALTER TABLE applicant_messages
                ADD COLUMN IF NOT EXISTS created_at
                TIMESTAMP
                NOT NULL DEFAULT CURRENT_TIMESTAMP
                """
            )


            # =================================================
            # CREATE INDEXES
            # =================================================

            cur.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_applicant_messages_application

                ON applicant_messages(application_id)
                """
            )


            cur.execute(
                """
                CREATE INDEX IF NOT EXISTS
                idx_applicant_messages_unread

                ON applicant_messages(
                    application_id,
                    is_read
                )
                """
            )


            # =================================================
            # COMMIT DATABASE MIGRATION
            # =================================================

            conn.commit()


            # =================================================
            # GET APPLICATION
            # =================================================

            cur.execute(
                """
                SELECT *
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (
                    applicant_id,
                )
            )

            application = cur.fetchone()


            # =================================================
            # APPLICATION NOT FOUND
            # =================================================

            if not application:

                session.pop(
                    "applicant_id",
                    None
                )

                session.pop(
                    "applicant_application_id",
                    None
                )

                session.pop(
                    "applicant_unread_count",
                    None
                )

                flash(
                    "Application account not found. "
                    "Please log in again.",
                    "error"
                )

                return redirect(
                    url_for("applicant_login")
                )


            # =================================================
            # GET APPLICANT COMMUNICATIONS
            #
            # NORMAL MESSAGE
            # INTERVIEW INVITATION
            # INTERVIEW DETAILS
            # =================================================

            cur.execute(
                """
                SELECT

                    id,

                    subject,

                    message,

                    message_type,

                    interview_date,

                    interview_time,

                    interview_location,

                    is_read,

                    created_at

                FROM applicant_messages

                WHERE application_id = %s

                ORDER BY
                    created_at DESC,
                    id DESC
                """,
                (
                    applicant_id,
                )
            )

            messages = cur.fetchall()


            # =================================================
            # COUNT UNREAD COMMUNICATIONS
            # =================================================

            cur.execute(
                """
                SELECT
                    COUNT(*) AS unread_count

                FROM applicant_messages

                WHERE application_id = %s

                AND is_read = FALSE
                """,
                (
                    applicant_id,
                )
            )

            unread_row = cur.fetchone()


            if unread_row:

                unread_count = int(
                    unread_row["unread_count"]
                    or 0
                )

            else:

                unread_count = 0


            # =================================================
            # TOTAL COMMUNICATIONS
            # =================================================

            cur.execute(
                """
                SELECT
                    COUNT(*) AS total_messages

                FROM applicant_messages

                WHERE application_id = %s
                """,
                (
                    applicant_id,
                )
            )

            total_row = cur.fetchone()


            if total_row:

                total_messages = int(
                    total_row["total_messages"]
                    or 0
                )

            else:

                total_messages = 0


            # =================================================
            # TOTAL INTERVIEW INVITATIONS
            # =================================================

            cur.execute(
                """
                SELECT
                    COUNT(*) AS interview_count

                FROM applicant_messages

                WHERE application_id = %s

                AND message_type = 'interview'
                """,
                (
                    applicant_id,
                )
            )

            interview_row = cur.fetchone()


            if interview_row:

                interview_count = int(
                    interview_row["interview_count"]
                    or 0
                )

            else:

                interview_count = 0


            # =================================================
            # UPDATE SIDEBAR UNREAD BADGE
            # =================================================

            session[
                "applicant_unread_count"
            ] = unread_count


    except Exception:

        # =====================================================
        # ROLLBACK
        # =====================================================

        try:

            conn.rollback()

        except Exception:

            pass


        # =====================================================
        # LOG COMPLETE ERROR
        # =====================================================

        app.logger.exception(
            "Error loading applicant portal."
        )


        # =====================================================
        # CLEAR POSSIBLE BAD SESSION DATA
        # =====================================================

        session.pop(
            "applicant_unread_count",
            None
        )


        flash(
            "We could not load your applicant portal. "
            "Please try again.",
            "error"
        )


        return redirect(
            url_for("applicant_login")
        )


    finally:

        # =====================================================
        # CLOSE CONNECTION
        # =====================================================

        conn.close()


    # =========================================================
    # RENDER APPLICANT PORTAL
    # =========================================================

    return render_template(
        "applicant_portal.html",

        application=application,

        messages=messages,

        unread_count=unread_count,

        total_messages=total_messages,

        interview_count=interview_count
    )

# ============================================================
# APPLICANT MESSAGES
# ============================================================
@app.route("/applicant/messages")
def applicant_messages():

    application_id = session.get(
        "application_id"
    )

    if not application_id:

        return redirect(
            url_for("applicant_login")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET ALL MESSAGES
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    message_type,
                    subject,
                    message,
                    interview_date,
                    interview_time,
                    interview_location,
                    is_read,
                    created_at
                FROM applicant_messages
                WHERE application_id = %s
                ORDER BY
                    created_at DESC,
                    id DESC
                """,
                (
                    application_id,
                )
            )

            messages = cur.fetchall()


            # =================================================
            # UNREAD COUNT
            # =================================================

            cur.execute(
                """
                SELECT COUNT(*) AS unread_count
                FROM applicant_messages
                WHERE application_id = %s
                AND is_read = FALSE
                """,
                (
                    application_id,
                )
            )

            unread_result = cur.fetchone()

            unread_count = (
                unread_result["unread_count"]
                if unread_result
                else 0
            )


    except Exception:

        app.logger.exception(
            "Error loading applicant messages"
        )

        messages = []
        unread_count = 0


    finally:

        conn.close()


    return render_template(
        "applicant_messages.html",
        messages=messages,
        unread_count=unread_count
    )



@app.route(
    "/applicant/messages/<int:message_id>/read",
    methods=["POST"]
)
def applicant_mark_message_read(message_id):

    application_id = session.get(
        "application_id"
    )

    if not application_id:

        return redirect(
            url_for("applicant_login")
        )

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                UPDATE applicant_messages

                SET
                    is_read = TRUE

                WHERE
                    id = %s
                    AND application_id = %s
                """,
                (
                    message_id,
                    application_id
                )
            )

        conn.commit()

    except Exception:

        conn.rollback()

        app.logger.exception(
            "Error marking applicant message as read"
        )

    finally:

        conn.close()


    return redirect(
        url_for(
            "applicant_messages"
        )
    )
# ============================================================
# VIEW SINGLE APPLICANT MESSAGE
# ============================================================

@app.route("/applicant/messages/<int:message_id>")
def applicant_message(message_id):

    # --------------------------------------------------------
    # CHECK APPLICANT LOGIN
    # --------------------------------------------------------

    if not session.get("applicant_logged_in"):
        return redirect(url_for("applicant_login"))

    # --------------------------------------------------------
    # GET APPLICATION ID
    # --------------------------------------------------------

    application_id = session.get("applicant_application_id")

    # --------------------------------------------------------
    # RECOVER APPLICATION ID IF NECESSARY
    # --------------------------------------------------------

    if not application_id:

        applicant_email = (
            session.get("applicant_email")
            or session.get("email")
        )

        applicant_number = (
            session.get("applicant_number")
            or session.get("application_number")
        )

        conn = get_db()

        try:

            with conn.cursor() as cur:

                # --------------------------------------------
                # FIND BY APPLICATION NUMBER
                # --------------------------------------------

                if applicant_number:

                    cur.execute(
                        """
                        SELECT id
                        FROM applications
                        WHERE application_number = %s
                        LIMIT 1
                        """,
                        (applicant_number,)
                    )

                    row = cur.fetchone()

                    if row:
                        application_id = row["id"]

                # --------------------------------------------
                # FIND BY EMAIL
                # --------------------------------------------

                if not application_id and applicant_email:

                    cur.execute(
                        """
                        SELECT id
                        FROM applications
                        WHERE LOWER(email) = LOWER(%s)
                        ORDER BY id DESC
                        LIMIT 1
                        """,
                        (applicant_email,)
                    )

                    row = cur.fetchone()

                    if row:
                        application_id = row["id"]

        finally:

            conn.close()

    # --------------------------------------------------------
    # APPLICATION ID STILL MISSING
    # --------------------------------------------------------

    if not application_id:

        flash(
            "Your applicant session could not be verified. Please login again.",
            "error"
        )

        session.pop("applicant_application_id", None)

        return redirect(
            url_for("applicant_login")
        )

    # --------------------------------------------------------
    # SAVE APPLICATION ID
    # --------------------------------------------------------

    session["applicant_application_id"] = application_id

    # --------------------------------------------------------
    # GET THE MESSAGE
    # --------------------------------------------------------

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    id,
                    application_id,
                    subject,
                    message,
                    message_type,
                    is_read,
                    created_at
                FROM applicant_messages
                WHERE id = %s
                  AND application_id = %s
                LIMIT 1
                """,
                (
                    message_id,
                    application_id
                )
            )

            message = cur.fetchone()

            # ------------------------------------------------
            # MESSAGE DOES NOT EXIST
            # ------------------------------------------------

            if not message:

                flash(
                    "Message not found.",
                    "error"
                )

                return redirect(
                    url_for("applicant_messages")
                )

            # ------------------------------------------------
            # MARK MESSAGE AS READ
            # ------------------------------------------------

            cur.execute(
                """
                UPDATE applicant_messages
                SET is_read = TRUE
                WHERE id = %s
                  AND application_id = %s
                """,
                (
                    message_id,
                    application_id
                )
            )

        conn.commit()

    except Exception:

        conn.rollback()
        raise

    finally:

        conn.close()

    # --------------------------------------------------------
    # GET UPDATED UNREAD COUNT
    # --------------------------------------------------------

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT COUNT(*) AS unread_count
                FROM applicant_messages
                WHERE application_id = %s
                  AND is_read = FALSE
                """,
                (application_id,)
            )

            row = cur.fetchone()

            unread_count = int(
                row["unread_count"]
            ) if row else 0

    finally:

        conn.close()

    # --------------------------------------------------------
    # UPDATE SESSION BADGE
    # --------------------------------------------------------

    session["applicant_unread_count"] = unread_count

    # --------------------------------------------------------
    # RENDER MESSAGE
    # --------------------------------------------------------

    return render_template(
        "applicant_message.html",
        message=message,
        unread_count=unread_count
    )

# ============================================================
# ADMIN SETTINGS
# ============================================================
@app.route(
    "/admin/settings",
    methods=["GET", "POST"]
)
def admin_settings():

    # =========================================================
    # ADMIN AUTHENTICATION
    # =========================================================

    if not admin_required():

        return redirect(
            url_for("admin_login")
        )


    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # GET EXISTING SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT *
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            existing_settings = cur.fetchone()


            # =================================================
            # POST
            # =================================================

            if request.method == "POST":

                # =================================================
                # COMPANY INFORMATION
                # =================================================

                company_name = (
                    request.form.get(
                        "company_name",
                        ""
                    ).strip()
                )

                company_email = (
                    request.form.get(
                        "company_email",
                        ""
                    ).strip()
                )

                company_phone = (
                    request.form.get(
                        "company_phone",
                        ""
                    ).strip()
                )

                company_address = (
                    request.form.get(
                        "company_address",
                        ""
                    ).strip()
                )

                company_website = (
                    request.form.get(
                        "company_website",
                        ""
                    ).strip()
                )

                footer_text = (
                    request.form.get(
                        "footer_text",
                        ""
                    ).strip()
                )


                # =================================================
                # APPLICATION SETTINGS
                # =================================================

                application_status = (
                    request.form.get(
                        "application_status",
                        "Open"
                    ).strip()
                )


                if application_status not in (
                    "Open",
                    "Closed"
                ):

                    application_status = "Open"


                application_deadline = (
                    request.form.get(
                        "application_deadline",
                        ""
                    ).strip()
                )


                if not application_deadline:

                    application_deadline = None


                # =================================================
                # ATTENDANCE ENABLED
                # =================================================

                attendance_enabled = (
                    request.form.get(
                        "attendance_enabled"
                    ) == "1"
                )


                # =================================================
                # COMPANY GPS LOCATION
                # =================================================

                latitude_raw = (
                    request.form.get(
                        "company_latitude",
                        ""
                    ).strip()
                )

                longitude_raw = (
                    request.form.get(
                        "company_longitude",
                        ""
                    ).strip()
                )


                try:

                    company_latitude = (
                        float(latitude_raw)
                        if latitude_raw
                        else None
                    )

                    company_longitude = (
                        float(longitude_raw)
                        if longitude_raw
                        else None
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    flash(
                        "Please enter valid company GPS coordinates.",
                        "error"
                    )

                    return redirect(
                        url_for("admin_settings")
                    )


                # =================================================
                # VALIDATE LATITUDE
                # =================================================

                if company_latitude is not None:

                    if not (
                        -90
                        <= company_latitude
                        <= 90
                    ):

                        flash(
                            "Company latitude must be between -90 and 90.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                # =================================================
                # VALIDATE LONGITUDE
                # =================================================

                if company_longitude is not None:

                    if not (
                        -180
                        <= company_longitude
                        <= 180
                    ):

                        flash(
                            "Company longitude must be between -180 and 180.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                # =================================================
                # ATTENDANCE RADIUS
                # =================================================

                attendance_radius_raw = (
                    request.form.get(
                        "attendance_radius",
                        "100"
                    ).strip()
                )


                try:

                    attendance_radius = int(
                        attendance_radius_raw
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    attendance_radius = 100


                if attendance_radius < 20:

                    attendance_radius = 20


                if attendance_radius > 5000:

                    attendance_radius = 5000


                # =================================================
                # WORKING HOURS
                # =================================================

                clock_in_start = (
                    request.form.get(
                        "clock_in_start",
                        "06:00"
                    ).strip()
                )

                clock_in_end = (
                    request.form.get(
                        "clock_in_end",
                        "10:00"
                    ).strip()
                )

                clock_out_start = (
                    request.form.get(
                        "clock_out_start",
                        "15:00"
                    ).strip()
                )

                clock_out_end = (
                    request.form.get(
                        "clock_out_end",
                        "23:00"
                    ).strip()
                )


                # =================================================
                # VALIDATE WORKING HOURS
                # =================================================

                try:

                    datetime.strptime(
                        clock_in_start,
                        "%H:%M"
                    )

                    datetime.strptime(
                        clock_in_end,
                        "%H:%M"
                    )

                    datetime.strptime(
                        clock_out_start,
                        "%H:%M"
                    )

                    datetime.strptime(
                        clock_out_end,
                        "%H:%M"
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    flash(
                        "Please provide valid attendance times.",
                        "error"
                    )

                    return redirect(
                        url_for("admin_settings")
                    )


                # =================================================
                # LATE THRESHOLD
                # =================================================

                late_after_raw = (
                    request.form.get(
                        "late_after_minutes",
                        "15"
                    ).strip()
                )


                try:

                    late_after_minutes = int(
                        late_after_raw
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    late_after_minutes = 15


                if late_after_minutes < 0:

                    late_after_minutes = 0


                # =================================================
                # EARLY CLOCK-OUT THRESHOLD
                # =================================================

                early_before_raw = (
                    request.form.get(
                        "early_before_minutes",
                        "15"
                    ).strip()
                )


                try:

                    early_before_minutes = int(
                        early_before_raw
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    early_before_minutes = 15


                if early_before_minutes < 0:

                    early_before_minutes = 0


                # =================================================
                # ADMIN ACCOUNT
                # =================================================

                admin_username = (
                    request.form.get(
                        "admin_username",
                        ""
                    ).strip()
                )

                admin_password = (
                    request.form.get(
                        "admin_password",
                        ""
                    ).strip()
                )

                confirm_admin_password = (
                    request.form.get(
                        "confirm_admin_password",
                        ""
                    ).strip()
                )


                # =================================================
                # ADMIN USERNAME VALIDATION
                # =================================================

                if not admin_username:

                    flash(
                        "Admin username cannot be empty.",
                        "error"
                    )

                    return redirect(
                        url_for("admin_settings")
                    )


                # =================================================
                # ADMIN PASSWORD
                # =================================================

                admin_password_hash = None


                if admin_password:

                    if len(admin_password) < 6:

                        flash(
                            "Admin password must contain at least 6 characters.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                    if (
                        admin_password
                        != confirm_admin_password
                    ):

                        flash(
                            "Admin passwords do not match.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                    admin_password_hash = (
                        generate_password_hash(
                            admin_password
                        )
                    )


                # =================================================
                # CLOUDINARY LOGO
                # =================================================

                logo_url = None

                logo_file = request.files.get(
                    "logo"
                )


                if (
                    logo_file
                    and logo_file.filename
                ):

                    # -------------------------------------------------
                    # Validate extension
                    # -------------------------------------------------

                    original_filename = secure_filename(
                        logo_file.filename
                    )


                    if "." not in original_filename:

                        flash(
                            "Invalid logo file.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                    extension = (
                        original_filename
                        .rsplit(".", 1)[1]
                        .lower()
                    )


                    allowed_extensions = {
                        "png",
                        "jpg",
                        "jpeg",
                        "webp"
                    }


                    if extension not in allowed_extensions:

                        flash(
                            "Invalid logo format. "
                            "Please upload PNG, JPG, JPEG or WEBP.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                    # -------------------------------------------------
                    # Check Cloudinary configuration
                    # -------------------------------------------------

                    cloud_name = os.environ.get(
                        "CLOUDINARY_CLOUD_NAME"
                    )

                    api_key = os.environ.get(
                        "CLOUDINARY_API_KEY"
                    )

                    api_secret = os.environ.get(
                        "CLOUDINARY_API_SECRET"
                    )


                    if not all([
                        cloud_name,
                        api_key,
                        api_secret
                    ]):

                        app.logger.error(
                            "Cloudinary environment variables are missing."
                        )

                        flash(
                            "Cloudinary is not configured correctly. "
                            "Please check the Render environment variables.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                    # -------------------------------------------------
                    # Upload to Cloudinary
                    # -------------------------------------------------

                    try:

                        upload_result = (
                            cloudinary.uploader.upload(
                                logo_file,
                                folder="avking/company",
                                public_id="company_logo",
                                overwrite=True,
                                resource_type="image",
                                transformation=[
                                    {
                                        "width": 1000,
                                        "height": 1000,
                                        "crop": "limit",
                                        "quality": "auto"
                                    }
                                ]
                            )
                        )


                        logo_url = (
                            upload_result.get(
                                "secure_url"
                            )
                        )


                        if not logo_url:

                            raise RuntimeError(
                                "Cloudinary did not return a secure URL."
                            )


                        app.logger.info(
                            "Company logo uploaded successfully to Cloudinary."
                        )


                    except Exception:

                        app.logger.exception(
                            "Cloudinary logo upload failed."
                        )

                        flash(
                            "Unable to upload company logo. "
                            "Please try again.",
                            "error"
                        )

                        return redirect(
                            url_for("admin_settings")
                        )


                # =================================================
                # CREATE SETTINGS
                # =================================================

                if not existing_settings:

                    cur.execute(
                        """
                        INSERT INTO company_settings
                        (
                            company_name,
                            company_email,
                            company_phone,
                            company_address,
                            company_website,
                            footer_text,

                            application_status,
                            application_deadline,

                            logo,

                            admin_username,
                            admin_password_hash,

                            attendance_enabled,

                            company_latitude,
                            company_longitude,

                            attendance_radius,

                            clock_in_start,
                            clock_in_end,

                            clock_out_start,
                            clock_out_end,

                            late_after_minutes,
                            early_before_minutes
                        )

                        VALUES
                        (
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,

                            %s,
                            %s,

                            %s,

                            %s,
                            %s,

                            %s,

                            %s,
                            %s,

                            %s,

                            %s,
                            %s,

                            %s,
                            %s,

                            %s,
                            %s
                        )
                        """,
                        (
                            company_name,
                            company_email,
                            company_phone,
                            company_address,
                            company_website,
                            footer_text,

                            application_status,
                            application_deadline,

                            logo_url,

                            admin_username,
                            admin_password_hash,

                            attendance_enabled,

                            company_latitude,
                            company_longitude,

                            attendance_radius,

                            clock_in_start,
                            clock_in_end,

                            clock_out_start,
                            clock_out_end,

                            late_after_minutes,
                            early_before_minutes
                        )
                    )


                # =================================================
                # UPDATE EXISTING SETTINGS
                # =================================================

                else:

                    settings_id = (
                        existing_settings["id"]
                    )


                    # =================================================
                    # UPDATE WITH NEW LOGO
                    # =================================================

                    if logo_url:

                        cur.execute(
                            """
                            UPDATE company_settings

                            SET

                                company_name = %s,
                                company_email = %s,
                                company_phone = %s,
                                company_address = %s,
                                company_website = %s,
                                footer_text = %s,

                                application_status = %s,
                                application_deadline = %s,

                                logo = %s,

                                admin_username = %s,

                                attendance_enabled = %s,

                                company_latitude = %s,
                                company_longitude = %s,

                                attendance_radius = %s,

                                clock_in_start = %s,
                                clock_in_end = %s,

                                clock_out_start = %s,
                                clock_out_end = %s,

                                late_after_minutes = %s,
                                early_before_minutes = %s

                            WHERE id = %s
                            """,
                            (
                                company_name,
                                company_email,
                                company_phone,
                                company_address,
                                company_website,
                                footer_text,

                                application_status,
                                application_deadline,

                                logo_url,

                                admin_username,

                                attendance_enabled,

                                company_latitude,
                                company_longitude,

                                attendance_radius,

                                clock_in_start,
                                clock_in_end,

                                clock_out_start,
                                clock_out_end,

                                late_after_minutes,
                                early_before_minutes,

                                settings_id
                            )
                        )


                    # =================================================
                    # UPDATE WITHOUT CHANGING LOGO
                    # =================================================

                    else:

                        cur.execute(
                            """
                            UPDATE company_settings

                            SET

                                company_name = %s,
                                company_email = %s,
                                company_phone = %s,
                                company_address = %s,
                                company_website = %s,
                                footer_text = %s,

                                application_status = %s,
                                application_deadline = %s,

                                admin_username = %s,

                                attendance_enabled = %s,

                                company_latitude = %s,
                                company_longitude = %s,

                                attendance_radius = %s,

                                clock_in_start = %s,
                                clock_in_end = %s,

                                clock_out_start = %s,
                                clock_out_end = %s,

                                late_after_minutes = %s,
                                early_before_minutes = %s

                            WHERE id = %s
                            """,
                            (
                                company_name,
                                company_email,
                                company_phone,
                                company_address,
                                company_website,
                                footer_text,

                                application_status,
                                application_deadline,

                                admin_username,

                                attendance_enabled,

                                company_latitude,
                                company_longitude,

                                attendance_radius,

                                clock_in_start,
                                clock_in_end,

                                clock_out_start,
                                clock_out_end,

                                late_after_minutes,
                                early_before_minutes,

                                settings_id
                            )
                        )


                    # =================================================
                    # UPDATE ADMIN PASSWORD
                    # =================================================

                    if admin_password_hash:

                        cur.execute(
                            """
                            UPDATE company_settings

                            SET
                                admin_password_hash = %s

                            WHERE id = %s
                            """,
                            (
                                admin_password_hash,
                                settings_id
                            )
                        )


                # =================================================
                # COMMIT
                # =================================================

                conn.commit()


                # =================================================
                # SUCCESS
                # =================================================

                flash(
                    "Settings updated successfully.",
                    "success"
                )


                return redirect(
                    url_for(
                        "admin_settings"
                    )
                )


            # =================================================
            # GET CURRENT SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT *
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            settings = cur.fetchone()


    except Exception:

        try:
            conn.rollback()
        except Exception:
            pass

        app.logger.exception(
            "Error updating company settings"
        )

        flash(
            "Unable to update company settings. "
            "Please try again.",
            "error"
        )

        settings = None


    finally:

        conn.close()


    # =========================================================
    # RENDER SETTINGS
    # =========================================================

    return render_template(
        "admin_settings.html",
        settings=settings
    )
# ============================================================
# ADMIN APPLICATION DOCUMENT ROUTES
# CLOUDINARY
# ============================================================
# ============================================================
# ADMIN APPLICATION DOCUMENT VIEW / DOWNLOAD
# CLOUDINARY + POSTGRESQL + RENDER
# ============================================================

import os
import io
import requests

from flask import (
    redirect,
    url_for,
    flash,
    send_file,
    abort
)


# ============================================================
# GET APPLICANT DOCUMENT URL
# ============================================================

def get_application_document_url(application_id, document_type):

    allowed_documents = {
        "cv": "cv_filename",
        "passport": "passport_filename",
        "result": "qualification_filename",
    }

    column = allowed_documents.get(
        str(document_type).strip().lower()
    )

    if not column:
        return None

    conn = None

    try:

        conn = get_db()

        with conn.cursor() as cur:

            cur.execute(
                f"""
                SELECT {column}
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            row = cur.fetchone()

            if not row:
                return None

            value = row[column]

            if not value:
                return None

            value = str(value).strip()

            if not value:
                return None

            return value

    except Exception:

        app.logger.exception(
            "Error retrieving %s document for application %s",
            document_type,
            application_id
        )

        return None

    finally:

        if conn:

            try:
                conn.close()
            except Exception:
                pass


# ============================================================
# VIEW APPLICATION DOCUMENT
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/document/<document_type>/view",
    methods=["GET"]
)
def admin_view_application_document(
    application_id,
    document_type
):

    # --------------------------------------------------------
    # ADMIN LOGIN
    # --------------------------------------------------------

    if not admin_required():

        return redirect(
            url_for("admin_login")
        )

    document_type = str(
        document_type
    ).strip().lower()

    if document_type not in {
        "cv",
        "passport",
        "result"
    }:

        flash(
            "Invalid document requested.",
            "error"
        )

        return redirect(
            url_for(
                "admin_applications"
            )
        )

    # --------------------------------------------------------
    # GET CLOUDINARY URL
    # --------------------------------------------------------

    document_url = get_application_document_url(
        application_id,
        document_type
    )

    if not document_url:

        flash(
            "This document is not available.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )

    # --------------------------------------------------------
    # CLOUDINARY URL
    # --------------------------------------------------------

    if document_url.startswith(
        (
            "https://",
            "http://"
        )
    ):

        app.logger.info(
            "Viewing %s for application %s",
            document_type,
            application_id
        )

        return redirect(
            document_url
        )

    # --------------------------------------------------------
    # INVALID URL
    # --------------------------------------------------------

    flash(
        "The stored document URL is invalid.",
        "error"
    )

    return redirect(
        url_for(
            "admin_application_details",
            application_id=application_id
        )
    )


# ============================================================
# DOWNLOAD APPLICATION DOCUMENT
# ============================================================

@app.route(
    "/admin/applications/<int:application_id>/document/<document_type>/download",
    methods=["GET"]
)
def admin_download_application_document(
    application_id,
    document_type
):

    # --------------------------------------------------------
    # ADMIN LOGIN
    # --------------------------------------------------------

    if not admin_required():

        return redirect(
            url_for("admin_login")
        )

    document_type = str(
        document_type
    ).strip().lower()

    if document_type not in {
        "cv",
        "passport",
        "result"
    }:

        flash(
            "Invalid document requested.",
            "error"
        )

        return redirect(
            url_for(
                "admin_applications"
            )
        )

    # --------------------------------------------------------
    # GET CLOUDINARY URL
    # --------------------------------------------------------

    document_url = get_application_document_url(
        application_id,
        document_type
    )

    if not document_url:

        flash(
            "This document is not available.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )

    # --------------------------------------------------------
    # MAKE SURE IT IS A CLOUDINARY URL
    # --------------------------------------------------------

    if not document_url.startswith(
        (
            "https://",
            "http://"
        )
    ):

        flash(
            "The stored document URL is invalid.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )

    # --------------------------------------------------------
    # DOWNLOAD FROM CLOUDINARY
    #
    # We download the file through the Flask server and send
    # it to the administrator as an attachment.
    #
    # This avoids depending on Render's temporary filesystem.
    # --------------------------------------------------------

    try:

        response = requests.get(
            document_url,
            timeout=30,
            allow_redirects=True
        )

        response.raise_for_status()

    except requests.RequestException:

        app.logger.exception(
            "Unable to download %s document for application %s",
            document_type,
            application_id
        )

        flash(
            "Unable to download the document from Cloudinary.",
            "error"
        )

        return redirect(
            url_for(
                "admin_application_details",
                application_id=application_id
            )
        )

    # --------------------------------------------------------
    # DETERMINE FILE NAME
    # --------------------------------------------------------

    extension = ""

    content_type = (
        response.headers.get(
            "Content-Type",
            ""
        )
        .split(";")[0]
        .lower()
    )

    extension_map = {

        "application/pdf": ".pdf",

        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",

        "application/msword": ".doc",

        "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            ".docx",

        "application/vnd.ms-excel": ".xls",

        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            ".xlsx",
    }

    extension = extension_map.get(
        content_type,
        ""
    )

    # --------------------------------------------------------
    # FRIENDLY DOWNLOAD NAME
    # --------------------------------------------------------

    document_names = {

        "cv": "Applicant_CV",

        "passport": "Applicant_Passport",

        "result": "Applicant_Result",
    }

    download_name = (
        document_names.get(
            document_type,
            "Applicant_Document"
        )
        + extension
    )

    # --------------------------------------------------------
    # SEND FILE
    # --------------------------------------------------------

    return send_file(

        io.BytesIO(
            response.content
        ),

        mimetype=(
            content_type
            or "application/octet-stream"
        ),

        as_attachment=True,

        download_name=download_name
    )

@app.route("/admin/reports/export/pdf")
def export_admin_reports_pdf():

    # =========================================================
    # CHECK ADMIN LOGIN
    # =========================================================

    if not admin_required():
        return redirect(url_for("admin_login"))

    # =========================================================
    # IMPORT PDF COMPONENTS
    # =========================================================

    from io import BytesIO
    from datetime import datetime

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle
    )
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image
    )

    # =========================================================
    # DATABASE
    # =========================================================

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =====================================================
            # COMPANY SETTINGS
            # =====================================================

            cur.execute(
                """
                SELECT
                    company_name,
                    company_email,
                    company_phone,
                    company_address,
                    company_website,
                    footer_text,
                    logo
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
                """
            )

            settings = cur.fetchone()

            # =====================================================
            # APPLICATION REPORT
            #
            # created_at REMOVED
            # because it does not exist in PostgreSQL table
            # =====================================================

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    gender,
                    phone,
                    email,
                    state,
                    lga,
                    position_applied,
                    highest_qualification,
                    status
                FROM applications
                ORDER BY id DESC
                """
            )

            applications = cur.fetchall()

    finally:

        conn.close()

    # =========================================================
    # COMPANY INFORMATION
    # =========================================================

    if settings:

        company_name = (
            settings["company_name"]
            or "AV KING VET DRUG VENTURE"
        )

        company_email = (
            settings["company_email"]
            or ""
        )

        company_phone = (
            settings["company_phone"]
            or ""
        )

        company_address = (
            settings["company_address"]
            or ""
        )

        company_website = (
            settings["company_website"]
            or ""
        )

        footer_text = (
            settings["footer_text"]
            or ""
        )

        logo_filename = (
            settings["logo"]
            or ""
        )

    else:

        company_name = "AV KING VET DRUG VENTURE"
        company_email = ""
        company_phone = ""
        company_address = ""
        company_website = ""
        footer_text = ""
        logo_filename = ""

    # =========================================================
    # PDF BUFFER
    # =========================================================

    buffer = BytesIO()

    # =========================================================
    # PDF DOCUMENT
    # =========================================================

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=14 * mm
    )

    # =========================================================
    # STYLES
    # =========================================================

    styles = getSampleStyleSheet()

    company_name_style = ParagraphStyle(
        "CompanyName",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=4
    )

    company_info_style = ParagraphStyle(
        "CompanyInfo",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#4b5563")
    )

    report_title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceBefore=7,
        spaceAfter=4
    )

    date_style = ParagraphStyle(
        "DateStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#6b7280"),
        spaceAfter=10
    )

    body_style = ParagraphStyle(
        "BodyStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.4,
        leading=8,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#374151")
    )

    center_body_style = ParagraphStyle(
        "CenterBodyStyle",
        parent=body_style,
        alignment=TA_CENTER
    )

    header_style = ParagraphStyle(
        "HeaderStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=8,
        alignment=TA_CENTER,
        textColor=colors.white
    )

    # =========================================================
    # STORY
    # =========================================================

    story = []

    # =========================================================
    # COMPANY LOGO
    # =========================================================

    if logo_filename:

        logo_path = os.path.join(
            app.root_path,
            "static",
            "uploads",
            logo_filename
        )

        if os.path.exists(logo_path):

            try:

                logo = Image(
                    logo_path,
                    width=23 * mm,
                    height=23 * mm,
                    kind="proportional"
                )

                logo.hAlign = "CENTER"

                story.append(logo)

                story.append(
                    Spacer(1, 3 * mm)
                )

            except Exception:

                app.logger.warning(
                    "Unable to load company logo for PDF.",
                    exc_info=True
                )

    # =========================================================
    # COMPANY NAME
    # =========================================================

    story.append(
        Paragraph(
            company_name,
            company_name_style
        )
    )

    # =========================================================
    # ADDRESS
    # =========================================================

    if company_address:

        story.append(
            Paragraph(
                company_address,
                company_info_style
            )
        )

    # =========================================================
    # CONTACT INFORMATION
    # =========================================================

    contact_parts = []

    if company_phone:

        contact_parts.append(
            f"Phone: {company_phone}"
        )

    if company_email:

        contact_parts.append(
            f"Email: {company_email}"
        )

    if contact_parts:

        story.append(
            Paragraph(
                " &nbsp;&nbsp; | &nbsp;&nbsp; ".join(
                    contact_parts
                ),
                company_info_style
            )
        )

    # =========================================================
    # WEBSITE
    # =========================================================

    if company_website:

        story.append(
            Paragraph(
                company_website,
                company_info_style
            )
        )

    story.append(
        Spacer(1, 4 * mm)
    )

    # =========================================================
    # DIVIDER
    # =========================================================

    divider = Table(
        [[""]],
        colWidths=[277 * mm],
        rowHeights=[1]
    )

    divider.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    colors.HexColor("#d1d5db")
                )
            ]
        )
    )

    story.append(divider)

    # =========================================================
    # REPORT TITLE
    # =========================================================

    story.append(
        Paragraph(
            "APPLICANT RECRUITMENT REPORT",
            report_title_style
        )
    )

    # =========================================================
    # REPORT DATE
    # =========================================================

    generated_date = datetime.now().strftime(
        "%d %B %Y, %I:%M %p"
    )

    story.append(
        Paragraph(
            f"Generated on {generated_date}",
            date_style
        )
    )

    # =========================================================
    # CALCULATE SUMMARY
    # =========================================================

    total_applications = len(
        applications
    )

    pending_count = 0
    review_count = 0
    shortlisted_count = 0
    approved_count = 0
    rejected_count = 0

    for application in applications:

        status = (
            application["status"]
            or "Pending"
        )

        status = str(status).strip().lower()

        if status == "pending":

            pending_count += 1

        elif status == "under review":

            review_count += 1

        elif status == "shortlisted":

            shortlisted_count += 1

        elif status == "approved":

            approved_count += 1

        elif status == "rejected":

            rejected_count += 1

    # =========================================================
    # SUMMARY TABLE
    # =========================================================

    summary_data = [
        [
            "TOTAL",
            "PENDING",
            "UNDER REVIEW",
            "SHORTLISTED",
            "APPROVED",
            "REJECTED"
        ],
        [
            str(total_applications),
            str(pending_count),
            str(review_count),
            str(shortlisted_count),
            str(approved_count),
            str(rejected_count)
        ]
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            46 * mm,
            46 * mm,
            46 * mm,
            46 * mm,
            46 * mm,
            46 * mm
        ]
    )

    summary_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#111827")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, 1),
                    colors.HexColor("#f8fafc")
                ),

                (
                    "FONTNAME",
                    (0, 1),
                    (-1, 1),
                    "Helvetica-Bold"
                ),

                (
                    "FONTSIZE",
                    (0, 1),
                    (-1, 1),
                    12
                ),

                (
                    "TEXTCOLOR",
                    (0, 1),
                    (-1, 1),
                    colors.HexColor("#111827")
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#d1d5db")
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7
                )
            ]
        )
    )

    story.append(
        summary_table
    )

    story.append(
        Spacer(1, 7 * mm)
    )

    # =========================================================
    # APPLICATION TABLE
    # =========================================================

    table_data = [
        [
            "S/N",
            "APPLICATION NO.",
            "APPLICANT NAME",
            "GENDER",
            "PHONE",
            "EMAIL",
            "STATE",
            "LGA",
            "POSITION",
            "QUALIFICATION",
            "STATUS"
        ]
    ]

    for index, application in enumerate(
        applications,
        start=1
    ):

        first_name = (
            application["first_name"]
            or ""
        )

        middle_name = (
            application["middle_name"]
            or ""
        )

        last_name = (
            application["last_name"]
            or ""
        )

        full_name = " ".join(
            part
            for part in [
                first_name,
                middle_name,
                last_name
            ]
            if part
        )

        status = (
            application["status"]
            or "Pending"
        )

        table_data.append(
            [
                str(index),

                application["application_number"]
                or "—",

                full_name or "—",

                application["gender"]
                or "—",

                application["phone"]
                or "—",

                application["email"]
                or "—",

                application["state"]
                or "—",

                application["lga"]
                or "—",

                application["position_applied"]
                or "—",

                application["highest_qualification"]
                or "—",

                status
            ]
        )

    # =========================================================
    # FORMAT TABLE
    # =========================================================

    formatted_table_data = []

    for row_index, row in enumerate(
        table_data
    ):

        formatted_row = []

        for column_index, value in enumerate(row):

            if row_index == 0:

                formatted_row.append(
                    Paragraph(
                        str(value),
                        header_style
                    )
                )

            else:

                if column_index in [
                    0,
                    1,
                    3,
                    10
                ]:

                    formatted_row.append(
                        Paragraph(
                            str(value),
                            center_body_style
                        )
                    )

                else:

                    formatted_row.append(
                        Paragraph(
                            str(value),
                            body_style
                        )
                    )

        formatted_table_data.append(
            formatted_row
        )

    # =========================================================
    # APPLICATION TABLE
    # =========================================================

    application_table = Table(
        formatted_table_data,
        repeatRows=1,
        colWidths=[
            9 * mm,
            27 * mm,
            40 * mm,
            17 * mm,
            27 * mm,
            43 * mm,
            20 * mm,
            20 * mm,
            31 * mm,
            31 * mm,
            25 * mm
        ]
    )

    application_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#111827")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    colors.HexColor("#d1d5db")
                ),

                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#f8fafc")
                    ]
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                )
            ]
        )
    )

    story.append(
        application_table
    )

    # =========================================================
    # FOOTER
    # =========================================================

    if footer_text:

        story.append(
            Spacer(1, 6 * mm)
        )

        story.append(
            Paragraph(
                footer_text,
                company_info_style
            )
        )

    # =========================================================
    # BUILD PDF
    # =========================================================

    document.build(
        story
    )

    # =========================================================
    # PREPARE DOWNLOAD
    # =========================================================

    buffer.seek(0)

    filename = (
        "applicant_recruitment_report_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".pdf"
    )

    # =========================================================
    # RETURN PDF
    # =========================================================

    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

@app.route("/admin/reports/export/excel")
def export_admin_reports_excel():

    # =========================================================
    # CHECK ADMIN LOGIN
    # =========================================================

    if not admin_required():
        return redirect(url_for("admin_login"))


    conn = get_db()

    try:

        with conn.cursor() as cur:

            # =================================================
            # COMPANY SETTINGS
            # =================================================

            cur.execute("""
                SELECT
                    company_name,
                    company_address,
                    company_phone,
                    company_email,
                    logo
                FROM company_settings
                ORDER BY id ASC
                LIMIT 1
            """)

            settings = cur.fetchone()


            # =================================================
            # APPLICATIONS
            # =================================================

            cur.execute("""
                SELECT
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    position_applied,
                    phone,
                    email,
                    gender,
                    state,
                    lga,
                    highest_qualification,
                    status,
                    submitted_at
                FROM applications
                ORDER BY submitted_at DESC NULLS LAST
            """)

            applications = cur.fetchall()

    finally:

        conn.close()


    # =========================================================
    # COMPANY INFORMATION
    # =========================================================

    company_name = (
        settings["company_name"]
        if settings and settings["company_name"]
        else "AV KING VET DRUG VENTURE"
    )

    company_address = (
        settings["company_address"]
        if settings and settings["company_address"]
        else ""
    )

    company_phone = (
        settings["company_phone"]
        if settings and settings["company_phone"]
        else ""
    )

    company_email = (
        settings["company_email"]
        if settings and settings["company_email"]
        else ""
    )

    company_logo = (
        settings["logo"]
        if settings and settings["logo"]
        else None
    )


    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Applicant Report"


    # =========================================================
    # PAGE SETTINGS
    # =========================================================

    worksheet.page_setup.orientation = "landscape"

    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )

    worksheet.page_setup.fitToWidth = 1

    worksheet.page_setup.fitToHeight = 0

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins = PageMargins(

        left=0.25,
        right=0.25,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2
    )


    # =========================================================
    # COLORS
    # =========================================================

    dark = "111827"

    white = "FFFFFF"

    light_gray = "F3F4F6"

    border_gray = "D1D5DB"

    text_gray = "4B5563"

    green = "059669"

    orange = "EA580C"

    blue = "2563EB"

    red = "DC2626"


    # =========================================================
    # BORDER
    # =========================================================

    thin_border = Border(

        left=Side(
            style="thin",
            color=border_gray
        ),

        right=Side(
            style="thin",
            color=border_gray
        ),

        top=Side(
            style="thin",
            color=border_gray
        ),

        bottom=Side(
            style="thin",
            color=border_gray
        )
    )


    # =========================================================
    # LOGO
    # =========================================================

    current_row = 1

    if company_logo:

        try:

            logo_path = company_logo

            # ---------------------------------------------
            # Convert relative upload path to filesystem path
            # ---------------------------------------------

            if not os.path.isabs(logo_path):

                logo_path = os.path.join(
                    os.getcwd(),
                    logo_path.lstrip("/")
                )


            if os.path.exists(logo_path):

                logo = XLImage(logo_path)

                logo.width = 75

                logo.height = 75

                worksheet.add_image(
                    logo,
                    "A1"
                )

                current_row = 2

        except Exception:

            pass


    # =========================================================
    # COMPANY NAME
    # =========================================================

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=10
    )

    company_cell = worksheet.cell(
        row=1,
        column=1
    )

    company_cell.value = company_name

    company_cell.font = Font(
        name="Calibri",
        size=20,
        bold=True,
        color=dark
    )

    company_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.row_dimensions[1].height = 30


    # =========================================================
    # COMPANY ADDRESS
    # =========================================================

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=10
    )

    address_cell = worksheet.cell(
        row=2,
        column=1
    )

    contact_text = company_address


    if company_phone:

        if contact_text:

            contact_text += "  |  "

        contact_text += (
            "Phone: "
            + str(company_phone)
        )


    if company_email:

        if contact_text:

            contact_text += "  |  "

        contact_text += (
            "Email: "
            + str(company_email)
        )


    address_cell.value = contact_text

    address_cell.font = Font(
        name="Calibri",
        size=10,
        color=text_gray
    )

    address_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    worksheet.row_dimensions[2].height = 30


    # =========================================================
    # REPORT TITLE
    # =========================================================

    worksheet.merge_cells(
        start_row=4,
        start_column=1,
        end_row=4,
        end_column=10
    )

    title_cell = worksheet.cell(
        row=4,
        column=1
    )

    title_cell.value = (
        "APPLICANT RECRUITMENT REPORT"
    )

    title_cell.font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color=dark
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.row_dimensions[4].height = 26


    # =========================================================
    # REPORT SUBTITLE
    # =========================================================

    worksheet.merge_cells(
        start_row=5,
        start_column=1,
        end_row=5,
        end_column=10
    )

    subtitle_cell = worksheet.cell(
        row=5,
        column=1
    )

    subtitle_cell.value = (
        "Complete recruitment application summary"
    )

    subtitle_cell.font = Font(
        name="Calibri",
        size=10,
        italic=True,
        color=text_gray
    )

    subtitle_cell.alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # GENERATED DATE
    # =========================================================

    worksheet.merge_cells(
        start_row=6,
        start_column=1,
        end_row=6,
        end_column=10
    )

    generated_cell = worksheet.cell(
        row=6,
        column=1
    )

    generated_cell.value = (
        "Generated: "
        + datetime.now().strftime(
            "%d %B %Y at %I:%M %p"
        )
    )

    generated_cell.font = Font(
        name="Calibri",
        size=9,
        color=text_gray
    )

    generated_cell.alignment = Alignment(
        horizontal="center"
    )


    # =========================================================
    # SUMMARY
    # =========================================================

    total = len(applications)

    pending = sum(
        1
        for row in applications
        if row["status"] == "Pending"
    )

    under_review = sum(
        1
        for row in applications
        if row["status"] == "Under Review"
    )

    shortlisted = sum(
        1
        for row in applications
        if row["status"] == "Shortlisted"
    )

    approved = sum(
        1
        for row in applications
        if row["status"] == "Approved"
    )

    rejected = sum(
        1
        for row in applications
        if row["status"] == "Rejected"
    )


    summary_row = 8


    summary_headers = [

        "TOTAL",
        "PENDING",
        "UNDER REVIEW",
        "SHORTLISTED",
        "APPROVED",
        "REJECTED"

    ]


    summary_values = [

        total,
        pending,
        under_review,
        shortlisted,
        approved,
        rejected

    ]


    summary_columns = [
        1,
        2,
        3,
        4,
        5,
        6
    ]


    for index, column in enumerate(
        summary_columns
    ):

        header = worksheet.cell(
            row=summary_row,
            column=column
        )

        header.value = summary_headers[index]

        header.font = Font(
            name="Calibri",
            size=9,
            bold=True,
            color=white
        )

        header.fill = PatternFill(
            fill_type="solid",
            fgColor=dark
        )

        header.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        header.border = thin_border


        value = worksheet.cell(
            row=summary_row + 1,
            column=column
        )

        value.value = summary_values[index]

        value.font = Font(
            name="Calibri",
            size=14,
            bold=True,
            color=dark
        )

        value.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        value.border = thin_border


    worksheet.row_dimensions[
        summary_row
    ].height = 22

    worksheet.row_dimensions[
        summary_row + 1
    ].height = 28


    # =========================================================
    # APPLICATION TABLE
    # =========================================================

    header_row = 11


    headers = [

        "Application Number",
        "Applicant Name",
        "Position Applied",
        "Phone",
        "Email",
        "Gender",
        "State",
        "LGA",
        "Highest Qualification",
        "Status"

    ]


    for column, header_text in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column
        )

        cell.value = header_text

        cell.font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color=white
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=dark
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = thin_border


    worksheet.row_dimensions[
        header_row
    ].height = 32


    # =========================================================
    # APPLICATION DATA
    # =========================================================

    row_number = header_row + 1


    for application in applications:

        full_name = " ".join(

            part

            for part in [

                application["first_name"],
                application["middle_name"],
                application["last_name"]

            ]

            if part
        )


        data = [

            application["application_number"]
            or "—",

            full_name
            or "—",

            application["position_applied"]
            or "—",

            application["phone"]
            or "—",

            application["email"]
            or "—",

            application["gender"]
            or "—",

            application["state"]
            or "—",

            application["lga"]
            or "—",

            application["highest_qualification"]
            or "—",

            application["status"]
            or "Pending"

        ]


        for column, value in enumerate(
            data,
            start=1
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column
            )

            cell.value = value

            cell.font = Font(
                name="Calibri",
                size=9,
                color=dark
            )

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

            cell.border = thin_border


            # ---------------------------------------------
            # Status formatting
            # ---------------------------------------------

            if column == 10:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                status = str(value).lower()


                if status == "approved":

                    cell.font = Font(
                        name="Calibri",
                        size=9,
                        bold=True,
                        color=green
                    )


                elif status == "shortlisted":

                    cell.font = Font(
                        name="Calibri",
                        size=9,
                        bold=True,
                        color=green
                    )


                elif status == "rejected":

                    cell.font = Font(
                        name="Calibri",
                        size=9,
                        bold=True,
                        color=red
                    )


                elif status == "under review":

                    cell.font = Font(
                        name="Calibri",
                        size=9,
                        bold=True,
                        color=blue
                    )


                elif status == "pending":

                    cell.font = Font(
                        name="Calibri",
                        size=9,
                        bold=True,
                        color=orange
                    )


        # Alternating row background

        if row_number % 2 == 0:

            for column in range(
                1,
                len(headers) + 1
            ):

                worksheet.cell(
                    row=row_number,
                    column=column
                ).fill = PatternFill(
                    fill_type="solid",
                    fgColor=light_gray
                )


        worksheet.row_dimensions[
            row_number
        ].height = 25


        row_number += 1


    # =========================================================
    # FILTER
    # =========================================================

    if row_number > header_row + 1:

        worksheet.auto_filter.ref = (

            f"A{header_row}:"
            f"J{row_number - 1}"

        )


    # =========================================================
    # FREEZE HEADER
    # =========================================================

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )


    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    widths = {

        "A": 21,
        "B": 28,
        "C": 23,
        "D": 17,
        "E": 30,
        "F": 12,
        "G": 18,
        "H": 20,
        "I": 25,
        "J": 18

    }


    for column, width in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width


    # =========================================================
    # PRINT AREA
    # =========================================================

    if row_number > header_row:

        worksheet.print_area = (
            f"A1:J{row_number - 1}"
        )


    # =========================================================
    # REPEAT HEADER WHEN PRINTING
    # =========================================================

    worksheet.print_title_rows = (
        f"1:{header_row}"
    )


    # =========================================================
    # HEADER / FOOTER
    # =========================================================

    worksheet.oddHeader.center.text = (
        "&B" + company_name
    )

    worksheet.oddFooter.center.text = (
        "Confidential Recruitment Document"
    )

    worksheet.oddFooter.right.text = (
        "Page &P of &N"
    )


    # =========================================================
    # ACTIVE CELL
    # =========================================================

    worksheet.sheet_view.selection[0].activeCell = (
        "A1"
    )

    worksheet.sheet_view.selection[0].sqref = (
        "A1"
    )


    # =========================================================
    # SAVE TO MEMORY
    # =========================================================

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)


    # =========================================================
    # DOWNLOAD
    # =========================================================

    return send_file(

        output,

        as_attachment=True,

        download_name=(
            "Applicant_Recruitment_Report.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        )
    )

@app.route("/admin/reports")
def admin_reports():

    if not admin_required():
        return redirect(url_for("admin_login"))

    conn = get_db()

    try:

        with conn.cursor() as cur:

            # Total applications
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM applications
            """)
            total = cur.fetchone()["total"]


            # Pending
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Pending'
            """)
            pending = cur.fetchone()["total"]


            # Under Review
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Under Review'
            """)
            under_review = cur.fetchone()["total"]


            # Shortlisted
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Shortlisted'
            """)
            shortlisted = cur.fetchone()["total"]


            # Approved
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Approved'
            """)
            approved = cur.fetchone()["total"]


            # Rejected
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM applications
                WHERE status = 'Rejected'
            """)
            rejected = cur.fetchone()["total"]


            # Applications by position
            cur.execute("""
                SELECT
                    position_applied,
                    COUNT(*) AS total
                FROM applications
                GROUP BY position_applied
                ORDER BY total DESC
            """)

            position_report = cur.fetchall()


            # Applications by gender
            cur.execute("""
                SELECT
                    gender,
                    COUNT(*) AS total
                FROM applications
                GROUP BY gender
                ORDER BY total DESC
            """)

            gender_report = cur.fetchall()


    finally:

        conn.close()


    return render_template(
        "admin_reports.html",

        total=total,
        pending=pending,
        under_review=under_review,
        shortlisted=shortlisted,
        approved=approved,
        rejected=rejected,

        position_report=position_report,
        gender_report=gender_report
        )


def format_date(date_value):

    if not date_value:
        return "Not Available"

    try:

        if isinstance(date_value, datetime):
            return date_value.strftime("%d %B %Y")

        if hasattr(date_value, "strftime"):
            return date_value.strftime("%d %B %Y")

        if isinstance(date_value, str):

            try:
                parsed_date = datetime.fromisoformat(
                    date_value.replace("Z", "+00:00")
                )

                return parsed_date.strftime(
                    "%d %B %Y"
                )

            except ValueError:

                try:
                    parsed_date = datetime.strptime(
                        date_value,
                        "%Y-%m-%d"
                    )

                    return parsed_date.strftime(
                        "%d %B %Y"
                    )

                except ValueError:
                    return date_value

        return str(date_value)

    except Exception:

        app.logger.exception(
            "Error formatting application date"
        )

        return str(date_value)
        
@app.route(
    "/admin/applications/<int:application_id>/biodata/pdf"
)
def download_applicant_biodata(application_id):

    # =========================================================
    # ADMIN LOGIN
    # =========================================================

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    # =========================================================
    # IMPORTS
    # =========================================================

    import os
    import html
    import requests

    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle
    )
    from reportlab.lib.units import mm

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image,
        HRFlowable
    )

    # =========================================================
    # DATABASE
    # =========================================================

    conn = None

    applicant = None
    settings = None

    try:

        conn = get_db()

        with conn.cursor() as cur:

            # -------------------------------------------------
            # GET APPLICANT
            # -------------------------------------------------

            cur.execute(
                """
                SELECT *
                FROM applications
                WHERE id = %s
                LIMIT 1
                """,
                (application_id,)
            )

            applicant = cur.fetchone()

            # -------------------------------------------------
            # GET COMPANY SETTINGS
            # -------------------------------------------------

            cur.execute(
                """
                SELECT
                    company_name,
                    company_email,
                    company_phone,
                    company_address,
                    company_website,
                    footer_text,
                    logo
                FROM company_settings
                ORDER BY id DESC
                LIMIT 1
                """
            )

            settings = cur.fetchone()

    except Exception:

        app.logger.exception(
            "Error loading applicant biodata for application ID %s",
            application_id
        )

        flash(
            "Unable to load applicant biodata.",
            "error"
        )

        return redirect(
            url_for("admin_applications")
        )

    finally:

        if conn:

            try:
                conn.close()
            except Exception:
                pass

    # =========================================================
    # APPLICANT NOT FOUND
    # =========================================================

    if not applicant:

        flash(
            "Applicant not found.",
            "error"
        )

        return redirect(
            url_for("admin_applications")
        )

    # =========================================================
    # SAFE VALUE HELPER
    # =========================================================

    def value(
        field,
        default="—"
    ):

        try:

            result = applicant[field]

        except (
            KeyError,
            TypeError,
            IndexError
        ):

            result = None

        if result is None:

            return default

        result = str(
            result
        ).strip()

        if not result:

            return default

        return result

    # =========================================================
    # SAFE HTML HELPER
    # =========================================================

    def safe_text(text):

        if text is None:
            return ""

        return (
            html.escape(
                str(text)
            )
            .replace(
                "\n",
                "<br/>"
            )
        )

    # =========================================================
    # COMPANY SETTINGS
    # =========================================================

    if settings:

        try:

            company_name = (
                settings["company_name"]
                or "AV KING VET DRUG VENTURE"
            )

        except Exception:

            company_name = (
                "AV KING VET DRUG VENTURE"
            )

        try:

            company_email = (
                settings["company_email"]
                or ""
            )

        except Exception:

            company_email = ""

        try:

            company_phone = (
                settings["company_phone"]
                or ""
            )

        except Exception:

            company_phone = ""

        try:

            company_address = (
                settings["company_address"]
                or ""
            )

        except Exception:

            company_address = ""

        try:

            company_website = (
                settings["company_website"]
                or ""
            )

        except Exception:

            company_website = ""

        try:

            footer_text = (
                settings["footer_text"]
                or ""
            )

        except Exception:

            footer_text = ""

        try:

            logo_value = (
                settings["logo"]
                or ""
            )

        except Exception:

            logo_value = ""

    else:

        company_name = (
            "AV KING VET DRUG VENTURE"
        )

        company_email = ""
        company_phone = ""
        company_address = ""
        company_website = ""
        footer_text = ""
        logo_value = ""

    # =========================================================
    # IMAGE LOADER
    #
    # Supports:
    #
    # 1. Cloudinary URL
    # 2. Other HTTPS URL
    # 3. /static/...
    # 4. static/...
    # 5. filename stored locally
    #
    # =========================================================

    def load_image(
        image_value,
        label="image"
    ):

        if not image_value:
            return None

        image_value = str(
            image_value
        ).strip()

        if not image_value:
            return None

        # -----------------------------------------------------
        # REMOTE / CLOUDINARY IMAGE
        # -----------------------------------------------------

        if (
            image_value.startswith("http://")
            or
            image_value.startswith("https://")
        ):

            try:

                response = requests.get(
                    image_value,
                    timeout=25,
                    headers={
                        "User-Agent":
                            "AV-KING-Recruitment-Portal/1.0"
                    }
                )

                response.raise_for_status()

                if not response.content:

                    raise ValueError(
                        "Remote image returned empty content."
                    )

                image_buffer = BytesIO(
                    response.content
                )

                image_buffer.seek(0)

                app.logger.info(
                    "%s loaded successfully from remote URL.",
                    label
                )

                return image_buffer

            except Exception:

                app.logger.exception(
                    "Unable to load %s from URL: %s",
                    label,
                    image_value
                )

                return None

        # -----------------------------------------------------
        # LOCAL FILE
        # -----------------------------------------------------

        if image_value.startswith(
            "/static/"
        ):

            local_path = os.path.join(
                app.root_path,
                image_value.lstrip("/")
            )

        elif image_value.startswith(
            "static/"
        ):

            local_path = os.path.join(
                app.root_path,
                image_value
            )

        else:

            local_path = os.path.join(
                app.root_path,
                "static",
                "uploads",
                os.path.basename(
                    image_value
                )
            )

        local_path = os.path.abspath(
            local_path
        )

        if not os.path.isfile(
            local_path
        ):

            app.logger.warning(
                "%s not found locally: %s",
                label,
                local_path
            )

            return None

        try:

            with open(
                local_path,
                "rb"
            ) as file:

                image_buffer = BytesIO(
                    file.read()
                )

            image_buffer.seek(0)

            app.logger.info(
                "%s loaded successfully from local file.",
                label
            )

            return image_buffer

        except Exception:

            app.logger.exception(
                "Unable to read %s: %s",
                label,
                local_path
            )

            return None

    # =========================================================
    # LOAD COMPANY LOGO
    # =========================================================

    logo_buffer = load_image(
        logo_value,
        "Company logo"
    )

    # =========================================================
    # LOAD APPLICANT PASSPORT
    #
    # Your database should contain the Cloudinary URL in
    # passport_filename.
    #
    # If your actual column is "passport", change the field
    # below from passport_filename to passport.
    #
    # =========================================================

    passport_value = value(
        "passport_filename",
        ""
    )

    if passport_value == "—":

        passport_value = ""

    passport_buffer = load_image(
        passport_value,
        "Applicant passport"
    )

    # =========================================================
    # PDF BUFFER
    # =========================================================

    buffer = BytesIO()

    # =========================================================
    # PDF DOCUMENT
    # =========================================================

    document = SimpleDocTemplate(

        buffer,

        pagesize=A4,

        rightMargin=15 * mm,

        leftMargin=15 * mm,

        topMargin=12 * mm,

        bottomMargin=18 * mm
    )

    # =========================================================
    # STYLES
    # =========================================================

    styles = getSampleStyleSheet()

    # ---------------------------------------------------------
    # COMPANY NAME
    # ---------------------------------------------------------

    company_style = ParagraphStyle(
        "CompanyStyle",

        parent=styles["Normal"],

        fontName="Helvetica-Bold",

        fontSize=14,

        leading=17,

        alignment=TA_CENTER,

        textColor=colors.HexColor(
            "#111827"
        ),

        spaceAfter=2
    )

    # ---------------------------------------------------------
    # COMPANY INFORMATION
    # ---------------------------------------------------------

    company_info_style = ParagraphStyle(
        "CompanyInfo",

        parent=styles["Normal"],

        fontName="Helvetica",

        fontSize=7.5,

        leading=10,

        alignment=TA_CENTER,

        textColor=colors.HexColor(
            "#4b5563"
        )
    )

    # ---------------------------------------------------------
    # MAIN TITLE
    # ---------------------------------------------------------

    title_style = ParagraphStyle(
        "TitleStyle",

        parent=styles["Normal"],

        fontName="Helvetica-Bold",

        fontSize=15,

        leading=19,

        alignment=TA_CENTER,

        textColor=colors.HexColor(
            "#111827"
        )
    )

    # ---------------------------------------------------------
    # SECTION HEADER
    # ---------------------------------------------------------

    section_style = ParagraphStyle(
        "SectionStyle",

        parent=styles["Normal"],

        fontName="Helvetica-Bold",

        fontSize=9,

        leading=11,

        textColor=colors.white
    )

    # ---------------------------------------------------------
    # LABEL
    # ---------------------------------------------------------

    label_style = ParagraphStyle(
        "LabelStyle",

        parent=styles["Normal"],

        fontName="Helvetica-Bold",

        fontSize=7.5,

        leading=10,

        textColor=colors.HexColor(
            "#374151"
        )
    )

    # ---------------------------------------------------------
    # DATA
    # ---------------------------------------------------------

    data_style = ParagraphStyle(
        "DataStyle",

        parent=styles["Normal"],

        fontName="Helvetica",

        fontSize=8.5,

        leading=11,

        textColor=colors.HexColor(
            "#111827"
        )
    )

    # ---------------------------------------------------------
    # SMALL
    # ---------------------------------------------------------

    small_style = ParagraphStyle(
        "SmallStyle",

        parent=styles["Normal"],

        fontName="Helvetica",

        fontSize=7.5,

        leading=10,

        textColor=colors.HexColor(
            "#6b7280"
        )
    )

    # =========================================================
    # STORY
    # =========================================================

    story = []

    # =========================================================
    # COMPANY LOGO - LEFT
    # =========================================================

    if logo_buffer:

        try:

            logo_buffer.seek(0)

            header_logo = Image(

                logo_buffer,

                width=30 * mm,

                height=30 * mm,

                kind="proportional"
            )

        except Exception:

            app.logger.exception(
                "Unable to create company logo image."
            )

            header_logo = Paragraph(
                "AV KING",
                company_style
            )

    else:

        header_logo = Paragraph(
            "AV KING",
            company_style
        )

    # =========================================================
    # APPLICANT PASSPORT - RIGHT
    # =========================================================

    if passport_buffer:

        try:

            passport_buffer.seek(0)

            header_passport = Image(

                passport_buffer,

                width=30 * mm,

                height=38 * mm,

                kind="proportional"
            )

        except Exception:

            app.logger.exception(
                "Unable to create applicant passport image."
            )

            header_passport = Paragraph(
                "PASSPORT<br/>PHOTO",
                small_style
            )

    else:

        header_passport = Paragraph(
            "PASSPORT<br/>PHOTO",
            small_style
        )

    # =========================================================
    # COMPANY INFORMATION - CENTER
    # =========================================================

    company_details = []

    company_details.append(
        Paragraph(
            safe_text(company_name),
            company_style
        )
    )

    if company_address:

        company_details.append(
            Paragraph(
                safe_text(company_address),
                company_info_style
            )
        )

    if company_phone:

        company_details.append(
            Paragraph(
                "Phone: "
                + safe_text(company_phone),
                company_info_style
            )
        )

    if company_email:

        company_details.append(
            Paragraph(
                "Email: "
                + safe_text(company_email),
                company_info_style
            )
        )

    if company_website:

        company_details.append(
            Paragraph(
                safe_text(company_website),
                company_info_style
            )
        )

    # =========================================================
    # PROFESSIONAL HEADER
    #
    # LEFT:
    # COMPANY LOGO
    #
    # CENTER:
    # COMPANY DETAILS
    #
    # RIGHT:
    # APPLICANT PASSPORT
    #
    # =========================================================

    header_table = Table(

        [
            [
                header_logo,
                company_details,
                header_passport
            ]
        ],

        colWidths=[
            40 * mm,
            100 * mm,
            40 * mm
        ],

        hAlign="CENTER"
    )

    header_table.setStyle(
        TableStyle(
            [

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),

                # -------------------------------------------------
                # LEFT LOGO
                # -------------------------------------------------

                (
                    "ALIGN",
                    (0, 0),
                    (0, 0),
                    "LEFT"
                ),

                # -------------------------------------------------
                # CENTER COMPANY DETAILS
                # -------------------------------------------------

                (
                    "ALIGN",
                    (1, 0),
                    (1, 0),
                    "CENTER"
                ),

                # -------------------------------------------------
                # RIGHT PASSPORT
                # -------------------------------------------------

                (
                    "ALIGN",
                    (2, 0),
                    (2, 0),
                    "RIGHT"
                ),

                # -------------------------------------------------
                # PADDING
                # -------------------------------------------------

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    0
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    0
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    0
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    0
                )
            ]
        )
    )

    story.append(
        header_table
    )

    story.append(
        Spacer(
            1,
            5 * mm
        )
    )

    # =========================================================
    # HEADER LINE
    # =========================================================

    story.append(
        HRFlowable(
            width="100%",

            thickness=1,

            color=colors.HexColor(
                "#d1d5db"
            )
        )
    )

    story.append(
        Spacer(
            1,
            4 * mm
        )
    )

    # =========================================================
    # TITLE
    # =========================================================

    story.append(
        Paragraph(
            "APPLICANT BIODATA",
            title_style
        )
    )

    story.append(
        Spacer(
            1,
            2 * mm
        )
    )

    story.append(
        Paragraph(
            (
                "Application No.: "
                "<b>"
                + safe_text(
                    value(
                        "application_number"
                    )
                )
                + "</b>"
            ),
            company_info_style
        )
    )

    story.append(
        Spacer(
            1,
            5 * mm
        )
    )

    # =========================================================
    # SECTION HEADER FUNCTION
    # =========================================================

    def section_header(title):

        table = Table(

            [
                [
                    Paragraph(
                        safe_text(title),
                        section_style
                    )
                ]
            ],

            colWidths=[
                180 * mm
            ]
        )

        table.setStyle(
            TableStyle(
                [

                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        colors.HexColor(
                            "#111827"
                        )
                    ),

                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        7
                    ),

                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        7
                    ),

                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    ),

                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ]
            )
        )

        story.append(
            table
        )

        story.append(
            Spacer(
                1,
                2 * mm
            )
        )

    # =========================================================
    # INFORMATION TABLE FUNCTION
    # =========================================================

    def information_table(rows):

        table_data = []

        for label, data in rows:

            table_data.append(
                [

                    Paragraph(
                        safe_text(label),
                        label_style
                    ),

                    Paragraph(
                        safe_text(data),
                        data_style
                    )
                ]
            )

        table = Table(

            table_data,

            colWidths=[
                48 * mm,
                132 * mm
            ],

            repeatRows=0
        )

        table.setStyle(
            TableStyle(
                [

                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.4,
                        colors.HexColor(
                            "#d1d5db"
                        )
                    ),

                    (
                        "BACKGROUND",
                        (0, 0),
                        (0, -1),
                        colors.HexColor(
                            "#f3f4f6"
                        )
                    ),

                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP"
                    ),

                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    ),

                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    ),

                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    ),

                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ]
            )
        )

        story.append(
            table
        )

        story.append(
            Spacer(
                1,
                5 * mm
            )
        )

    # =========================================================
    # APPLICATION INFORMATION
    # =========================================================

    section_header(
        "APPLICATION INFORMATION"
    )

    created_at = value(
        "created_at",
        ""
    )

    if created_at:

        try:

            formatted_created_at = format_date(
                created_at
            )

        except Exception:

            formatted_created_at = created_at

    else:

        formatted_created_at = "—"

    information_table(
        [

            (
                "Application Number",
                value(
                    "application_number"
                )
            ),

            (
                "Position Applied For",
                value(
                    "position_applied"
                )
            ),

            (
                "Application Date",
                formatted_created_at
            ),

            (
                "Application Status",
                value(
                    "status",
                    "Pending"
                )
            )
        ]
    )

    # =========================================================
    # PERSONAL INFORMATION
    # =========================================================

    section_header(
        "PERSONAL INFORMATION"
    )

    first_name = value(
        "first_name",
        ""
    )

    middle_name = value(
        "middle_name",
        ""
    )

    last_name = value(
        "last_name",
        ""
    )

    name_parts = []

    for part in [
        first_name,
        middle_name,
        last_name
    ]:

        if part and part != "—":

            name_parts.append(
                part.strip()
            )

    full_name = " ".join(
        name_parts
    )

    if not full_name:

        full_name = "—"

    information_table(
        [

            (
                "Full Name",
                full_name
            ),

            (
                "Gender",
                value(
                    "gender"
                )
            ),

            (
                "Date of Birth",
                value(
                    "date_of_birth"
                )
            ),

            (
                "Phone Number",
                value(
                    "phone"
                )
            ),

            (
                "Email Address",
                value(
                    "email"
                )
            ),

            (
                "Residential Address",
                value(
                    "address"
                )
            ),

            (
                "State",
                value(
                    "state"
                )
            ),

            (
                "Local Government Area",
                value(
                    "lga"
                )
            )
        ]
    )

    # =========================================================
    # EDUCATION & QUALIFICATION
    # =========================================================

    section_header(
        "EDUCATION & QUALIFICATION"
    )

    information_table(
        [

            (
                "Highest Qualification",
                value(
                    "highest_qualification"
                )
            ),

            (
                "Course of Study",
                value(
                    "course_of_study"
                )
            ),

            (
                "Institution",
                value(
                    "institution"
                )
            ),

            (
                "Graduation Year",
                value(
                    "graduation_year"
                )
            )
        ]
    )

    # =========================================================
    # WORK EXPERIENCE
    # =========================================================

    section_header(
        "WORK EXPERIENCE"
    )

    information_table(
        [

            (
                "Previous Employer",
                value(
                    "previous_employer"
                )
            ),

            (
                "Previous Position",
                value(
                    "previous_position"
                )
            ),

            (
                "Work Experience",
                value(
                    "work_experience"
                )
            )
        ]
    )

    # =========================================================
    # APPLICATION STATEMENT
    # =========================================================

    section_header(
        "APPLICATION STATEMENT"
    )

    information_table(
        [

            (
                "Reason for Applying",
                value(
                    "reason_for_applying"
                )
            ),

            (
                "Additional Information",
                value(
                    "additional_information"
                )
            )
        ]
    )

    # =========================================================
    # SUBMITTED DOCUMENTS
    # =========================================================

    section_header(
        "SUBMITTED DOCUMENTS"
    )

    information_table(
        [

            (
                "Passport Photograph",
                value(
                    "passport_filename"
                )
            ),

            (
                "Curriculum Vitae",
                value(
                    "cv_filename"
                )
            ),

            (
                "Qualification Document",
                value(
                    "qualification_filename"
                )
            )
        ]
    )

    # =========================================================
    # DECLARATION
    # =========================================================

    section_header(
        "APPLICANT DECLARATION"
    )

    declaration_text = (
        "I confirm that the information provided in this "
        "application is true and correct to the best of my "
        "knowledge. I understand that any false or misleading "
        "information may affect the consideration of my "
        "application."
    )

    declaration_table = Table(

        [
            [
                Paragraph(
                    declaration_text,
                    data_style
                )
            ]
        ],

        colWidths=[
            180 * mm
        ]
    )

    declaration_table.setStyle(
        TableStyle(
            [

                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#d1d5db"
                    )
                ),

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    colors.HexColor(
                        "#f9fafb"
                    )
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                )
            ]
        )
    )

    story.append(
        declaration_table
    )

    story.append(
        Spacer(
            1,
            12 * mm
        )
    )

    # =========================================================
    # OFFICIAL USE
    # =========================================================

    section_header(
        "FOR OFFICIAL USE ONLY"
    )

    official_use_table = Table(

        [

            [

                Paragraph(
                    "<b>Application Reviewed By:</b>",
                    label_style
                ),

                Paragraph(
                    "________________________________________",
                    data_style
                )
            ],

            [

                Paragraph(
                    "<b>Review Date:</b>",
                    label_style
                ),

                Paragraph(
                    "________________________________________",
                    data_style
                )
            ],

            [

                Paragraph(
                    "<b>Decision:</b>",
                    label_style
                ),

                Paragraph(
                    "Approved     Shortlisted     "
                    "Rejected     Pending",
                    data_style
                )
            ],

            [

                Paragraph(
                    "<b>Comments:</b>",
                    label_style
                ),

                Paragraph(
                    "________________________________________"
                    "<br/>"
                    "________________________________________"
                    "<br/>"
                    "________________________________________",
                    data_style
                )
            ]
        ],

        colWidths=[
            48 * mm,
            132 * mm
        ]
    )

    official_use_table.setStyle(
        TableStyle(
            [

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#d1d5db"
                    )
                ),

                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor(
                        "#f3f4f6"
                    )
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7
                )
            ]
        )
    )

    story.append(
        official_use_table
    )

    story.append(
        Spacer(
            1,
            10 * mm
        )
    )

    # =========================================================
    # FOOTER
    # =========================================================

    story.append(
        HRFlowable(
            width="100%",

            thickness=0.7,

            color=colors.HexColor(
                "#d1d5db"
            )
        )
    )

    story.append(
        Spacer(
            1,
            3 * mm
        )
    )

    if footer_text:

        footer_content = str(
            footer_text
        )

    else:

        footer_content = (
            f"{company_name} — "
            "Applicant Recruitment Portal"
        )

    story.append(
        Paragraph(
            safe_text(
                footer_content
            ),
            company_info_style
        )
    )

    story.append(
        Spacer(
            1,
            2 * mm
        )
    )

    story.append(
        Paragraph(
            "This document was generated electronically "
            "from the recruitment portal.",
            small_style
        )
    )

    # =========================================================
    # BUILD PDF
    # =========================================================

    try:

        document.build(
            story
        )

    except Exception:

        app.logger.exception(
            "Error generating applicant biodata PDF "
            "for application ID %s",
            application_id
        )

        flash(
            "Unable to generate applicant biodata PDF.",
            "error"
        )

        return redirect(
            url_for(
                "admin_applications"
            )
        )

    # =========================================================
    # PREPARE PDF
    # =========================================================

    buffer.seek(0)

    # =========================================================
    # SAFE APPLICATION NUMBER
    # =========================================================

    safe_application_number = value(
        "application_number",
        f"APP-{application_id}"
    )

    safe_application_number = (
        safe_application_number
        .replace(
            "/",
            "-"
        )
        .replace(
            "\\",
            "-"
        )
        .replace(
            " ",
            "_"
        )
    )

    # =========================================================
    # FILE NAME
    # =========================================================

    filename = (
        f"{safe_application_number}_"
        f"Applicant_Biodata.pdf"
    )

    # =========================================================
    # RETURN PDF
    # =========================================================

    return send_file(

        buffer,

        mimetype="application/pdf",

        as_attachment=True,

        download_name=filename
    )
# ============================================================
# APPLICANT ATTENDANCE
# Applicant clocks in/out from their own portal
# Location must be within the company's configured radius
# ============================================================

from datetime import datetime, date
import math


# ============================================================
# HELPER: DISTANCE BETWEEN TWO GPS COORDINATES
# Returns distance in metres
# ============================================================
import math


def calculate_distance_meters(
    lat1,
    lon1,
    lat2,
    lon2
):
    """
    Calculate the straight-line distance between
    two GPS coordinates in metres.
    """

    lat1 = float(lat1)
    lon1 = float(lon1)
    lat2 = float(lat2)
    lon2 = float(lon2)

    earth_radius = 6371000

    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)

    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)

    a = (
        math.sin(delta_lat / 2) ** 2
        +
        math.cos(lat1_rad)
        * math.cos(lat2_rad)
        * math.sin(delta_lon / 2) ** 2
    )

    # Protect against tiny floating-point errors
    a = max(0, min(1, a))

    c = 2 * math.atan2(
        math.sqrt(a),
        math.sqrt(1 - a)
    )

    return earth_radius * c


# ============================================================
# HELPER: GET TODAY'S ATTENDANCE
# ============================================================

def get_today_attendance(
    worker_id
):

    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT *
                FROM attendance

                WHERE worker_id = %s
                AND attendance_date = CURRENT_DATE

                LIMIT 1
                """,
                (worker_id,)
            )

            return cur.fetchone()

    finally:

        conn.close()


# ============================================================
# APPLICANT ATTENDANCE PAGE
# ============================================================
# =========================================================
# APPLICANT ATTENDANCE PAGE
# =========================================================
# =========================================================
# APPLICANT ATTENDANCE PAGE
# =========================================================

@app.route("/applicant/attendance")
def applicant_attendance():

    from datetime import datetime, time

    # =====================================================
    # NIGERIA TIMEZONE
    # =====================================================

    nigeria_now = get_nigeria_now()
    nigeria_today = nigeria_now.date()

    # =====================================================
    # AUTHENTICATION
    # =====================================================

    applicant_id = session.get("applicant_id")

    if not applicant_id:

        return redirect(
            url_for("applicant_login")
        )

    # =====================================================
    # DEFAULT VALUES
    #
    # Prevent variables from being undefined if a
    # database error occurs.
    # =====================================================

    conn = None

    applicant = None
    settings = None
    today_attendance = None
    attendance_history = []

    # =====================================================
    # DATABASE
    # =====================================================

    try:

        conn = get_db()

        with conn.cursor() as cur:

            # =================================================
            # GET APPLICANT
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    application_number,
                    first_name,
                    middle_name,
                    last_name,
                    position_applied,
                    status,
                    portal_active

                FROM applications

                WHERE id = %s

                LIMIT 1
                """,
                (
                    applicant_id,
                )
            )

            applicant = cur.fetchone()

            # =================================================
            # APPLICANT NOT FOUND
            # =================================================

            if not applicant:

                session.clear()

                flash(
                    "Applicant account not found.",
                    "error"
                )

                return redirect(
                    url_for("applicant_login")
                )

            # =================================================
            # APPROVAL CHECK
            # =================================================

            applicant_status = str(
                applicant["status"] or ""
            ).strip().lower()

            if applicant_status != "approved":

                flash(
                    "Attendance is available only to approved applicants.",
                    "error"
                )

                return redirect(
                    url_for("applicant_portal")
                )

            # =================================================
            # PORTAL ACTIVE CHECK
            # =================================================

            portal_active = applicant["portal_active"]

            if not portal_active:

                flash(
                    "Your applicant portal has been disabled.",
                    "error"
                )

                session.clear()

                return redirect(
                    url_for("applicant_login")
                )

            # =================================================
            # GET ATTENDANCE SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT
                    id,

                    attendance_enabled,

                    company_latitude,
                    company_longitude,

                    attendance_radius,

                    clock_in_start,
                    clock_in_end,

                    clock_out_start,
                    clock_out_end,

                    late_after_minutes,
                    early_before_minutes

                FROM company_settings

                ORDER BY id ASC

                LIMIT 1
                """
            )

            settings = cur.fetchone()

            # =================================================
            # TODAY'S ATTENDANCE
            #
            # IMPORTANT:
            # Do NOT use CURRENT_DATE because PostgreSQL's
            # timezone may not be Africa/Lagos.
            #
            # We explicitly use Nigeria's date.
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    worker_id,
                    attendance_date,

                    clock_in,
                    clock_out,

                    total_hours,

                    status,

                    clock_in_latitude,
                    clock_in_longitude,

                    clock_out_latitude,
                    clock_out_longitude,

                    clock_in_location_verified,
                    clock_out_location_verified,

                    created_at,
                    updated_at

                FROM attendance

                WHERE worker_id = %s

                AND attendance_date = %s

                ORDER BY id DESC

                LIMIT 1
                """,
                (
                    applicant_id,
                    nigeria_today
                )
            )

            today_attendance = cur.fetchone()

            # =================================================
            # ATTENDANCE HISTORY
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    attendance_date,

                    clock_in,
                    clock_out,

                    total_hours,

                    status,

                    clock_in_location_verified,
                    clock_out_location_verified

                FROM attendance

                WHERE worker_id = %s

                ORDER BY
                    attendance_date DESC,
                    id DESC

                LIMIT 30
                """,
                (
                    applicant_id,
                )
            )

            attendance_history = cur.fetchall()

    except Exception:

        app.logger.exception(
            "Error loading applicant attendance"
        )

        flash(
            "Unable to load attendance information. Please try again.",
            "error"
        )

    finally:

        if conn is not None:

            try:
                conn.close()
            except Exception:
                pass

    # =====================================================
    # RENDER
    # =====================================================

    return render_template(
        "applicant_attendance.html",

        application=applicant,

        applicant=applicant,

        attendance=today_attendance,

        today_attendance=today_attendance,

        settings=settings,

        attendance_history=attendance_history,

        nigeria_now=nigeria_now,

        nigeria_today=nigeria_today
    )


# =========================================================
# APPLICANT CLOCK IN
# =========================================================

@app.route(
    "/applicant/attendance/clock-in",
    methods=["POST"]
)
def applicant_clock_in():

    from datetime import time

    # =====================================================
    # AUTHENTICATION
    # =====================================================

    applicant_id = session.get("applicant_id")

    if not applicant_id:

        return jsonify({
            "success": False,
            "message": (
                "Your applicant session has expired. "
                "Please log in again."
            )
        }), 401

    # =====================================================
    # NIGERIA TIME
    # =====================================================

    nigeria_now = get_nigeria_now()

    nigeria_date = nigeria_now.date()

    nigeria_time = nigeria_now.time()

    # =====================================================
    # GPS INPUT
    # =====================================================

    latitude_raw = (
        request.form.get(
            "latitude",
            ""
        ) or ""
    ).strip()

    longitude_raw = (
        request.form.get(
            "longitude",
            ""
        ) or ""
    ).strip()

    # =====================================================
    # CONVERT GPS
    # =====================================================

    try:

        latitude = float(
            latitude_raw
        )

        longitude = float(
            longitude_raw
        )

    except (
        ValueError,
        TypeError
    ):

        return jsonify({
            "success": False,
            "message": (
                "Unable to determine your location. "
                "Please allow location access and try again."
            )
        }), 400

    # =====================================================
    # VALIDATE GPS
    # =====================================================

    if not (-90 <= latitude <= 90):

        return jsonify({
            "success": False,
            "message": (
                "Invalid latitude received from your device."
            )
        }), 400

    if not (-180 <= longitude <= 180):

        return jsonify({
            "success": False,
            "message": (
                "Invalid longitude received from your device."
            )
        }), 400

    # =====================================================
    # DATABASE
    # =====================================================

    conn = None

    try:

        conn = get_db()

        with conn.cursor() as cur:

            # =================================================
            # GET APPLICANT
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    first_name,
                    last_name,
                    status,
                    portal_active

                FROM applications

                WHERE id = %s

                LIMIT 1
                """,
                (
                    applicant_id,
                )
            )

            applicant = cur.fetchone()

            # =================================================
            # APPLICANT NOT FOUND
            # =================================================

            if not applicant:

                return jsonify({
                    "success": False,
                    "message": "Applicant account not found."
                }), 404

            # =================================================
            # APPROVAL
            # =================================================

            applicant_status = str(
                applicant["status"] or ""
            ).strip().lower()

            if applicant_status != "approved":

                return jsonify({
                    "success": False,
                    "message": (
                        "Only approved applicants can "
                        "clock attendance."
                    )
                }), 403

            # =================================================
            # PORTAL ACTIVE
            # =================================================

            if not applicant["portal_active"]:

                return jsonify({
                    "success": False,
                    "message": (
                        "Your applicant portal has been disabled."
                    )
                }), 403

            # =================================================
            # GET SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT
                    attendance_enabled,

                    company_latitude,
                    company_longitude,

                    attendance_radius,

                    clock_in_start,
                    clock_in_end,

                    clock_out_start,
                    clock_out_end,

                    late_after_minutes,
                    early_before_minutes

                FROM company_settings

                ORDER BY id ASC

                LIMIT 1
                """
            )

            settings = cur.fetchone()

            if not settings:

                return jsonify({
                    "success": False,
                    "message": (
                        "Attendance settings have not "
                        "been configured."
                    )
                }), 400

            # =================================================
            # ATTENDANCE ENABLED
            # =================================================

            if not settings["attendance_enabled"]:

                return jsonify({
                    "success": False,
                    "message": (
                        "Attendance is currently disabled "
                        "by the administrator."
                    )
                }), 403

            # =================================================
            # COMPANY LOCATION
            # =================================================

            company_latitude = (
                settings["company_latitude"]
            )

            company_longitude = (
                settings["company_longitude"]
            )

            if (
                company_latitude is None
                or company_longitude is None
            ):

                return jsonify({
                    "success": False,
                    "message": (
                        "The company attendance location "
                        "has not been configured."
                    )
                }), 400

            # =================================================
            # ATTENDANCE RADIUS
            # =================================================

            try:

                radius = int(
                    settings["attendance_radius"]
                    or 200
                )

            except (
                ValueError,
                TypeError
            ):

                radius = 200

            if radius <= 0:

                radius = 200

            # =================================================
            # CLOCK-IN WINDOW
            # =================================================

            clock_in_start = normalize_db_time(
                settings["clock_in_start"],
                time(6, 0)
            )

            clock_in_end = normalize_db_time(
                settings["clock_in_end"],
                time(10, 0)
            )

            # =================================================
            # CHECK CLOCK-IN WINDOW
            # =================================================

            if not is_time_in_window(
                nigeria_time,
                clock_in_start,
                clock_in_end
            ):

                return jsonify({
                    "success": False,

                    "message": (
                        "Clock-in is not currently available. "
                        f"Clock-in window is "
                        f"{clock_in_start.strftime('%I:%M %p')} "
                        f"to "
                        f"{clock_in_end.strftime('%I:%M %p')}."
                    ),

                    "current_time": (
                        nigeria_now.strftime(
                            "%I:%M:%S %p"
                        )
                    ),

                    "clock_in_start": (
                        clock_in_start.strftime(
                            "%I:%M %p"
                        )
                    ),

                    "clock_in_end": (
                        clock_in_end.strftime(
                            "%I:%M %p"
                        )
                    ),

                    "timezone": "Africa/Lagos"

                }), 403

            # =================================================
            # CALCULATE GPS DISTANCE
            # =================================================

            distance = calculate_distance_meters(
                latitude,
                longitude,
                float(company_latitude),
                float(company_longitude)
            )

            # =================================================
            # LOCATION CHECK
            # =================================================

            if distance > radius:

                return jsonify({
                    "success": False,

                    "message": (
                        "You are outside the company "
                        "attendance area. "
                        f"Distance: {round(distance)} metres. "
                        f"Allowed radius: {radius} metres."
                    ),

                    "distance": round(distance),

                    "allowed_radius": radius,

                    "timezone": "Africa/Lagos"

                }), 403

            # =================================================
            # FIND TODAY'S RECORD
            #
            # IMPORTANT:
            # Use Nigeria date rather than CURRENT_DATE.
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    clock_in,
                    clock_out,
                    status

                FROM attendance

                WHERE worker_id = %s

                AND attendance_date = %s

                ORDER BY id DESC

                LIMIT 1
                """,
                (
                    applicant_id,
                    nigeria_date
                )
            )

            existing = cur.fetchone()

            # =================================================
            # ALREADY CLOCKED IN
            # =================================================

            if (
                existing
                and existing["clock_in"] is not None
            ):

                existing_time = (
                    existing["clock_in"]
                )

                if hasattr(
                    existing_time,
                    "strftime"
                ):

                    formatted_time = (
                        existing_time.strftime(
                            "%I:%M %p"
                        )
                    )

                else:

                    formatted_time = str(
                        existing_time
                    )

                return jsonify({
                    "success": False,

                    "already_clocked_in": True,

                    "message": (
                        "You have already clocked in today."
                    ),

                    "clock_in": formatted_time,

                    "timezone": "Africa/Lagos"

                }), 409

            # =================================================
            # CALCULATE LATE STATUS
            # =================================================

            try:

                late_after_minutes = int(
                    settings["late_after_minutes"]
                    or 15
                )

            except (
                ValueError,
                TypeError
            ):

                late_after_minutes = 15

            if late_after_minutes < 0:

                late_after_minutes = 15

            late_minutes = calculate_lateness_minutes(
                nigeria_time,
                clock_in_start
            )

            if late_minutes >= late_after_minutes:

                attendance_status = "Late"

            else:

                attendance_status = "Present"

            # =================================================
            # DATABASE TIMESTAMP
            #
            # Store Nigeria wall-clock time as naive
            # timestamp because the attendance columns
            # are currently TIMESTAMP.
            # =================================================

            db_clock_in = (
                nigeria_now.replace(
                    tzinfo=None
                )
            )

            # =================================================
            # SAVE ATTENDANCE
            # =================================================

            if existing:

                cur.execute(
                    """
                    UPDATE attendance

                    SET
                        clock_in = %s,

                        clock_in_latitude = %s,

                        clock_in_longitude = %s,

                        clock_in_location_verified = TRUE,

                        status = %s,

                        updated_at = %s

                    WHERE id = %s
                    """,
                    (
                        db_clock_in,

                        latitude,
                        longitude,

                        attendance_status,

                        db_clock_in,

                        existing["id"]
                    )
                )

            else:

                cur.execute(
                    """
                    INSERT INTO attendance
                    (
                        worker_id,
                        attendance_date,

                        clock_in,

                        clock_in_latitude,
                        clock_in_longitude,

                        clock_in_location_verified,

                        status,

                        created_at,
                        updated_at
                    )

                    VALUES
                    (
                        %s,
                        %s,

                        %s,

                        %s,
                        %s,

                        TRUE,

                        %s,

                        %s,
                        %s
                    )
                    """,
                    (
                        applicant_id,
                        nigeria_date,

                        db_clock_in,

                        latitude,
                        longitude,

                        attendance_status,

                        db_clock_in,
                        db_clock_in
                    )
                )

            # =================================================
            # COMMIT
            # =================================================

            conn.commit()

            # =================================================
            # SUCCESS
            # =================================================

            return jsonify({

                "success": True,

                "message": (
                    "Clock-in recorded successfully."
                ),

                "clock_in": (
                    nigeria_now.strftime(
                        "%I:%M %p"
                    )
                ),

                "status": attendance_status,

                "clock_in_location_verified": True,

                "distance": round(
                    distance
                ),

                "allowed_radius": radius,

                "current_time": (
                    nigeria_now.strftime(
                        "%I:%M:%S %p"
                    )
                ),

                "date": nigeria_date.strftime(
                    "%Y-%m-%d"
                ),

                "timezone": "Africa/Lagos"

            }), 200

    except Exception:

        if conn is not None:

            try:
                conn.rollback()
            except Exception:
                pass

        app.logger.exception(
            "Applicant clock-in error"
        )

        return jsonify({
            "success": False,
            "message": (
                "Unable to record clock-in. "
                "Please try again."
            )
        }), 500

    finally:

        if conn is not None:

            try:
                conn.close()
            except Exception:
                pass


# =========================================================
# APPLICANT CLOCK OUT
# =========================================================

@app.route(
    "/applicant/attendance/clock-out",
    methods=["POST"]
)
def applicant_clock_out():

    from datetime import datetime, time

    # =====================================================
    # AUTHENTICATION
    # =====================================================

    applicant_id = session.get(
        "applicant_id"
    )

    if not applicant_id:

        return jsonify({
            "success": False,
            "message": (
                "Your applicant session has expired. "
                "Please log in again."
            )
        }), 401

    # =====================================================
    # NIGERIA TIME
    # =====================================================

    nigeria_now = get_nigeria_now()

    nigeria_date = nigeria_now.date()

    nigeria_time = nigeria_now.time()

    # =====================================================
    # GPS INPUT
    # =====================================================

    latitude_raw = (
        request.form.get(
            "latitude",
            ""
        ) or ""
    ).strip()

    longitude_raw = (
        request.form.get(
            "longitude",
            ""
        ) or ""
    ).strip()

    # =====================================================
    # CONVERT GPS
    # =====================================================

    try:

        latitude = float(
            latitude_raw
        )

        longitude = float(
            longitude_raw
        )

    except (
        ValueError,
        TypeError
    ):

        return jsonify({
            "success": False,
            "message": (
                "Unable to determine your location. "
                "Please allow location access and try again."
            )
        }), 400

    # =====================================================
    # VALIDATE GPS
    # =====================================================

    if not (-90 <= latitude <= 90):

        return jsonify({
            "success": False,
            "message": (
                "Invalid latitude received from your device."
            )
        }), 400

    if not (-180 <= longitude <= 180):

        return jsonify({
            "success": False,
            "message": (
                "Invalid longitude received from your device."
            )
        }), 400

    # =====================================================
    # DATABASE
    # =====================================================

    conn = None

    try:

        conn = get_db()

        with conn.cursor() as cur:

            # =================================================
            # GET APPLICANT
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    first_name,
                    last_name,
                    status,
                    portal_active

                FROM applications

                WHERE id = %s

                LIMIT 1
                """,
                (
                    applicant_id,
                )
            )

            applicant = cur.fetchone()

            # =================================================
            # APPLICANT NOT FOUND
            # =================================================

            if not applicant:

                return jsonify({
                    "success": False,
                    "message": "Applicant account not found."
                }), 404

            # =================================================
            # APPROVAL
            # =================================================

            applicant_status = str(
                applicant["status"] or ""
            ).strip().lower()

            if applicant_status != "approved":

                return jsonify({
                    "success": False,
                    "message": (
                        "Only approved applicants can "
                        "clock attendance."
                    )
                }), 403

            # =================================================
            # PORTAL ACTIVE
            # =================================================

            if not applicant["portal_active"]:

                return jsonify({
                    "success": False,
                    "message": (
                        "Your applicant portal has been disabled."
                    )
                }), 403

            # =================================================
            # GET SETTINGS
            # =================================================

            cur.execute(
                """
                SELECT
                    attendance_enabled,

                    company_latitude,
                    company_longitude,

                    attendance_radius,

                    clock_in_start,
                    clock_in_end,

                    clock_out_start,
                    clock_out_end,

                    late_after_minutes,
                    early_before_minutes

                FROM company_settings

                ORDER BY id ASC

                LIMIT 1
                """
            )

            settings = cur.fetchone()

            if not settings:

                return jsonify({
                    "success": False,
                    "message": (
                        "Attendance settings have not "
                        "been configured."
                    )
                }), 400

            # =================================================
            # ATTENDANCE ENABLED
            # =================================================

            if not settings["attendance_enabled"]:

                return jsonify({
                    "success": False,
                    "message": (
                        "Attendance is currently disabled "
                        "by the administrator."
                    )
                }), 403

            # =================================================
            # COMPANY LOCATION
            # =================================================

            company_latitude = (
                settings["company_latitude"]
            )

            company_longitude = (
                settings["company_longitude"]
            )

            if (
                company_latitude is None
                or company_longitude is None
            ):

                return jsonify({
                    "success": False,
                    "message": (
                        "The company attendance location "
                        "has not been configured."
                    )
                }), 400

            # =================================================
            # RADIUS
            # =================================================

            try:

                radius = int(
                    settings["attendance_radius"]
                    or 200
                )

            except (
                ValueError,
                TypeError
            ):

                radius = 200

            if radius <= 0:

                radius = 200

            # =================================================
            # CLOCK-OUT WINDOW
            # =================================================

            clock_out_start = normalize_db_time(
                settings["clock_out_start"],
                time(15, 0)
            )

            clock_out_end = normalize_db_time(
                settings["clock_out_end"],
                time(23, 0)
            )

            # =================================================
            # CHECK CLOCK-OUT WINDOW
            # =================================================

            if not is_time_in_window(
                nigeria_time,
                clock_out_start,
                clock_out_end
            ):

                return jsonify({
                    "success": False,

                    "message": (
                        "Clock-out is not currently available. "
                        f"Clock-out window is "
                        f"{clock_out_start.strftime('%I:%M %p')} "
                        f"to "
                        f"{clock_out_end.strftime('%I:%M %p')}."
                    ),

                    "current_time": (
                        nigeria_now.strftime(
                            "%I:%M:%S %p"
                        )
                    ),

                    "clock_out_start": (
                        clock_out_start.strftime(
                            "%I:%M %p"
                        )
                    ),

                    "clock_out_end": (
                        clock_out_end.strftime(
                            "%I:%M %p"
                        )
                    ),

                    "timezone": "Africa/Lagos"

                }), 403

            # =================================================
            # GET TODAY'S ATTENDANCE
            #
            # Again, use Nigeria's date rather than
            # PostgreSQL CURRENT_DATE.
            # =================================================

            cur.execute(
                """
                SELECT
                    id,
                    clock_in,
                    clock_out,
                    status

                FROM attendance

                WHERE worker_id = %s

                AND attendance_date = %s

                ORDER BY id DESC

                LIMIT 1
                """,
                (
                    applicant_id,
                    nigeria_date
                )
            )

            attendance = cur.fetchone()

            # =================================================
            # NO ATTENDANCE
            # =================================================

            if not attendance:

                return jsonify({
                    "success": False,
                    "message": (
                        "You cannot clock out because "
                        "you have not clocked in today."
                    )
                }), 400

            # =================================================
            # NO CLOCK-IN
            # =================================================

            if attendance["clock_in"] is None:

                return jsonify({
                    "success": False,
                    "message": (
                        "You cannot clock out because "
                        "you have not clocked in today."
                    )
                }), 400

            # =================================================
            # ALREADY CLOCKED OUT
            # =================================================

            if attendance["clock_out"] is not None:

                existing_clock_out = (
                    attendance["clock_out"]
                )

                if hasattr(
                    existing_clock_out,
                    "strftime"
                ):

                    formatted = (
                        existing_clock_out.strftime(
                            "%I:%M %p"
                        )
                    )

                else:

                    formatted = str(
                        existing_clock_out
                    )

                return jsonify({
                    "success": False,

                    "already_clocked_out": True,

                    "message": (
                        "You have already clocked out today."
                    ),

                    "clock_out": formatted,

                    "timezone": "Africa/Lagos"

                }), 409

            # =================================================
            # GPS DISTANCE
            # =================================================

            distance = calculate_distance_meters(
                latitude,
                longitude,
                float(company_latitude),
                float(company_longitude)
            )

            # =================================================
            # LOCATION CHECK
            # =================================================

            if distance > radius:

                return jsonify({
                    "success": False,

                    "message": (
                        "You are outside the company "
                        "attendance area. "
                        f"Distance: {round(distance)} metres. "
                        f"Allowed radius: {radius} metres."
                    ),

                    "distance": round(distance),

                    "allowed_radius": radius,

                    "timezone": "Africa/Lagos"

                }), 403

            # =================================================
            # EARLY CLOCK-OUT SETTING
            # =================================================

            try:

                early_before_minutes = int(
                    settings["early_before_minutes"]
                    or 15
                )

            except (
                ValueError,
                TypeError
            ):

                early_before_minutes = 15

            if early_before_minutes < 0:

                early_before_minutes = 15

            # =================================================
            # CALCULATE EARLY DEPARTURE
            # =================================================

            early_minutes = (
                calculate_early_clockout_minutes(
                    nigeria_time,
                    clock_out_end
                )
            )

            # =================================================
            # PRESERVE EXISTING STATUS
            #
            # This fixes an important issue in the original
            # route:
            #
            # If an applicant clocked in late, clocking out
            # should not automatically change "Late" to
            # "Present".
            # =================================================

            existing_status = str(
                attendance["status"] or ""
            ).strip()

            if early_minutes >= early_before_minutes:

                current_status = "Early Departure"

            elif existing_status:

                current_status = existing_status

            else:

                current_status = "Present"

            # =================================================
            # DATABASE CLOCK-OUT TIME
            # =================================================

            db_clock_out = (
                nigeria_now.replace(
                    tzinfo=None
                )
            )

            # =================================================
            # CONVERT CLOCK-IN VALUE
            # =================================================

            clock_in_value = (
                attendance["clock_in"]
            )

            # -------------------------------------------------
            # CASE 1:
            # PostgreSQL returned datetime
            # -------------------------------------------------

            if isinstance(
                clock_in_value,
                datetime
            ):

                clock_in_naive = (
                    clock_in_value.replace(
                        tzinfo=None
                    )
                )

            # -------------------------------------------------
            # CASE 2:
            # PostgreSQL returned time
            # -------------------------------------------------

            elif isinstance(
                clock_in_value,
                time
            ):

                clock_in_naive = datetime.combine(
                    nigeria_date,
                    clock_in_value
                )

            # -------------------------------------------------
            # CASE 3:
            # PostgreSQL returned string
            # -------------------------------------------------

            else:

                clock_in_text = str(
                    clock_in_value
                ).strip()

                clock_in_naive = None

                # Try common datetime formats
                datetime_formats = [
                    "%Y-%m-%d %H:%M:%S.%f",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%d %H:%M"
                ]

                for fmt in datetime_formats:

                    try:

                        clock_in_naive = (
                            datetime.strptime(
                                clock_in_text,
                                fmt
                            )
                        )

                        break

                    except ValueError:

                        continue

                # If it was only a time
                if clock_in_naive is None:

                    time_formats = [
                        "%H:%M:%S.%f",
                        "%H:%M:%S",
                        "%H:%M"
                    ]

                    for fmt in time_formats:

                        try:

                            parsed_time = (
                                datetime.strptime(
                                    clock_in_text,
                                    fmt
                                ).time()
                            )

                            clock_in_naive = (
                                datetime.combine(
                                    nigeria_date,
                                    parsed_time
                                )
                            )

                            break

                        except ValueError:

                            continue

                # -------------------------------------------------
                # INVALID CLOCK-IN VALUE
                # -------------------------------------------------

                if clock_in_naive is None:

                    return jsonify({
                        "success": False,
                        "message": (
                            "The clock-in time could not "
                            "be processed. Please contact "
                            "the administrator."
                        )
                    }), 500

            # =================================================
            # CALCULATE TOTAL HOURS
            # =================================================

            total_seconds = (
                db_clock_out
                - clock_in_naive
            ).total_seconds()

            # =================================================
            # INVALID NEGATIVE TIME PROTECTION
            # =================================================

            if total_seconds < 0:

                return jsonify({
                    "success": False,
                    "message": (
                        "The clock-out time appears to be "
                        "earlier than the clock-in time. "
                        "Please contact the administrator."
                    )
                }), 400

            # =================================================
            # TOTAL HOURS
            # =================================================

            total_hours = (
                total_seconds / 3600
            )

            total_hours = round(
                total_hours,
                2
            )

            # =================================================
            # UPDATE ATTENDANCE
            # =================================================

            cur.execute(
                """
                UPDATE attendance

                SET
                    clock_out = %s,

                    clock_out_latitude = %s,

                    clock_out_longitude = %s,

                    clock_out_location_verified = TRUE,

                    total_hours = %s,

                    status = %s,

                    updated_at = %s

                WHERE id = %s
                """,
                (
                    db_clock_out,

                    latitude,
                    longitude,

                    total_hours,

                    current_status,

                    db_clock_out,

                    attendance["id"]
                )
            )

            # =================================================
            # VERIFY UPDATE
            # =================================================

            if cur.rowcount != 1:

                conn.rollback()

                return jsonify({
                    "success": False,
                    "message": (
                        "The attendance record could not "
                        "be updated. Please try again."
                    )
                }), 500

            # =================================================
            # COMMIT
            # =================================================

            conn.commit()

            # =================================================
            # SUCCESS
            # =================================================

            return jsonify({

                "success": True,

                "message": (
                    "Clock-out recorded successfully."
                ),

                "clock_out": (
                    nigeria_now.strftime(
                        "%I:%M %p"
                    )
                ),

                "total_hours": total_hours,

                "status": current_status,

                "clock_out_location_verified": True,

                "distance": round(
                    distance
                ),

                "allowed_radius": radius,

                "current_time": (
                    nigeria_now.strftime(
                        "%I:%M:%S %p"
                    )
                ),

                "date": nigeria_date.strftime(
                    "%Y-%m-%d"
                ),

                "timezone": "Africa/Lagos"

            }), 200

    except Exception:

        if conn is not None:

            try:
                conn.rollback()
            except Exception:
                pass

        app.logger.exception(
            "Applicant clock-out error"
        )

        return jsonify({
            "success": False,
            "message": (
                "Unable to record clock-out. "
                "Please try again."
            )
        }), 500

    finally:

        if conn is not None:

            try:
                conn.close()
            except Exception:
                pass
# ============================================================
# APPLICANT ATTENDANCE HISTORY
# ============================================================

@app.route("/applicant/attendance/history")
def applicant_attendance_history():

    applicant_id = session.get(
        "applicant_id"
    )

    if not applicant_id:

        return redirect(
            url_for("applicant_login")
        )


    conn = get_db()

    try:

        with conn.cursor() as cur:

            cur.execute(
                """
                SELECT
                    attendance_date,
                    clock_in,
                    clock_out,
                    total_hours,
                    status,
                    clock_in_location_verified,
                    clock_out_location_verified

                FROM attendance

                WHERE worker_id = %s

                ORDER BY attendance_date DESC

                """,
                (applicant_id,)
            )

            attendance_history = cur.fetchall()


    finally:

        conn.close()


    return render_template(
        "applicant_attendance_history.html",

        attendance_history=attendance_history
    )

# =========================================================
# CHECK WHETHER A TIME IS WITHIN AN ATTENDANCE WINDOW
# =========================================================

def is_time_in_window(current_time, start_time, end_time):
    """
    Return True when current_time falls within the
    configured attendance window.

    Supports normal windows such as:
        06:00 AM - 10:00 AM

    Also supports windows that cross midnight, such as:
        10:00 PM - 02:00 AM
    """

    if current_time is None:
        return False

    if start_time is None or end_time is None:
        return False

    # -----------------------------------------------------
    # NORMAL SAME-DAY WINDOW
    # Example: 06:00 -> 10:00
    # -----------------------------------------------------

    if start_time <= end_time:

        return (
            start_time
            <= current_time
            <= end_time
        )

    # -----------------------------------------------------
    # OVERNIGHT WINDOW
    # Example: 22:00 -> 02:00
    # -----------------------------------------------------

    return (
        current_time >= start_time
        or
        current_time <= end_time
    )

# =========================================================
# CALCULATE CLOCK-IN LATENESS
# =========================================================

def calculate_lateness_minutes(
    current_time,
    scheduled_start
):
    """
    Calculate how many minutes the user is late
    compared with the configured clock-in start time.

    Example:
        Scheduled: 08:00 AM
        Actual:    08:15 AM
        Result:    15 minutes
    """

    if current_time is None:
        return 0

    if scheduled_start is None:
        return 0

    try:

        current_minutes = (
            current_time.hour * 60
            + current_time.minute
        )

        scheduled_minutes = (
            scheduled_start.hour * 60
            + scheduled_start.minute
        )

        lateness = (
            current_minutes
            - scheduled_minutes
        )

        # Never return a negative lateness.
        return max(
            0,
            lateness
        )

    except Exception:

        app.logger.exception(
            "Unable to calculate attendance lateness."
        )

        return 0

# =========================================================
# CALCULATE EARLY CLOCK-OUT MINUTES
# =========================================================

def calculate_early_clockout_minutes(
    current_time,
    scheduled_end
):
    """
    Calculate how many minutes early a worker clocks out
    compared with the configured clock-out end time.

    Example:
        Scheduled end: 05:00 PM
        Actual clock-out: 04:30 PM
        Result: 30 minutes early
    """

    if current_time is None:
        return 0

    if scheduled_end is None:
        return 0

    try:

        current_minutes = (
            current_time.hour * 60
            + current_time.minute
        )

        scheduled_end_minutes = (
            scheduled_end.hour * 60
            + scheduled_end.minute
        )

        early_minutes = (
            scheduled_end_minutes
            - current_minutes
        )

        # If the person clocks out after the scheduled
        # end time, they are not early.
        return max(
            0,
            early_minutes
        )

    except Exception:

        app.logger.exception(
            "Unable to calculate early clock-out minutes."
        )

        return 0
# ============================================================
# INITIALIZE DATABASE
# ============================================================
with app.app_context():
    init_db()
    create_initial_admin()
# ============================================================
# RUN APPLICATION
# ============================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=int(
            os.getenv(
                "PORT",
                5000
            )
        ),
        debug=True
    )
